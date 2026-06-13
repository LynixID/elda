import base64
import hashlib
import re
import json
import math
import calendar
from collections import defaultdict
from datetime import date

from django.http import JsonResponse
from django.shortcuts import render
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from datetime import timedelta
from django.shortcuts import render, redirect
from django.db import connection
from django.db import transaction as db_transaction
from django.views.decorators.http import require_http_methods
import midtransclient
from django.conf import settings

from .models import TbBarang
import csv
import io
try:
    import win32print
except ImportError:
    win32print = None
from collections import defaultdict
from datetime import datetime, date
from zoneinfo import ZoneInfo

from django.contrib import messages
from django.shortcuts import render, redirect
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.db import transaction as db_transaction
from django.db import connection
from django.db.models import Sum, Count
from django.db.models.functions import TruncMonth

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from .models import (
    TbUser,
    TbKategori,
    TbBarang,
    TbSupplier,
    TbTransaksi,
    TbTransaksiDetail,
    TbPembelian,
    TbDetailPembelian,
    TbRetur,
    TbPengembalian,
)

from decimal import Decimal, ROUND_HALF_UP

_MD5_RE = re.compile(r"^[a-fA-F0-9]{32}$")


# =========================================================
# AUTH / SESSION
# =========================================================
def _md5(raw: str) -> str:
    return hashlib.md5(raw.encode("utf-8")).hexdigest()


def _is_logged_in(request) -> bool:
    return bool(request.session.get("kasir_user_id"))


def _require_login(request):
    if not _is_logged_in(request):
        return redirect("kasir:login")
    return None


def _normalize_role(value: str) -> str:
    s = (value or "").strip().lower()
    return "pemilik" if s == "pemilik" else "kasir"


def _normalize_status(value: str) -> str:
    s = (value or "").strip().lower()
    s = s.replace("_", " ").replace("-", " ")
    s = re.sub(r"\s+", " ", s)

    if s == "aktif":
        return "aktif"
    if s in ("tidak aktif", "nonaktif"):
        return "tidak aktif"
    return "aktif"


def _is_pemilik(request) -> bool:
    return _normalize_role(request.session.get("kasir_role")) == "pemilik"


# =========================================================
# AUTO NONAKTIF KASIR JIKA TIDAK LOGIN 3 BULAN
# =========================================================
def _auto_nonaktif_kasir_jika_lama_tidak_login(user):
    if not user:
        return False

    role_val = _normalize_role(getattr(user, "role", "") or "")
    status_val = _normalize_status(getattr(user, "status", "") or "")

    if role_val == "pemilik":
        return False

    if role_val != "kasir":
        return False

    if status_val == "tidak aktif":
        return False

    sekarang = timezone.now()
    batas = sekarang - timezone.timedelta(days=90)

    patokan = getattr(user, "last_login", None) or getattr(user, "created_at", None)
    if not patokan:
        return False

    try:
        if patokan <= batas:
            user.status = "tidak aktif"
            user.updated_at = sekarang
            user.save(update_fields=["status", "updated_at"])
            return True
    except Exception:
        pass

    return False


# =========================================================
# MODEL / DB HELPER
# =========================================================
def _model_field_names(model_cls):
    try:
        return {f.name for f in model_cls._meta.get_fields()}
    except Exception:
        return set()


def _filter_create_kwargs(model_cls, data: dict) -> dict:
    valid = _model_field_names(model_cls)
    return {k: v for k, v in data.items() if k in valid}


def _db_has_column(table_name: str, column_name: str) -> bool:
    try:
        with connection.cursor() as cursor:
            desc = connection.introspection.get_table_description(cursor, table_name)
        cols = {c.name for c in desc}
        return column_name in cols
    except Exception:
        return False


def _force_update_db_column(table_name: str, pk_col: str, pk_val: int, col: str, val):
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                f"UPDATE `{table_name}` SET `{col}`=%s WHERE `{pk_col}`=%s",
                [val, pk_val],
            )
    except Exception:
        pass


def _ambil_kolom_opsional_db_text(table_name: str, pk_col: str, pk_val: int, kandidat_field) -> str:
    try:
        pk_val = int(pk_val or 0)
    except Exception:
        pk_val = 0

    if pk_val <= 0:
        return ""

    for nama_field in kandidat_field:
        try:
            if not _db_has_column(table_name, nama_field):
                continue

            with connection.cursor() as cursor:
                cursor.execute(
                    f"SELECT `{nama_field}` FROM `{table_name}` WHERE `{pk_col}`=%s LIMIT 1",
                    [pk_val],
                )
                row = cursor.fetchone()

            if row and row[0] not in (None, ""):
                return str(row[0]).strip()
        except Exception:
            pass

    return ""


def _ambil_kolom_opsional_db_int(table_name: str, pk_col: str, pk_val: int, kandidat_field, default=0) -> int:
    teks = _ambil_kolom_opsional_db_text(table_name, pk_col, pk_val, kandidat_field)
    if teks == "":
        return int(default or 0)
    return _to_int(teks, default)


def _simpan_ke_kolom_trx_opsional(trx, kandidat_field, nilai):
    field_trx = _model_field_names(TbTransaksi)
    table = TbTransaksi._meta.db_table

    for nama_field in kandidat_field:
        try:
            if nama_field in field_trx:
                TbTransaksi.objects.filter(pk=trx.pk).update(**{nama_field: int(nilai or 0)})
                return True
            if _db_has_column(table, nama_field):
                _force_update_db_column(
                    table_name=table,
                    pk_col="id_transaksi",
                    pk_val=int(trx.id_transaksi),
                    col=nama_field,
                    val=int(nilai or 0),
                )
                return True
        except Exception:
            pass
    return False


def _simpan_ke_kolom_trx_opsional_text(trx, kandidat_field, nilai):
    field_trx = _model_field_names(TbTransaksi)
    table = TbTransaksi._meta.db_table
    nilai = str(nilai or "").strip()

    for nama_field in kandidat_field:
        try:
            if nama_field in field_trx:
                TbTransaksi.objects.filter(pk=trx.pk).update(**{nama_field: nilai})
                return True
            if _db_has_column(table, nama_field):
                _force_update_db_column(
                    table_name=table,
                    pk_col="id_transaksi",
                    pk_val=int(trx.id_transaksi),
                    col=nama_field,
                    val=nilai,
                )
                return True
        except Exception:
            pass
    return False


def _simpan_ke_kolom_detail_opsional_text(detail_obj, kandidat_field, nilai):
    field_detail = _model_field_names(TbTransaksiDetail)
    table = TbTransaksiDetail._meta.db_table
    nilai = str(nilai or "").strip()

    try:
        pk_val = int(getattr(detail_obj, "id_detail", 0) or getattr(detail_obj, "pk", 0) or 0)
    except Exception:
        pk_val = 0

    if pk_val <= 0:
        return False

    for nama_field in kandidat_field:
        try:
            if nama_field in field_detail:
                TbTransaksiDetail.objects.filter(pk=detail_obj.pk).update(**{nama_field: nilai})
                return True
            if _db_has_column(table, nama_field):
                _force_update_db_column(
                    table_name=table,
                    pk_col="id_detail",
                    pk_val=pk_val,
                    col=nama_field,
                    val=nilai,
                )
                return True
        except Exception:
            pass
    return False

def _simpan_ke_kolom_detail_opsional_int(detail_obj, kandidat_field, nilai):
    field_detail = _model_field_names(TbTransaksiDetail)
    table = TbTransaksiDetail._meta.db_table

    try:
        pk_val = int(getattr(detail_obj, "id_detail", 0) or getattr(detail_obj, "pk", 0) or 0)
    except Exception:
        pk_val = 0

    if pk_val <= 0:
        return False

    nilai = int(nilai or 0)

    for nama_field in kandidat_field:
        try:
            if nama_field in field_detail:
                TbTransaksiDetail.objects.filter(pk=detail_obj.pk).update(**{nama_field: nilai})
                return True
            if _db_has_column(table, nama_field):
                _force_update_db_column(
                    table_name=table,
                    pk_col="id_detail",
                    pk_val=pk_val,
                    col=nama_field,
                    val=nilai,
                )
                return True
        except Exception:
            pass
    return False


def _ambil_snapshot_kode_barang_detail(detail_obj) -> str:
    for nama_field in ("kode_barang_snapshot", "kode_barang", "kode"):
        try:
            nilai = getattr(detail_obj, nama_field, None)
            if nilai not in (None, ""):
                nilai = str(nilai).strip()
                if nilai:
                    return nilai
        except Exception:
            pass
    try:
        pk_val = int(getattr(detail_obj, "id_detail", 0) or getattr(detail_obj, "pk", 0) or 0)
    except Exception:
        pk_val = 0

    if pk_val > 0:
        nilai_db = _ambil_kolom_opsional_db_text(
            TbTransaksiDetail._meta.db_table,
            "id_detail",
            pk_val,
            ["kode_barang_snapshot", "kode_barang", "kode"],
        )
        if nilai_db:
            return nilai_db

    return "-"


def _ambil_snapshot_nama_barang_detail(detail_obj, barang_map: dict) -> str:
    try:
        id_barang = int(getattr(detail_obj, "id_barang", 0) or 0)
    except Exception:
        id_barang = 0

    if id_barang > 0 and id_barang in barang_map:
        nama_master = str(barang_map[id_barang] or "").strip()
        if nama_master:
            return nama_master

    for nama_field in ("nama_barang_snapshot", "nama_barang", "barang_nama", "nama"):
        try:
            nilai = getattr(detail_obj, nama_field, None)
            if nilai not in (None, ""):
                nilai = str(nilai).strip()
                if nilai:
                    return nilai
        except Exception:
            pass
    try:
        pk_val = int(getattr(detail_obj, "id_detail", 0) or getattr(detail_obj, "pk", 0) or 0)
    except Exception:
        pk_val = 0

    if pk_val > 0:
        nilai_db = _ambil_kolom_opsional_db_text(
            TbTransaksiDetail._meta.db_table,
            "id_detail",
            pk_val,
            ["nama_barang_snapshot", "nama_barang", "barang_nama", "nama"],
        )
        if nilai_db:
            return nilai_db

    kode_snapshot = _ambil_snapshot_kode_barang_detail(detail_obj)
    if kode_snapshot and kode_snapshot != "-":
        return kode_snapshot

    return f"Barang #{id_barang}" if id_barang else "-"

def _ambil_snapshot_satuan_barang_detail(detail_obj, barang_satuan_map: dict) -> str:
    for nama_field in (
        "satuan_snapshot",
        "satuan_barang_snapshot",
        "besaran_snapshot",
        "satuan",
        "berat",
        "ukuran",
    ):
        try:
            nilai = getattr(detail_obj, nama_field, None)
            if nilai not in (None, ""):
                nilai = str(nilai).strip()
                if nilai:
                    return nilai
        except Exception:
            pass

    try:
        pk_val = int(getattr(detail_obj, "id_detail", 0) or getattr(detail_obj, "pk", 0) or 0)
    except Exception:
        pk_val = 0

    if pk_val > 0:
        nilai_db = _ambil_kolom_opsional_db_text(
            TbTransaksiDetail._meta.db_table,
            "id_detail",
            pk_val,
            [
                "satuan_snapshot",
                "satuan_barang_snapshot",
                "besaran_snapshot",
                "satuan",
                "berat",
                "ukuran",
            ],
        )
        if nilai_db:
            return nilai_db

    try:
        id_barang = int(getattr(detail_obj, "id_barang", 0) or 0)
    except Exception:
        id_barang = 0

    if id_barang > 0 and id_barang in barang_satuan_map:
        satuan_master = str(barang_satuan_map[id_barang] or "").strip()
        if satuan_master:
            return satuan_master

    return "-"
# =========================================================
# TIMEZONE WIB
# =========================================================
def _get_timezone_wib():
    try:
        return ZoneInfo("Asia/Jakarta")
    except Exception:
        try:
            return timezone.get_current_timezone()
        except Exception:
            return None


def _to_wib(dt_obj):
    if not dt_obj:
        return None
    try:
        tz_wib = _get_timezone_wib()
        if timezone.is_naive(dt_obj):
            if tz_wib:
                return timezone.make_aware(dt_obj, tz_wib)
            return dt_obj
        if tz_wib:
            return timezone.localtime(dt_obj, tz_wib)
        return timezone.localtime(dt_obj)
    except Exception:
        return dt_obj


def _filter_date_range_helper(qs, awal_date, akhir_date, field_name="tanggal_waktu"):
    from datetime import datetime, time
    from django.utils import timezone
    tz_wib = _get_timezone_wib()
    dt_awal = datetime.combine(awal_date, time.min)
    dt_akhir = datetime.combine(akhir_date, time.min)
    if tz_wib:
        try:
            dt_awal = timezone.make_aware(dt_awal, tz_wib)
            dt_akhir = timezone.make_aware(dt_akhir, tz_wib)
        except Exception:
            pass
    kwargs = {
        f"{field_name}__gte": dt_awal,
        f"{field_name}__lt": dt_akhir,
    }
    return qs.filter(**kwargs)


# =========================================================
# PARSER / FORMAT ANGKA
# =========================================================
def _to_int(v, default=0):
    try:
        if v is None:
            return default

        s = str(v).strip()
        if s == "":
            return default

        s = s.replace("Rp", "").replace("rp", "").strip()
        s = s.replace(" ", "")
        s = s.replace(".", "")
        if "," in s:
            s = s.split(",")[0]

        return int(float(s))
    except Exception:
        return default


def _safe_decimal_to_int(val):
    try:
        return int(
            Decimal(str(val or 0)).quantize(
                Decimal("1"),
                rounding=ROUND_HALF_UP
            )
        )
    except Exception:
        return 0


def _normalize_nominal_penuh(n: int) -> int:
    try:
        n = int(n or 0)
    except Exception:
        n = 0
    if 0 < n < 1000:
        return n * 1000
    return n


def _format_rupiah_display(n) -> str:
    try:
        val = int(float(n or 0))
    except Exception:
        val = 0

    if 0 < val < 1000:
        val *= 1000

    return f"Rp {val:,}".replace(",", ".")


def _format_angka_ribuan_display(n) -> str:
    try:
        val = int(float(n or 0))
    except Exception:
        val = 0
    if 0 < val < 1000:
        val *= 1000
    return f"{val:,}".replace(",", ".")


# =========================================================
# HELPER JENIS TRANSAKSI / MARKETPLACE
# =========================================================
def _ambil_text_trx_obj_atau_db(trx, kandidat_field) -> str:
    for nama_field in kandidat_field:
        try:
            if hasattr(trx, nama_field):
                nilai = getattr(trx, nama_field, None)
                if nilai not in (None, ""):
                    return str(nilai).strip()
        except Exception:
            pass

    return _ambil_kolom_opsional_db_text(
        TbTransaksi._meta.db_table,
        "id_transaksi",
        getattr(trx, "id_transaksi", 0),
        list(kandidat_field),
    )


def _normalize_marketplace_value(value: str) -> str:
    s = str(value or "").strip().lower()
    s = re.sub(r"\s+", " ", s)

    if s in ("", "-", "null", "none", "0"):
        return ""

    alias = {
        "tik tok": "tiktok",
        "tiktok shop": "tiktok",
        "tik tok shop": "tiktok",
        "shoppee": "shopee",
        "tokped": "tokopedia",
    }
    return alias.get(s, s)


def _ambil_sumber_penjualan_raw(trx) -> str:
    return _ambil_text_trx_obj_atau_db(
        trx,
        (
            "sumber_penjualan",
            "jenis_penjualan",
            "jenis_transaksi",
            "tipe_transaksi",
            "channel_penjualan",
        ),
    )


def _ambil_marketplace_raw(trx) -> str:
    return _ambil_text_trx_obj_atau_db(
        trx,
        (
            "marketplace",
            "nama_marketplace",
            "platform_marketplace",
            "platform",
            "market_place",
            "channel_marketplace",
        ),
    )


def _ambil_sumber_penjualan(trx) -> str:
    sumber = str(_ambil_sumber_penjualan_raw(trx) or "").strip().lower()
    sumber = re.sub(r"\s+", " ", sumber)

    marketplace = _normalize_marketplace_value(_ambil_marketplace_raw(trx))

    if sumber in ("online", "ecommerce", "e-commerce", "marketplace"):
        return "online"
    if sumber in ("offline", "toko", "langsung"):
        return "offline"

    if sumber in ("shopee", "tiktok", "tokopedia", "lazada", "blibli", "bukalapak"):
        return "online"

    if marketplace:
        return "online"

    return "offline"


def _ambil_marketplace(trx) -> str:
    marketplace = _normalize_marketplace_value(_ambil_marketplace_raw(trx))
    if marketplace:
        return marketplace

    sumber = _normalize_marketplace_value(_ambil_sumber_penjualan_raw(trx))
    if sumber in ("shopee", "tiktok", "tokopedia", "lazada", "blibli", "bukalapak"):
        return sumber

    return ""


def _ambil_metode_bayar_raw(trx) -> str:
    return _ambil_text_trx_obj_atau_db(
        trx,
        (
            "metode_bayar",
            "metode_pembayaran",
            "jenis_bayar",
            "payment_method",
            "cara_bayar",
        ),
    )


def _ambil_metode_bayar(trx) -> str:
    metode = str(_ambil_metode_bayar_raw(trx) or "").strip().lower()
    metode = metode.replace("-", " ").replace("_", " ")
    metode = re.sub(r"\s+", " ", metode)

    if metode in ("qris", "qr", "qr code", "midtrans qris", "shopeepay qris"):
        return "qris"

    if metode in ("tunai", "cash", "uang tunai"):
        return "tunai"

    if metode in ("marketplace", "online", "shopee", "tiktok", "tokopedia", "lazada"):
        return "marketplace"

    if _ambil_sumber_penjualan(trx) == "online":
        return "marketplace"

    return "tunai"


def _label_jenis_transaksi(trx) -> str:
    sumber = _ambil_sumber_penjualan(trx)
    metode = _ambil_metode_bayar(trx)
    marketplace = _ambil_marketplace(trx)

    if sumber == "online":
        if marketplace == "shopee":
            return "Online Shopee"
        if marketplace == "tiktok":
            return "Online TikTok Shop"
        if marketplace == "tokopedia":
            return "Online Tokopedia"
        if marketplace == "lazada":
            return "Online Lazada"
        return "Online Marketplace"

    if metode == "qris":
        return "Offline QRIS"

    return "Offline Tunai"


def _label_marketplace(trx) -> str:
    nilai = _ambil_marketplace(trx)

    if not nilai:
        return "-"
    if nilai == "shopee":
        return "Shopee"
    if nilai == "tiktok":
        return "TikTok"
    if nilai == "tokopedia":
        return "Tokopedia"
    if nilai == "lazada":
        return "Lazada"
    if nilai == "blibli":
        return "Blibli"
    if nilai == "bukalapak":
        return "Bukalapak"

    return nilai.title()

def _ambil_nominal_bayar_tampil(trx, total_belanja_final=0):
    trx_id = int(getattr(trx, "id_transaksi", 0) or 0)

    nominal_bayar = _normalize_nominal_penuh(
        _ambil_kolom_opsional_db_int(
            TbTransaksi._meta.db_table,
            "id_transaksi",
            trx_id,
            ["nominal_bayar", "uang_bayar", "bayar", "jumlah_bayar"],
            getattr(trx, "nominal_bayar", 0),
        )
    )

    kembalian = _normalize_nominal_penuh(
        _ambil_kolom_opsional_db_int(
            TbTransaksi._meta.db_table,
            "id_transaksi",
            trx_id,
            ["kembalian", "uang_kembali", "total_kembalian", "kembali"],
            getattr(trx, "kembalian", 0),
        )
    )

    sumber = _ambil_sumber_penjualan(trx)
    metode = _ambil_metode_bayar(trx)

    if sumber == "online":
        nominal_bayar = 0
        kembalian = 0

    elif sumber == "offline" and metode == "qris":
        nominal_bayar = int(total_belanja_final or 0)
        kembalian = 0

    return nominal_bayar, kembalian


# =========================================================
# TRANSAKSI: TANGGAL / NO JUAL
# =========================================================
def _build_tanggal_waktu_from_payload(payload):
    tz_wib = _get_timezone_wib()
    now_utc = timezone.now()

    try:
        now_wib = timezone.localtime(now_utc, tz_wib) if tz_wib else timezone.localtime(now_utc)
    except Exception:
        now_wib = now_utc

    tanggal_raw = str(payload.get("tanggal", "") or "").strip()
    if not tanggal_raw:
        return now_wib

    try:
        tanggal_obj = datetime.strptime(tanggal_raw, "%Y-%m-%d").date()
        dt_naive = datetime.combine(tanggal_obj, now_wib.time())

        if timezone.is_naive(dt_naive):
            try:
                if tz_wib:
                    return timezone.make_aware(dt_naive, tz_wib)
                return timezone.make_aware(dt_naive, timezone.get_current_timezone())
            except Exception:
                return now_wib

        return dt_naive
    except Exception:
        return now_wib


def _generate_no_jual_otomatis():
    tz_wib = _get_timezone_wib()
    now = timezone.now()
    try:
        now = timezone.localtime(now, tz_wib) if tz_wib else timezone.localtime(now)
    except Exception:
        pass

    return f"NOTA-{now.strftime('%d%m%y%H%M%S')}"


def _build_transaksi_create_kwargs(request, payload, total_item, total_harga, tanggal_waktu_transaksi):
    user_id = int(request.session.get("kasir_user_id") or 0)

    nama_kasir = (
        (request.session.get("kasir_nama") or "").strip()
        or (request.session.get("kasir_username") or "").strip()
    )

    no_jual = str(
        payload.get("no_jual")
        or payload.get("noJual")
        or payload.get("nomor_jual")
        or payload.get("nomorJual")
        or ""
    ).strip()

    if not no_jual:
        no_jual = _generate_no_jual_otomatis()

    sumber_penjualan = str(
        payload.get("sumber_penjualan")
        or payload.get("sumberPenjualan")
        or "offline"
    ).strip().lower()

    if sumber_penjualan not in ("offline", "online"):
        sumber_penjualan = "offline"

    metode_bayar = str(
        payload.get("metode_bayar")
        or payload.get("metodeBayar")
        or payload.get("metode")
        or "tunai"
    ).strip().lower()

    marketplace = str(payload.get("marketplace") or "").strip().lower()

    if sumber_penjualan == "online":
        metode_bayar = "marketplace"

        if marketplace not in ("shopee", "tiktok", "tokopedia", "lazada"):
            marketplace = ""

        nominal_bayar = 0
        kembalian = 0

    else:
        sumber_penjualan = "offline"
        marketplace = ""

        if metode_bayar not in ("tunai", "qris"):
            metode_bayar = "tunai"

        if metode_bayar == "qris":
            nominal_bayar = int(total_harga or 0)
            kembalian = 0
        else:
            nominal_bayar = _normalize_nominal_penuh(
                _to_int(payload.get("bayar", payload.get("nominal_bayar", 0)), 0)
            )

            kembalian = _normalize_nominal_penuh(
                _to_int(payload.get("kembali", payload.get("kembalian", 0)), 0)
            )

            if nominal_bayar > 0 and nominal_bayar >= total_harga:
                kembalian = nominal_bayar - total_harga

    base_data = {
        "no_jual": no_jual,
        "tanggal_waktu": tanggal_waktu_transaksi,
        "sumber_penjualan": sumber_penjualan,
        "metode_bayar": metode_bayar,
        "marketplace": marketplace,
        "id_kasir": user_id,
        "nama_kasir": nama_kasir,
        "total_item": int(total_item or 0),
        "total_harga": int(total_harga or 0),
        "nominal_bayar": int(nominal_bayar or 0),
        "kembalian": int(kembalian or 0),
    }

    field_names = _model_field_names(TbTransaksi)
    now = timezone.now()

    if "created_at" in field_names:
        base_data["created_at"] = now

    if "updated_at" in field_names:
        base_data["updated_at"] = now

    return _filter_create_kwargs(TbTransaksi, base_data)


# =========================================================
# HELPER DASHBOARD HOME
# =========================================================
def _ambil_kode_transaksi(trx) -> str:
    for nama_field in ("no_jual", "kode_transaksi", "no_transaksi", "nomor_jual"):
        try:
            nilai = getattr(trx, nama_field, None)
            if nilai:
                return str(nilai).strip()
        except Exception:
            pass

    nilai_db = _ambil_kolom_opsional_db_text(
        TbTransaksi._meta.db_table,
        "id_transaksi",
        getattr(trx, "id_transaksi", 0),
        ["no_jual", "kode_transaksi", "no_transaksi", "nomor_jual"],
    )
    if nilai_db:
        return nilai_db

    try:
        return f"TRX-{int(getattr(trx, 'id_transaksi', 0)):05d}"
    except Exception:
        return "-"


def _ambil_nama_kasir(trx, peta_user: dict) -> str:
    try:
        id_kasir = int(getattr(trx, "id_kasir", 0) or 0)
    except Exception:
        id_kasir = 0

    # PRIORITAS 1:
    # kalau akun kasirnya masih ada, selalu pakai nama terbaru dari tb_user
    if id_kasir and id_kasir in peta_user:
        nama_user = str(peta_user[id_kasir] or "").strip()
        if nama_user:
            return nama_user

    # PRIORITAS 2:
    # fallback ke snapshot nama di transaksi / kolom lain
    for nama_field in ("nama_kasir", "kasir_nama", "nama_user", "nama"):
        try:
            nilai = getattr(trx, nama_field, None)
            if nilai:
                return str(nilai).strip()
        except Exception:
            pass

    for nama_field in ("username_kasir", "kasir_username", "username"):
        try:
            nilai = getattr(trx, nama_field, None)
            if nilai:
                return str(nilai).strip()
        except Exception:
            pass

    nilai_db = _ambil_kolom_opsional_db_text(
        TbTransaksi._meta.db_table,
        "id_transaksi",
        getattr(trx, "id_transaksi", 0),
        [
            "nama_kasir",
            "kasir_nama",
            "nama_user",
            "nama",
            "username_kasir",
            "kasir_username",
            "username",
        ],
    )
    if nilai_db:
        return nilai_db

    return "-"


def _sinkronkan_nama_kasir_di_riwayat(id_kasir: int, nama_baru: str, username_baru: str = ""):
    try:
        id_kasir = int(id_kasir or 0)
    except Exception:
        id_kasir = 0

    nama_final = str(nama_baru or "").strip()
    username_final = str(username_baru or "").strip()

    if id_kasir <= 0:
        return

    if not nama_final and not username_final:
        return

    if not nama_final:
        nama_final = username_final
    if not username_final:
        username_final = nama_final

    # update kolom model utama
    try:
        TbTransaksi.objects.filter(id_kasir=id_kasir).update(nama_kasir=nama_final)
    except Exception:
        pass

    table = TbTransaksi._meta.db_table

    nama_cols = ["nama_kasir", "kasir_nama", "nama_user", "nama"]
    username_cols = ["username_kasir", "kasir_username", "username"]

    for col in nama_cols:
        try:
            if _db_has_column(table, col):
                with connection.cursor() as cursor:
                    cursor.execute(
                        f"UPDATE `{table}` SET `{col}`=%s WHERE `id_kasir`=%s",
                        [nama_final, id_kasir],
                    )
        except Exception:
            pass

    for col in username_cols:
        try:
            if _db_has_column(table, col):
                with connection.cursor() as cursor:
                    cursor.execute(
                        f"UPDATE `{table}` SET `{col}`=%s WHERE `id_kasir`=%s",
                        [username_final, id_kasir],
                    )
        except Exception:
            pass


# =========================================================
# HELPER UMUM
# =========================================================
def _angka_aman(v, default=0):
    try:
        if v is None:
            return default
        if isinstance(v, bool):
            return int(v)
        s = str(v).strip()
        if s == "":
            return default
        s = s.replace("Rp", "").replace("rp", "").replace(".", "").replace(",", "")
        return int(float(s))
    except Exception:
        return default


def _teks_aman(v, default=""):
    try:
        return str(v or "").strip()
    except Exception:
        return default

def _nama_supplier_display(supplier):
    if not supplier:
        return ""

    nama_perusahaan = _teks_aman(getattr(supplier, "nama_perusahaan", ""))
    if not nama_perusahaan:
        nama_perusahaan = _ambil_kolom_opsional_db_text(
            TbSupplier._meta.db_table,
            "id_supplier",
            getattr(supplier, "id_supplier", 0),
            ["nama_perusahaan", "perusahaan", "nama_mitra", "mitra_supplier"],
        )

    nama_supplier = _teks_aman(getattr(supplier, "nama_supplier", ""))

    if nama_perusahaan and nama_supplier:
        return f"{nama_perusahaan} - {nama_supplier}"

    if nama_perusahaan:
        return nama_perusahaan

    if nama_supplier:
        return nama_supplier

    return ""


def _ambil_id_supplier_dari_barang(barang):
    """
    Ambil id_supplier dari TbBarang.
    Aman untuk 2 kondisi:
    1. models.py sudah punya field id_supplier
    2. database sudah punya kolom id_supplier, tapi models.py belum ditambah field
    """

    if not barang:
        return 0
    
    try:
        nilai = getattr(barang, "id_supplier", None)

        if nilai and hasattr(nilai, "id_supplier"):
            return int(getattr(nilai, "id_supplier", 0) or 0)

        if nilai not in (None, ""):
            return _angka_aman(nilai, 0)
    except Exception:
        pass

    try:
        id_barang = int(getattr(barang, "id_barang", 0) or 0)
    except Exception:
        id_barang = 0

    if id_barang <= 0:
        return 0

    return _ambil_kolom_opsional_db_int(
        TbBarang._meta.db_table,
        "id_barang",
        id_barang,
        ["id_supplier"],
        0,
    )


def _ambil_supplier_barang_info(barang):
    id_supplier = _ambil_id_supplier_dari_barang(barang)

    if id_supplier <= 0:
        return {
            "id_supplier": 0,
            "nama_supplier": "",
            "nama_perusahaan": "",
            "supplier_label": "",
        }

    supplier = TbSupplier.objects.filter(id_supplier=id_supplier).first()

    if not supplier:
        return {
            "id_supplier": id_supplier,
            "nama_supplier": "",
            "nama_perusahaan": "",
            "supplier_label": f"Supplier #{id_supplier}",
        }

    nama_supplier = _teks_aman(getattr(supplier, "nama_supplier", ""))
    nama_perusahaan = _teks_aman(getattr(supplier, "nama_perusahaan", ""))

    if not nama_perusahaan:
        nama_perusahaan = _ambil_kolom_opsional_db_text(
            TbSupplier._meta.db_table,
            "id_supplier",
            id_supplier,
            ["nama_perusahaan", "perusahaan", "nama_mitra", "mitra_supplier"],
        )

    supplier_label = _nama_supplier_display(supplier)

    return {
        "id_supplier": id_supplier,
        "nama_supplier": nama_supplier,
        "nama_perusahaan": nama_perusahaan,
        "supplier_label": supplier_label,
    }


def _format_rupiah_angka(n):
    n = max(0, _angka_aman(n))
    return "Rp {:,}".format(n).replace(",", ".")


def _format_angka_desimal(x, digit=4):
    try:
        x = float(x)
        if x.is_integer():
            return str(int(x))
        return f"{x:.{digit}f}".rstrip("0").rstrip(".")
    except Exception:
        return str(x)


def _bulan_indonesia(n):
    data = {
        1: "Januari",
        2: "Februari",
        3: "Maret",
        4: "April",
        5: "Mei",
        6: "Juni",
        7: "Juli",
        8: "Agustus",
        9: "September",
        10: "Oktober",
        11: "November",
        12: "Desember",
    }
    return data.get(int(n or 0), "-")


def _tanggal_periode_bulanan(bulan, tahun):
    bulan = int(bulan)
    tahun = int(tahun)
    hari_terakhir = calendar.monthrange(tahun, bulan)[1]
    return date(tahun, bulan, 1), date(tahun, bulan, hari_terakhir)


def _triangular(x, a, b, c):
    x = float(x)
    a = float(a)
    b = float(b)
    c = float(c)

    if a == b and x <= b:
        return 1.0
    if b == c and x >= b:
        return 1.0

    if x <= a or x >= c:
        return 0.0
    if x == b:
        return 1.0
    if a < x < b:
        return 0.0 if b == a else (x - a) / (b - a)
    if b < x < c:
        return 0.0 if c == b else (c - x) / (c - b)
    return 0.0


def _pembulatan_standar(nilai):
    try:
        return int(
            Decimal(str(nilai or 0)).quantize(
                Decimal("1"),
                rounding=ROUND_HALF_UP
            )
        )
    except Exception:
        return 0


def _rapikan_domain_fuzzy(domain):
    def rapikan_triplet(a, b, c):
        a = float(a)
        b = float(b)
        c = float(c)

        if b < a:
            b = a
        if c < b:
            c = b

        return (a, b, c)

    return {
        "rendah": rapikan_triplet(*domain["rendah"]),
        "sedang": rapikan_triplet(*domain["sedang"]),
        "tinggi": rapikan_triplet(*domain["tinggi"]),
    }


def _buat_domain_fuzzy_stabil(daftar_nilai, fallback_maks=10, tipe="penjualan"):
    nilai = [max(0.0, float(x or 0)) for x in daftar_nilai]
    maks_data = max(nilai) if nilai else 0.0
    maks = max(float(fallback_maks), maks_data, 1.0)

    if tipe == "stok":
        batas1 = max(2.0, round(maks * 0.25))
        batas2 = max(batas1 + 1.0, round(maks * 0.50))
        batas3 = max(batas2 + 1.0, round(maks * 0.75))
    else:
        batas1 = max(2.0, round(maks * 0.30))
        batas2 = max(batas1 + 1.0, round(maks * 0.60))
        batas3 = max(batas2 + 1.0, round(maks * 0.85))

    return {
        "rendah": (0.0, 0.0, batas1),
        "sedang": (0.0, batas2, batas3),
        "tinggi": (batas2, maks, maks),
    }


def _ambil_domain_global_rekomendasi(periode_mulai, periode_selesai):
    semua_barang = _ambil_data_barang_rekomendasi(
        periode_mulai,
        periode_selesai,
        "SEMUA"
    )

    semua_penjualan = []
    semua_stok = []

    for barang in semua_barang:
        semua_penjualan.append(
            _angka_aman(barang.get("terjual_periode"), 0)
        )
        semua_stok.append(
            _angka_aman(barang.get("stok_saat_ini"), 0)
        )

    domain_penjualan = _rapikan_domain_fuzzy(
        _buat_domain_fuzzy_stabil(
            semua_penjualan,
            fallback_maks=10,
            tipe="penjualan"
        )
    )

    domain_stok = _rapikan_domain_fuzzy(
        _buat_domain_fuzzy_stabil(
            semua_stok,
            fallback_maks=10,
            tipe="stok"
        )
    )

    return {
        "penjualan": domain_penjualan,
        "stok": domain_stok,
    }


def _membership_input(nilai, domain):
    return {
        "rendah": round(_triangular(nilai, *domain["rendah"]), 6),
        "sedang": round(_triangular(nilai, *domain["sedang"]), 6),
        "tinggi": round(_triangular(nilai, *domain["tinggi"]), 6),
    }


def _membership_variabel(nilai, domain):
    x = max(0.0, float(nilai or 0))
    return {
        "rendah": round(_triangular(x, *domain["rendah"]), 6),
        "sedang": round(_triangular(x, *domain["sedang"]), 6),
        "tinggi": round(_triangular(x, *domain["tinggi"]), 6),
    }


# =========================================================
# FUZZY
# =========================================================
def _hitung_kebutuhan_hitung_manual(terjual_periode, stok_saat_ini, jumlah_hari, stok_minimal=0):
    terjual_periode = max(0, int(terjual_periode or 0))
    stok_saat_ini = max(0, int(stok_saat_ini or 0))
    stok_minimal = max(0, int(stok_minimal or 0))
    jumlah_hari = max(1, int(jumlah_hari or 1))

    rata_harian = terjual_periode / jumlah_hari
    proyeksi_30_hari = _pembulatan_standar(rata_harian * 30)
    kebutuhan_stok_minimal = max(0, stok_minimal - stok_saat_ini)
    target_stok_aman = proyeksi_30_hari + stok_minimal
    kebutuhan_total = max(0, target_stok_aman - stok_saat_ini)
    kebutuhan_total = max(kebutuhan_total, kebutuhan_stok_minimal)

    if stok_saat_ini <= 0 and kebutuhan_total <= 0 and stok_minimal > 0:
        kebutuhan_total = stok_minimal

    target_fuzzy = max(
        1,
        kebutuhan_total,
        target_stok_aman,
        proyeksi_30_hari,
        stok_minimal
    )

    return {
        "rata_harian": round(rata_harian, 4),
        "proyeksi_30_hari": proyeksi_30_hari,
        "stok_pengaman": stok_minimal,
        "stok_minimal": stok_minimal,
        "kebutuhan_stok_minimal": kebutuhan_stok_minimal,
        "target_stok_aman": target_stok_aman,
        "target_fuzzy": target_fuzzy,
        "kebutuhan_manual": kebutuhan_total,
        "kebutuhan_selisih": kebutuhan_total,
        "kebutuhan_total": kebutuhan_total,
    }


def _aturan_fuzzy(alpha_penjualan, alpha_stok, target_maks):
    aturan = []

    def tambah(
        nama_rule,
        label_penjualan,
        label_stok,
        label_mamdani,
        label_tsukamoto,
        konstanta_sugeno
    ):
        mu1 = float(alpha_penjualan.get(label_penjualan, 0))
        mu2 = float(alpha_stok.get(label_stok, 0))
        alpha = round(min(mu1, mu2), 6)

        if alpha > 0:
            aturan.append({
                "nama": nama_rule,
                "label_penjualan": label_penjualan,
                "label_stok": label_stok,
                "alpha": alpha,

                "label_output_mamdani": label_mamdani,
                "label_output_tsukamoto": label_tsukamoto,
                "konstanta_output_sugeno": float(konstanta_sugeno),

                "rumus_alpha": (
                    f"α-predikat = min(μ_penjualan_{label_penjualan}, μ_stok_{label_stok}) = "
                    f"min({_format_angka_desimal(mu1)}, {_format_angka_desimal(mu2)}) = {_format_angka_desimal(alpha)}"
                )
            })

    target_maks = max(1, int(target_maks or 1))

    z_kecil = max(0, _pembulatan_standar(0.25 * target_maks))
    z_sedang = max(1, _pembulatan_standar(0.60 * target_maks))
    z_besar = max(1, _pembulatan_standar(1.00 * target_maks))

    tambah("Jika penjualan tinggi dan stok rendah maka pemesanan besar",  "tinggi", "rendah", "besar",  "banyak",  z_besar)
    tambah("Jika penjualan tinggi dan stok sedang maka pemesanan besar",  "tinggi", "sedang", "besar",  "banyak",  z_besar)
    tambah("Jika penjualan tinggi dan stok tinggi maka pemesanan sedang", "tinggi", "tinggi", "sedang", "banyak",  z_sedang)

    tambah("Jika penjualan sedang dan stok rendah maka pemesanan besar",  "sedang", "rendah", "besar",  "banyak",  z_besar)
    tambah("Jika penjualan sedang dan stok sedang maka pemesanan sedang", "sedang", "sedang", "sedang", "banyak",  z_sedang)
    tambah("Jika penjualan sedang dan stok tinggi maka pemesanan kecil",  "sedang", "tinggi", "kecil",  "sedikit", z_kecil)

    tambah("Jika penjualan rendah dan stok rendah maka pemesanan sedang", "rendah", "rendah", "sedang", "banyak",  z_sedang)
    tambah("Jika penjualan rendah dan stok sedang maka pemesanan kecil",  "rendah", "sedang", "kecil",  "sedikit", z_kecil)
    tambah("Jika penjualan rendah dan stok tinggi maka pemesanan kecil",  "rendah", "tinggi", "kecil",  "sedikit", z_kecil)

    return aturan


def _fungsi_output_mamdani(label, y, maks_y):
    domain = {
        "kecil": (0, 0.25 * maks_y, 0.45 * maks_y),
        "sedang": (0.30 * maks_y, 0.55 * maks_y, 0.80 * maks_y),
        "besar": (0.65 * maks_y, 0.85 * maks_y, maks_y),
    }
    return _triangular(y, *domain[label])


def _hitung_mamdani_detail(aturan, maks_target):
    if not aturan:
        return {
            "nilai": 0,
            "pembilang": 0,
            "penyebut": 0,
            "rumus": "Tidak ada aturan aktif, maka hasil Mamdani = 0.",
        }

    maks_target = max(float(maks_target), 1.0)
    titik_uji = list(range(0, int(math.ceil(maks_target)) + 1))

    pembilang = 0.0
    penyebut = 0.0

    for y in titik_uji:
        mu_agregat = 0.0
        for rule in aturan:
            mu_output = _fungsi_output_mamdani(rule["label_output_mamdani"], y, maks_target)
            mu_terpotong = min(rule["alpha"], mu_output)
            mu_agregat = max(mu_agregat, mu_terpotong)

        pembilang += y * mu_agregat
        penyebut += mu_agregat

    if penyebut <= 0:
        return {
            "nilai": 0,
            "pembilang": pembilang,
            "penyebut": penyebut,
            "rumus": "Σμ(z) = 0, maka hasil Mamdani = 0.",
        }

    nilai_asli = pembilang / penyebut
    nilai = max(0, _pembulatan_standar(nilai_asli))

    rumus = (
        f"Z_mamdani = Σ(z × μ(z)) / Σμ(z) = "
        f"{_format_angka_desimal(pembilang)} / {_format_angka_desimal(penyebut)} = "
        f"{_format_angka_desimal(nilai_asli)} ≈ {nilai}"
    )

    return {
        "nilai": nilai,
        "pembilang": pembilang,
        "penyebut": penyebut,
        "rumus": rumus,
    }


def _inverse_tsukamoto_detail(label, alpha, maks_target):
    alpha = max(0.0, min(1.0, float(alpha)))
    maks_target = max(float(maks_target), 1.0)

    z_min = 0.0
    z_max = maks_target

    if label == "sedikit":
        zi = z_max - (alpha * (z_max - z_min))
        rumus = (
            f"Karena output monoton turun (sedikit), "
            f"z = z_max - α(z_max - z_min) = "
            f"{_format_angka_desimal(z_max)} - "
            f"{_format_angka_desimal(alpha)} × "
            f"({_format_angka_desimal(z_max)} - {_format_angka_desimal(z_min)}) = "
            f"{_format_angka_desimal(zi)}"
        )
        return zi, rumus

    zi = z_min + (alpha * (z_max - z_min))
    rumus = (
        f"Karena output monoton naik (banyak), "
        f"z = z_min + α(z_max - z_min) = "
        f"{_format_angka_desimal(z_min)} + "
        f"{_format_angka_desimal(alpha)} × "
        f"({_format_angka_desimal(z_max)} - {_format_angka_desimal(z_min)}) = "
        f"{_format_angka_desimal(zi)}"
    )
    return zi, rumus


def _hitung_tsukamoto_detail(aturan, maks_target):
    if not aturan:
        return {
            "nilai": 0,
            "pembilang": 0,
            "penyebut": 0,
            "rumus": "Tidak ada aturan aktif, maka hasil Tsukamoto = 0.",
            "detail_baris": [],
        }

    pembilang = 0.0
    penyebut = 0.0
    detail_baris = []

    for rule in aturan:
        zi, rumus_zi = _inverse_tsukamoto_detail(
            rule["label_output_tsukamoto"],
            rule["alpha"],
            maks_target
        )
        pembilang += rule["alpha"] * zi
        penyebut += rule["alpha"]

        detail_baris.append(
            f"{rule['nama']} → {rumus_zi}; "
            f"α × z = {_format_angka_desimal(rule['alpha'])} × {_format_angka_desimal(zi)} = "
            f"{_format_angka_desimal(rule['alpha'] * zi)}"
        )

    if penyebut <= 0:
        return {
            "nilai": 0,
            "pembilang": pembilang,
            "penyebut": penyebut,
            "rumus": "Σα = 0, maka hasil Tsukamoto = 0.",
            "detail_baris": detail_baris,
        }

    nilai_asli = pembilang / penyebut
    nilai = max(0, _pembulatan_standar(nilai_asli))

    rumus = (
        f"Z_tsukamoto = Σ(αn × zn) / Σαn = "
        f"{_format_angka_desimal(pembilang)} / {_format_angka_desimal(penyebut)} = "
        f"{_format_angka_desimal(nilai_asli)} ≈ {nilai}"
    )

    return {
        "nilai": nilai,
        "pembilang": pembilang,
        "penyebut": penyebut,
        "rumus": rumus,
        "detail_baris": detail_baris,
    }


def _hitung_sugeno_detail(aturan):
    if not aturan:
        return {
            "nilai": 0,
            "pembilang": 0,
            "penyebut": 0,
            "rumus": "Tidak ada aturan aktif, maka hasil Sugeno = 0.",
            "detail_baris": [],
        }

    pembilang = sum(r["alpha"] * r["konstanta_output_sugeno"] for r in aturan)
    penyebut = sum(r["alpha"] for r in aturan)

    if penyebut <= 0:
        return {
            "nilai": 0,
            "pembilang": pembilang,
            "penyebut": penyebut,
            "rumus": "Σα = 0, maka hasil Sugeno = 0.",
            "detail_baris": [],
        }

    nilai_asli = pembilang / penyebut
    nilai = max(0, _pembulatan_standar(nilai_asli))

    detail_baris = [
        f"{r['nama']} → α × z = {_format_angka_desimal(r['alpha'])} × {_format_angka_desimal(r['konstanta_output_sugeno'])} = {_format_angka_desimal(r['alpha'] * r['konstanta_output_sugeno'])}"
        for r in aturan
    ]

    rumus = (
        f"Z_sugeno = Σ(αi × zi) / Σαi = "
        f"{_format_angka_desimal(pembilang)} / {_format_angka_desimal(penyebut)} = "
        f"{_format_angka_desimal(nilai_asli)} ≈ {nilai}"
    )

    return {
        "nilai": nilai,
        "pembilang": pembilang,
        "penyebut": penyebut,
        "rumus": rumus,
        "detail_baris": detail_baris,
    }


def _evaluasi_metode_item(item):
    kebutuhan_total = max(0, _angka_aman(item.get("kebutuhan_total"), 0))

    hasil_metode = {
        "mamdani": max(0, _angka_aman(item.get("hasil_mamdani"), 0)),
        "tsukamoto": max(0, _angka_aman(item.get("hasil_tsukamoto"), 0)),
        "sugeno": max(0, _angka_aman(item.get("hasil_sugeno"), 0)),
    }

    error_mamdani = abs(hasil_metode["mamdani"] - kebutuhan_total)
    error_tsukamoto = abs(hasil_metode["tsukamoto"] - kebutuhan_total)
    error_sugeno = abs(hasil_metode["sugeno"] - kebutuhan_total)

    errors_item = {
        "mamdani": error_mamdani,
        "tsukamoto": error_tsukamoto,
        "sugeno": error_sugeno,
    }

    error_min = min(errors_item.values())

    kandidat_terbaik = [
        nama for nama, nilai in errors_item.items()
        if nilai == error_min
    ]

    if len(kandidat_terbaik) == 1:
        metode_terbaik_item = kandidat_terbaik[0]
        rekomendasi_fuzzy = hasil_metode[metode_terbaik_item]
        status_metode = "unik"
    else:
        metode_terbaik_item = "setara"
        nilai_setara = [hasil_metode[nama] for nama in kandidat_terbaik]
        rekomendasi_fuzzy = _pembulatan_standar(
            sum(nilai_setara) / max(1, len(nilai_setara))
        )
        status_metode = "setara"

    return {
        "error_mamdani": error_mamdani,
        "error_tsukamoto": error_tsukamoto,
        "error_sugeno": error_sugeno,
        "error_min": error_min,
        "kandidat_terbaik": kandidat_terbaik,
        "metode_terbaik_item": metode_terbaik_item,
        "status_metode": status_metode,
        "rekomendasi_fuzzy": max(0, rekomendasi_fuzzy),
        "rekomendasi_item": max(0, rekomendasi_fuzzy),
    }

def _evaluasi_fuzzy_item(item, aturan):
    kebutuhan_total = max(0, _angka_aman(item.get("kebutuhan_total"), 0))
    kebutuhan_stok_minimal = max(0, _angka_aman(item.get("kebutuhan_stok_minimal"), 0))

    target_fuzzy = max(
        1,
        _angka_aman(item.get("target_fuzzy"), 0),
        _angka_aman(item.get("target_stok_aman"), 0),
        kebutuhan_total,
        kebutuhan_stok_minimal,
    )

    detail_mamdani = _hitung_mamdani_detail(aturan, target_fuzzy)
    detail_tsukamoto = _hitung_tsukamoto_detail(aturan, target_fuzzy)
    detail_sugeno = _hitung_sugeno_detail(aturan)

    hasil_mamdani = max(0, _angka_aman(detail_mamdani.get("nilai"), 0))
    hasil_tsukamoto = max(0, _angka_aman(detail_tsukamoto.get("nilai"), 0))
    hasil_sugeno = max(0, _angka_aman(detail_sugeno.get("nilai"), 0))

    evaluasi = _evaluasi_metode_item({
        "hasil_mamdani": hasil_mamdani,
        "hasil_tsukamoto": hasil_tsukamoto,
        "hasil_sugeno": hasil_sugeno,
        "kebutuhan_total": kebutuhan_total,
    })

    rekomendasi_fuzzy = max(
        0,
        _angka_aman(evaluasi.get("rekomendasi_fuzzy"), 0)
    )

    if rekomendasi_fuzzy > 0:
        rekomendasi_beli = rekomendasi_fuzzy
        dasar_rekomendasi = "hasil_metode_fuzzy_terbaik"
    else:
        rekomendasi_beli = kebutuhan_total
        dasar_rekomendasi = "fallback_kebutuhan_total_karena_fuzzy_tidak_aktif"

    rumus_rekomendasi_fuzzy = (
        f"Mamdani={hasil_mamdani}, Tsukamoto={hasil_tsukamoto}, Sugeno={hasil_sugeno}. "
        f"Kebutuhan pembanding={kebutuhan_total}. "
        f"Error Mamdani={evaluasi.get('error_mamdani', 0)}, "
        f"Tsukamoto={evaluasi.get('error_tsukamoto', 0)}, "
        f"Sugeno={evaluasi.get('error_sugeno', 0)}. "
        f"Metode terpilih={evaluasi.get('metode_terbaik_item')}. "
        f"Rekomendasi fuzzy={rekomendasi_fuzzy}. "
        f"Rekomendasi ideal akhir={rekomendasi_beli}."
    )

    return {
        "hasil_mamdani": hasil_mamdani,
        "hasil_tsukamoto": hasil_tsukamoto,
        "hasil_sugeno": hasil_sugeno,

        "rumus_mamdani": detail_mamdani.get("rumus", ""),
        "rumus_tsukamoto": detail_tsukamoto.get("rumus", ""),
        "rumus_sugeno": detail_sugeno.get("rumus", ""),

        "error_mamdani": evaluasi.get("error_mamdani", 0),
        "error_tsukamoto": evaluasi.get("error_tsukamoto", 0),
        "error_sugeno": evaluasi.get("error_sugeno", 0),
        "error_min": evaluasi.get("error_min", 0),

        "kandidat_terbaik": evaluasi.get("kandidat_terbaik", []),
        "metode_terbaik_item": evaluasi.get("metode_terbaik_item", "setara"),
        "status_metode": evaluasi.get("status_metode", "setara"),

        "rekomendasi_fuzzy": rekomendasi_fuzzy,
        "nilai_rekomendasi_fuzzy": rekomendasi_fuzzy,
        "rekomendasi_item": rekomendasi_fuzzy,

        "rekomendasi_beli": rekomendasi_beli,
        "rekomendasi_ideal": rekomendasi_beli,
        "target_restock": rekomendasi_beli,

        "kebutuhan_total": kebutuhan_total,
        "kebutuhan_stok_minimal": kebutuhan_stok_minimal,
        "target_fuzzy": target_fuzzy,

        "dasar_rekomendasi": dasar_rekomendasi,
        "rumus_rekomendasi_fuzzy": rumus_rekomendasi_fuzzy,
    }

def _evaluasi_metode_global(total_error_mamdani, total_error_tsukamoto, total_error_sugeno):
    errors_global = {
        "mamdani": total_error_mamdani,
        "tsukamoto": total_error_tsukamoto,
        "sugeno": total_error_sugeno,
    }

    error_min = min(errors_global.values())
    kandidat_global = [nama for nama, nilai in errors_global.items() if nilai == error_min]

    if len(kandidat_global) == 1:
        return {
            "metode_terbaik_global": kandidat_global[0],
            "status_global": "unik",
            "kandidat_global": kandidat_global,
        }

    return {
        "metode_terbaik_global": "setara",
        "status_global": "setara",
        "kandidat_global": kandidat_global,
    }


def _hitung_skor_urgent(item):
    stok = max(0, _angka_aman(item.get("stok_saat_ini"), 0))
    stok_minimal = max(0, _angka_aman(item.get("stok_minimal"), 0))
    terjual = max(0, _angka_aman(item.get("terjual_periode"), 0))
    kebutuhan = max(0, _angka_aman(item.get("kebutuhan_total"), 0))
    qty_final = max(0, _angka_aman(item.get("qty_final_beli"), 0))

    skor = 0

    if stok <= 0:
        skor += 45
    elif stok_minimal > 0 and stok < stok_minimal:
        skor += 35
    elif stok_minimal > 0 and stok == stok_minimal:
        skor += 25

    if stok_minimal > 0:
        defisit = max(0, stok_minimal - stok)
        skor += min(20, defisit * 4)
    skor += min(20, kebutuhan * 3)
    skor += min(15, terjual * 3)

    if qty_final > 0:
        skor += 5

    return min(100, skor)

def _hitung_qty_dan_dana_per_metode(item):
    """
    Menampilkan hasil masing-masing metode secara asli.
    Jangan dibatasi kebutuhan_total, supaya hasil fuzzy kelihatan.
    """
    harga_beli = max(0, _angka_aman(item.get("harga_beli"), 0))

    qty_mamdani = max(0, _angka_aman(item.get("hasil_mamdani"), 0))
    qty_tsukamoto = max(0, _angka_aman(item.get("hasil_tsukamoto"), 0))
    qty_sugeno = max(0, _angka_aman(item.get("hasil_sugeno"), 0))

    dana_mamdani = qty_mamdani * harga_beli
    dana_tsukamoto = qty_tsukamoto * harga_beli
    dana_sugeno = qty_sugeno * harga_beli

    return {
        "qty_mamdani": qty_mamdani,
        "qty_tsukamoto": qty_tsukamoto,
        "qty_sugeno": qty_sugeno,

        "dana_mamdani": dana_mamdani,
        "dana_tsukamoto": dana_tsukamoto,
        "dana_sugeno": dana_sugeno,

        "dana_mamdani_format": _format_rupiah_angka(dana_mamdani),
        "dana_tsukamoto_format": _format_rupiah_angka(dana_tsukamoto),
        "dana_sugeno_format": _format_rupiah_angka(dana_sugeno),
    }

def _norm_key_barang(nama_barang="", jenis_barang="", satuan=""):
    """
    Helper untuk mencocokkan barang transaksi dengan master barang
    ketika kode barang di transaksi berbeda dengan kode barang master.
    """
    nama = _teks_aman(nama_barang).strip().lower()
    jenis = _teks_aman(jenis_barang).strip().lower()
    satuan = _teks_aman(satuan).strip().lower()

    nama = re.sub(r"\s+", " ", nama)
    jenis = re.sub(r"\s+", " ", jenis)
    satuan = re.sub(r"\s+", " ", satuan)

    return f"{nama}|{jenis}|{satuan}"

def _norm_nama_barang_only(nama_barang=""):
    """
    Fallback terakhir untuk mencocokkan barang berdasarkan nama saja.
    Dipakai untuk kasus transaksi lama yang kode/jenis snapshot-nya tidak sama
    dengan master barang, seperti kasus hama.
    """
    nama = _teks_aman(nama_barang).strip().lower()
    nama = re.sub(r"\s+", " ", nama)
    return nama


def _norm_key_barang_tanpa_satuan(nama_barang="", jenis_barang=""):
    """
    Cadangan kalau data transaksi lama belum menyimpan satuan.
    Dipakai hati-hati untuk kasus seperti hama yang stok masternya ada,
    tapi transaksi tidak nyambung ke kode master.
    """
    nama = _teks_aman(nama_barang).strip().lower()
    jenis = _teks_aman(jenis_barang).strip().lower()

    nama = re.sub(r"\s+", " ", nama)
    jenis = re.sub(r"\s+", " ", jenis)

    return f"{nama}|{jenis}"

def _ambil_data_barang_rekomendasi(periode_mulai, periode_selesai, jenis_filter="SEMUA"):
    hasil_map = {}

    penjualan_map = _ambil_penjualan_periode_map(
        periode_mulai,
        periode_selesai,
        jenis_filter,
    )

    index_transaksi_nama_jenis = {}
    index_transaksi_nama_only = {}

    # 1. Masukkan data transaksi dulu.
    for key, row in penjualan_map.items():
        kode = _teks_aman(row.get("kode_barang"))
        nama = _teks_aman(row.get("nama_barang"))
        jenis = _teks_aman(row.get("jenis_barang"))
        satuan = _teks_aman(row.get("satuan"))

        key_nama_only = row.get("key_nama_only") or _norm_nama_barang_only(nama)
        key_nama_jenis = row.get("key_nama_jenis") or f"{key_nama_only}|{jenis.lower()}"

        key_final = kode if kode else key_nama_jenis

        hasil_map[key_final] = {
            "id_barang": 0,

            "kode": kode,
            "kode_barang": kode,

            "nama": nama,
            "nama_barang": nama,
            "jenis_barang": jenis,

            "satuan": satuan,
            "satuan_barang": satuan,
            "nama_barang_label": f"{nama} - {satuan}" if satuan else nama,

            "stok": 0,
            "stok_saat_ini": 0,
            "stok_minimal": 0,

            "harga_beli": _angka_aman(row.get("harga_beli"), 0),
            "terjual_periode": _angka_aman(row.get("total_qty"), 0),

            "id_supplier": 0,
            "nama_supplier": "",
            "nama_perusahaan": "",
            "supplier_label": "",

            "key_nama_only": key_nama_only,
            "key_nama_jenis": key_nama_jenis,
        }

        if key_nama_jenis:
            index_transaksi_nama_jenis[key_nama_jenis] = key_final

        if key_nama_only:
            index_transaksi_nama_only[key_nama_only] = key_final

    # 2. Hitung nama master yang unik.
    # Nama saja dipakai fallback hanya kalau tidak dobel.
    master_nama_count = defaultdict(int)

    for b in TbBarang.objects.all():
        nama_master = _norm_nama_barang_only(getattr(b, "nama_barang", ""))
        if nama_master:
            master_nama_count[nama_master] += 1

    # 3. Gabungkan transaksi dengan master barang.
    qs = TbBarang.objects.all().order_by("jenis_barang", "kode_barang")

    for b in qs:
        kode = _teks_aman(getattr(b, "kode_barang", ""))
        nama = _teks_aman(getattr(b, "nama_barang", ""))
        jenis = _teks_aman(getattr(b, "jenis_barang", ""))
        satuan = _teks_aman(getattr(b, "satuan", ""))

        if not kode and not nama:
            continue

        key_nama_only_master = _norm_nama_barang_only(nama)
        key_nama_jenis_master = f"{key_nama_only_master}|{jenis.lower()}"

        nama_master_unik = (
            key_nama_only_master
            and master_nama_count.get(key_nama_only_master, 0) == 1
        )

        ada_di_transaksi = (
            kode in hasil_map
            or key_nama_jenis_master in index_transaksi_nama_jenis
            or (
                nama_master_unik
                and key_nama_only_master in index_transaksi_nama_only
            )
        )

        if _teks_aman(jenis_filter).upper() != "SEMUA":
            jenis_master_bersih = jenis.strip().upper()
            jenis_filter_bersih = _teks_aman(jenis_filter).strip().upper()

            if jenis_master_bersih != jenis_filter_bersih and not ada_di_transaksi:
                continue

        supplier_info = _ambil_supplier_barang_info(b)

        id_barang = _angka_aman(getattr(b, "id_barang", 0), 0)
        stok_saat_ini = _angka_aman(getattr(b, "stok", 0), 0)
        stok_minimal = _angka_aman(getattr(b, "stok_minimal", 0), 0)
        harga_beli = _angka_aman(getattr(b, "harga_beli", 0), 0)

        if kode and kode in hasil_map:
            key_target = kode
        elif key_nama_jenis_master in index_transaksi_nama_jenis:
            key_target = index_transaksi_nama_jenis[key_nama_jenis_master]
        elif nama_master_unik and key_nama_only_master in index_transaksi_nama_only:
            key_target = index_transaksi_nama_only[key_nama_only_master]
        else:
            key_target = kode if kode else key_nama_jenis_master

        terjual_periode = 0
        harga_beli_lama = 0

        if key_target in hasil_map:
            terjual_periode = _angka_aman(hasil_map[key_target].get("terjual_periode"), 0)
            harga_beli_lama = _angka_aman(hasil_map[key_target].get("harga_beli"), 0)

        hasil_map[key_target] = {
            "id_barang": id_barang,

            "kode": kode,
            "kode_barang": kode,

            "nama": nama,
            "nama_barang": nama,
            "jenis_barang": jenis,

            "satuan": satuan,
            "satuan_barang": satuan,
            "nama_barang_label": f"{nama} - {satuan}" if satuan else nama,

            "stok": stok_saat_ini,
            "stok_saat_ini": stok_saat_ini,
            "stok_minimal": stok_minimal,

            "harga_beli": harga_beli if harga_beli > 0 else harga_beli_lama,
            "terjual_periode": terjual_periode,

            "id_supplier": supplier_info.get("id_supplier", 0),
            "nama_supplier": supplier_info.get("nama_supplier", ""),
            "nama_perusahaan": supplier_info.get("nama_perusahaan", ""),
            "supplier_label": supplier_info.get("supplier_label", ""),

            "key_nama_only": key_nama_only_master,
            "key_nama_jenis": key_nama_jenis_master,
        }

    return list(hasil_map.values())

def _ambil_penjualan_periode_map(periode_mulai, periode_selesai, jenis_filter="SEMUA"):
    hasil = {}

    query = """
        SELECT
            COALESCE(
                NULLIF(TRIM(td.kode_barang_snapshot), ''),
                TRIM(b_id.kode_barang),
                TRIM(b_kode.kode_barang),
                ''
            ) AS kode_barang_final,

            COALESCE(
                NULLIF(TRIM(td.nama_barang_snapshot), ''),
                TRIM(b_id.nama_barang),
                TRIM(b_kode.nama_barang),
                ''
            ) AS nama_barang_final,

            COALESCE(
                NULLIF(TRIM(td.jenis_barang_snapshot), ''),
                TRIM(b_id.jenis_barang),
                TRIM(b_kode.jenis_barang),
                ''
            ) AS jenis_barang_final,

            COALESCE(
                NULLIF(TRIM(td.satuan_snapshot), ''),
                TRIM(b_id.satuan),
                TRIM(b_kode.satuan),
                ''
            ) AS satuan_final,

            COALESCE(
                td.harga_beli_snapshot,
                b_id.harga_beli,
                b_kode.harga_beli,
                0
            ) AS harga_beli_final,

            SUM(COALESCE(td.qty, 0)) AS total_qty

        FROM tb_penjualan_detail td
        INNER JOIN tb_penjualan t
            ON t.id_transaksi = td.id_transaksi

        LEFT JOIN tb_barang b_id
            ON b_id.id_barang = td.id_barang

        LEFT JOIN tb_barang b_kode
            ON b_kode.kode_barang = td.kode_barang_snapshot

        WHERE DATE(t.tanggal_waktu) BETWEEN %s AND %s
    """

    params = [periode_mulai, periode_selesai]

    if _teks_aman(jenis_filter).upper() != "SEMUA":
        query += """
            AND UPPER(TRIM(
                COALESCE(
                    NULLIF(TRIM(td.jenis_barang_snapshot), ''),
                    TRIM(b_id.jenis_barang),
                    TRIM(b_kode.jenis_barang),
                    ''
                )
            )) = %s
        """
        params.append(_teks_aman(jenis_filter).strip().upper())

    query += """
        GROUP BY
            COALESCE(
                NULLIF(TRIM(td.kode_barang_snapshot), ''),
                TRIM(b_id.kode_barang),
                TRIM(b_kode.kode_barang),
                ''
            ),
            COALESCE(
                NULLIF(TRIM(td.nama_barang_snapshot), ''),
                TRIM(b_id.nama_barang),
                TRIM(b_kode.nama_barang),
                ''
            ),
            COALESCE(
                NULLIF(TRIM(td.jenis_barang_snapshot), ''),
                TRIM(b_id.jenis_barang),
                TRIM(b_kode.jenis_barang),
                ''
            ),
            COALESCE(
                NULLIF(TRIM(td.satuan_snapshot), ''),
                TRIM(b_id.satuan),
                TRIM(b_kode.satuan),
                ''
            ),
            COALESCE(
                td.harga_beli_snapshot,
                b_id.harga_beli,
                b_kode.harga_beli,
                0
            )
    """

    with connection.cursor() as cursor:
        cursor.execute(query, params)
        rows = cursor.fetchall()

    for kode_barang, nama_barang, jenis_barang, satuan, harga_beli, total_qty in rows:
        kode = _teks_aman(kode_barang)
        nama = _teks_aman(nama_barang)
        jenis = _teks_aman(jenis_barang)
        satuan = _teks_aman(satuan)

        if not kode and not nama:
            continue

        key_nama_only = _norm_nama_barang_only(nama)
        key_nama_jenis = f"{key_nama_only}|{jenis.lower()}"

        key = kode if kode else key_nama_jenis

        hasil[key] = {
            "kode_barang": kode,
            "nama_barang": nama,
            "jenis_barang": jenis,
            "satuan": satuan,
            "harga_beli": _angka_aman(harga_beli, 0),
            "total_qty": _angka_aman(total_qty, 0),
            "key_nama_only": key_nama_only,
            "key_nama_jenis": key_nama_jenis,
        }

    return hasil


def _ambil_daftar_jenis_rekomendasi():
    hasil = set()

    try:
        jenis_master = (
            TbBarang.objects.exclude(jenis_barang__isnull=True)
            .exclude(jenis_barang__exact="")
            .values_list("jenis_barang", flat=True)
            .distinct()
        )
        for j in jenis_master:
            teks = _teks_aman(j)
            if teks:
                hasil.add(teks)
    except Exception:
        pass

    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT DISTINCT
                    COALESCE(NULLIF(jenis_barang_snapshot, ''), '')
                FROM tb_penjualan_detail
                WHERE COALESCE(NULLIF(jenis_barang_snapshot, ''), '') <> ''
            """)
            for (j,) in cursor.fetchall():
                teks = _teks_aman(j)
                if teks:
                    hasil.add(teks)
    except Exception:
        pass

    return sorted(list(hasil), key=lambda x: x.lower())


def _ambil_stok_urgent_beranda(jumlah_data=10, hari_penjualan=30):
    try:
        hari_ini = timezone.localdate(_get_timezone_wib())
    except Exception:
        hari_ini = timezone.localdate()

    periode_mulai = hari_ini - timezone.timedelta(days=hari_penjualan - 1)
    periode_selesai = hari_ini

    penjualan_map = _ambil_penjualan_periode_map(
        periode_mulai,
        periode_selesai,
        "SEMUA",
    )

    hasil = []

    qs = TbBarang.objects.all().order_by("jenis_barang", "kode_barang")

    for b in qs:
        id_barang = int(getattr(b, "id_barang", 0) or 0)
        kode_barang = _teks_aman(getattr(b, "kode_barang", ""))
        nama_barang = _teks_aman(getattr(b, "nama_barang", ""))
        jenis_barang = _teks_aman(getattr(b, "jenis_barang", ""))
        satuan = _teks_aman(getattr(b, "satuan", ""))

        stok_saat_ini = _angka_aman(
            getattr(b, "stok_saat_ini", getattr(b, "stok", 0)),
            0
        )

        stok_minimal = _angka_aman(getattr(b, "stok_minimal", 0), 0)
        harga_beli = _angka_aman(getattr(b, "harga_beli", 0), 0)

        row_penjualan = penjualan_map.get(kode_barang, {})
        terjual_periode = _angka_aman(row_penjualan.get("total_qty"), 0)

        kebutuhan_info = _hitung_kebutuhan_hitung_manual(
            terjual_periode=terjual_periode,
            stok_saat_ini=stok_saat_ini,
            jumlah_hari=max(1, hari_penjualan),
            stok_minimal=stok_minimal,
        )

        target_restock = _angka_aman(kebutuhan_info.get("kebutuhan_total"), 0)
        proyeksi_30_hari = _angka_aman(kebutuhan_info.get("proyeksi_30_hari"), 0)
        target_stok_aman = _angka_aman(kebutuhan_info.get("target_stok_aman"), 0)

        perlu_restock = (
            stok_saat_ini <= 0
            or (stok_minimal > 0 and stok_saat_ini <= stok_minimal)
            or target_restock > 0
        )

        if not perlu_restock:
            continue

        if stok_saat_ini <= 0 and terjual_periode > 0:
            status_text = "Sangat Perlu"
            status_class = "urgent"
            alasan = "Stok habis dan masih ada penjualan dalam 30 hari terakhir."
            skor_urgent = 100 + terjual_periode

        elif stok_saat_ini <= 0:
            status_text = "Stok Habis"
            status_class = "urgent"
            alasan = "Stok sudah habis, perlu segera restock."
            skor_urgent = 90

        elif stok_minimal > 0 and stok_saat_ini <= stok_minimal:
            status_text = "Perlu Restock"
            status_class = "perlu-restock"
            alasan = f"Stok saat ini sudah berada pada atau di bawah batas minimal {stok_minimal}."
            skor_urgent = 70 + min(30, terjual_periode)

        elif target_restock > 0:
            status_text = "Stok Menipis"
            status_class = "perlu"
            alasan = "Berdasarkan penjualan 30 hari terakhir, stok perlu ditambah agar aman."
            skor_urgent = 50 + min(30, terjual_periode)

        else:
            status_text = "Aman"
            status_class = "aman"
            alasan = "Stok masih aman."
            skor_urgent = 0

        hasil.append({
            "id_barang": id_barang,

            "kode": kode_barang,
            "kode_barang": kode_barang,

            "nama": nama_barang,
            "nama_barang": nama_barang,

            "jenis_barang": jenis_barang,
            "satuan": satuan,

            "stok": stok_saat_ini,
            "stok_saat_ini": stok_saat_ini,
            "stok_minimal": stok_minimal,

            "terjual_periode": terjual_periode,
            "proyeksi_30_hari": proyeksi_30_hari,
            "target_stok_aman": target_stok_aman,

            "target_restock": target_restock,
            "kebutuhan_total": target_restock,

            "harga_beli": harga_beli,

            "status_text": status_text,
            "status_label": status_text,
            "status_class": status_class,

            "alasan": alasan,
            "skor_urgent": skor_urgent,
        })

    hasil = sorted(
        hasil,
        key=lambda x: (
            -_angka_aman(x.get("skor_urgent"), 0),
            _angka_aman(x.get("stok_saat_ini"), 0),
            -_angka_aman(x.get("terjual_periode"), 0),
            x.get("kode_barang", "")
        )
    )

    return hasil[:jumlah_data]


@require_http_methods(["GET"])
def rekomendasi_halaman(request):
    gate = _require_login(request)
    if gate:
        return gate

    if not _is_pemilik(request):
        messages.error(request, "Akses ditolak.")
        return redirect("kasir:home")

    hari_ini = timezone.localdate()
    periode_mulai = hari_ini - timedelta(days=30)
    periode_selesai = hari_ini

    daftar_barang = _ambil_data_barang_rekomendasi(periode_mulai, periode_selesai)
    domain_global = _ambil_domain_global_rekomendasi(periode_mulai, periode_selesai)

    for item in daftar_barang:
        # Fuzzy
        alpha_penjualan = _membership_input(item.get("terjual_periode", 0), domain_global["penjualan"])
        alpha_stok = _membership_input(item.get("stok_saat_ini", 0), domain_global["stok"])
        aturan = _aturan_fuzzy(alpha_penjualan, alpha_stok, item.get("target_stok_aman", 0))

        fuzzy = _evaluasi_fuzzy_item(item, aturan)
        item.update(fuzzy)

        item["prioritas"] = _hitung_skor_urgent(item)

        jenis_barang_list = _ambil_daftar_jenis_rekomendasi()

        return render(request, "kasir/rekomendasi.html", {
            "role": request.session.get("kasir_role", ""),
            "nama": request.session.get("kasir_nama", ""),
            "daftar_barang": daftar_barang,
            "jenis_barang_list": jenis_barang_list,
        })

def _status_kebutuhan_dari_item(item):
    kebutuhan_total = _angka_aman(item.get("kebutuhan_total"), 0)
    stok = _angka_aman(item.get("stok_saat_ini"), 0)
    terjual = _angka_aman(item.get("terjual_periode"), 0)

    if kebutuhan_total <= 0:
        return "Aman"
    if stok <= 0 and terjual > 0:
        return "Sangat Perlu"
    return "Perlu"


def _label_prioritas_dari_skor(nilai_prioritas):
    nilai_prioritas = _angka_aman(nilai_prioritas, 0)

    if nilai_prioritas >= 70:
        return "Tinggi"
    if nilai_prioritas >= 35:
        return "Sedang"
    return "Rendah"


def _alasan_singkat_rekomendasi(item):
    stok = _angka_aman(item.get("stok_saat_ini"), 0)
    terjual = _angka_aman(item.get("terjual_periode"), 0)
    target = _angka_aman(item.get("rekomendasi_beli", item.get("target_restock", 0)), 0)
    qty_final = _angka_aman(item.get("qty_final_beli"), 0)

    if target <= 0:
        return "Stok masih mencukupi pada periode ini."
    if qty_final <= 0:
        return "Perlu restock, tetapi dana belum cukup."
    if qty_final < target:
        return "Perlu restock, pembelian disesuaikan dengan batas dana."
    if stok <= 0 and terjual > 0:
        return "Stok habis, penjualan masih berjalan."
    if stok < terjual and terjual > 0:
        return "Stok rendah dibanding penjualan periode ini."
    return "Ada kebutuhan tambahan untuk menjaga stok."


def _penjelasan_sederhana_rekomendasi(item):
    stok = _angka_aman(item.get("stok_saat_ini"), 0)
    terjual = _angka_aman(item.get("terjual_periode"), 0)
    rekom = _angka_aman(item.get("rekomendasi_beli", item.get("target_restock", 0)), 0)
    qty = _angka_aman(item.get("qty_final_beli"), 0)

    kondisi = "stok masih mencukupi pada periode yang dipilih"
    if stok <= 0 and terjual > 0:
        kondisi = "stok saat ini sudah habis sementara penjualan masih berjalan"
    elif stok < terjual and terjual > 0:
        kondisi = "stok lebih rendah daripada penjualan pada periode yang dipilih"

    if rekom <= 0:
        return (
            "Pada periode yang dipilih, stok barang ini masih mencukupi "
            "sehingga sistem belum menyarankan pembelian tambahan."
        )

    if qty <= 0:
        return (
            f"Barang ini tetap diperhatikan karena {kondisi}. "
            f"Sistem membaca kebutuhan awal {rekom} unit, tetapi pembelian belum dilakukan "
            f"karena sisa dana belum cukup."
        )

    if qty < rekom:
        return (
            f"Barang ini dipilih karena {kondisi}. "
            f"Sistem menyarankan pembelian {rekom} unit, lalu menyesuaikannya menjadi {qty} unit "
            f"agar tetap sesuai dengan batas dana."
        )

    return (
        f"Barang ini dipilih karena {kondisi}. "
        f"Sistem menyarankan pembelian {rekom} unit dan seluruh kebutuhan tersebut masih bisa dipenuhi."
    )


def _prioritas_dari_item(item):
    skor = _hitung_skor_urgent(item)

    stok = _angka_aman(item.get("stok_saat_ini"), 0)
    stok_minimal = _angka_aman(item.get("stok_minimal"), 0)
    terjual = _angka_aman(item.get("terjual_periode"), 0)
    kebutuhan = _angka_aman(item.get("kebutuhan_total"), 0)
    qty_final = _angka_aman(item.get("qty_final_beli"), 0)

    if stok <= 0 and terjual > 0:
        penjelasan = (
            f"Stok habis dan masih ada penjualan {terjual} barang pada periode ini. "
            f"Kebutuhan ideal dihitung {kebutuhan} barang agar stok tidak kosong lagi."
        )
    elif stok <= 0:
        penjelasan = (
            f"Stok habis. Kebutuhan ideal dihitung {kebutuhan} barang berdasarkan stok minimal."
        )
    elif stok_minimal > 0 and stok < stok_minimal:
        penjelasan = (
            f"Stok saat ini {stok} berada di bawah stok minimal {stok_minimal}. "
            f"Kebutuhan ideal dihitung {kebutuhan} barang."
        )
    elif stok_minimal > 0 and stok == stok_minimal:
        penjelasan = (
            f"Stok saat ini sama dengan stok minimal {stok_minimal}. "
            f"Barang perlu dipantau agar tidak kehabisan."
        )
    elif terjual > 0:
        penjelasan = (
            f"Barang masih terjual {terjual} pada periode ini. "
            f"Sistem mempertimbangkan penjualan dan stok minimal untuk menjaga ketersediaan."
        )
    else:
        penjelasan = (
            "Prioritas dihitung dari stok saat ini, stok minimal, kebutuhan restock, dan penjualan periode."
        )

    if kebutuhan > 0 and qty_final <= 0:
        penjelasan += " Barang belum masuk pembelian karena dana belum mencukupi."
    elif qty_final > 0 and qty_final < kebutuhan:
        penjelasan += f" Final beli menjadi {qty_final} barang karena menyesuaikan batas dana."
    elif qty_final > 0:
        penjelasan += f" Final beli {qty_final} barang."

    return skor, penjelasan


def _alokasi_budget_final(daftar_item, budget):
   
    budget = max(0, _angka_aman(budget))
    sisa = budget

    urut = sorted(
        daftar_item,
        key=lambda x: (
            -_angka_aman(x.get("prioritas"), 0),
            -_angka_aman(x.get("rekomendasi_beli"), 0),
            -_angka_aman(x.get("terjual_periode"), 0),
            _angka_aman(x.get("harga_beli"), 0),
            x.get("kode_barang", "")
        )
    )

    for item in urut:
        item["qty_final_beli"] = 0
        item["dana_final_beli"] = 0
        item["rumus_alokasi"] = []
        item["info_alokasi"] = []

        target = max(
            0,
            _angka_aman(
                item.get("rekomendasi_beli")
                or item.get("rekomendasi_ideal")
                or item.get("target_restock"),
                0,
            )
        )

        item["rekomendasi_beli"] = target
        item["rekomendasi_ideal"] = target
        item["target_restock"] = target

    for item in urut:
        harga_beli = max(0, _angka_aman(item.get("harga_beli"), 0))
        target = max(0, _angka_aman(item.get("rekomendasi_beli"), 0))

        if target <= 0:
            item["rumus_alokasi"].append(
                "Tidak dialokasikan karena rekomendasi fuzzy bernilai 0."
            )
            item["info_alokasi"].append(
                "Tidak dialokasikan karena rekomendasi fuzzy bernilai 0."
            )
            continue

        if harga_beli <= 0:
            item["rumus_alokasi"].append(
                "Tidak dialokasikan karena harga beli tidak valid."
            )
            item["info_alokasi"].append(
                "Tidak dialokasikan karena harga beli tidak valid."
            )
            continue

        if sisa < harga_beli:
            item["rumus_alokasi"].append(
                f"Sisa dana {_format_rupiah_angka(sisa)} belum cukup untuk membeli 1 barang "
                f"dengan harga {_format_rupiah_angka(harga_beli)}."
            )
            item["info_alokasi"].append(
                "Belum masuk pembelian karena sisa dana tidak cukup untuk pembelian minimal 1 barang."
            )
            continue

        qty_mampu = sisa // harga_beli
        qty_bisa = min(target, qty_mampu)
        dana = qty_bisa * harga_beli

        item["qty_final_beli"] = int(qty_bisa)
        item["dana_final_beli"] = int(dana)

        sisa -= dana

        item["rumus_alokasi"].append(
            f"Rekomendasi ideal fuzzy = {target} barang."
        )
        item["rumus_alokasi"].append(
            f"Harga beli = {_format_rupiah_angka(harga_beli)}."
        )
        item["rumus_alokasi"].append(
            f"Final beli = min({target}, {sisa + dana} // {harga_beli}) = {qty_bisa} barang."
        )
        item["rumus_alokasi"].append(
            f"Dana dipakai = {qty_bisa} × {_format_rupiah_angka(harga_beli)} = {_format_rupiah_angka(dana)}."
        )

        if qty_bisa < target:
            item["rumus_alokasi"].append(
                "Final beli lebih kecil dari rekomendasi ideal karena menyesuaikan batas dana."
            )
        else:
            item["rumus_alokasi"].append(
                "Final beli sudah memenuhi rekomendasi ideal fuzzy."
            )

        item["info_alokasi"] = list(item["rumus_alokasi"])

    total_digunakan = budget - sisa
    return int(total_digunakan), int(sisa)


def _ambil_prioritas_utama(hasil_items):
    if not hasil_items:
        return None

    kandidat = sorted(
        hasil_items,
        key=lambda x: (
            -_angka_aman(x.get("prioritas"), 0),
            -_angka_aman(x.get("kebutuhan_total"), 0),
            -_angka_aman(x.get("terjual_periode"), 0),
            x.get("kode_barang", "")
        )
    )
    return kandidat[0]


def _simpan_header_rekomendasi_db(*args, **kwargs):
    return 0


def _simpan_detail_rekomendasi_db(*args, **kwargs):
    return None

def _label_metode_terpilih(status_metode, metode_terbaik_item, kandidat_terbaik):
    kandidat_terbaik = kandidat_terbaik or []

    if status_metode == "setara":
        if kandidat_terbaik:
            return f"Setara ({', '.join([x.title() for x in kandidat_terbaik])})"
        return "Setara"

    return (metode_terbaik_item or "-").title()


def _penjelasan_metode_sederhana(status_metode, metode_terbaik_item, kandidat_terbaik, rekomendasi_item):
    kandidat_terbaik = kandidat_terbaik or []

    if status_metode == "setara":
        if kandidat_terbaik:
            daftar = ", ".join([x.title() for x in kandidat_terbaik])
            return (
                f"Sistem menghitung 3 metode fuzzy. Pada item ini, hasil terbaik muncul setara pada {daftar}, "
                f"sehingga sistem memakai hasil yang paling aman sesuai kebutuhan barang."
            )
        return (
            "Sistem menghitung 3 metode fuzzy. Pada item ini terdapat hasil yang setara, "
            "sehingga sistem memakai hasil yang paling aman sesuai kebutuhan barang."
        )

    return (
        f"Sistem menghitung 3 metode fuzzy dan memilih metode {metode_terbaik_item.title()} "
        f"karena hasilnya paling sesuai untuk kondisi barang ini. "
        f"Hasil rekomendasi awal dari metode terpilih adalah {rekomendasi_item} unit."
    )


def _ringkasan_metode_periode(hasil_items):
    hitung = {
        "mamdani": 0,
        "tsukamoto": 0,
        "sugeno": 0,
        "setara": 0,
    }

    for item in hasil_items:
        metode = _teks_aman(item.get("metode_terbaik_item"), "setara").lower()
        if metode in hitung:
            hitung[metode] += 1
        else:
            hitung["setara"] += 1

    urut = sorted(hitung.items(), key=lambda x: (-x[1], x[0]))
    nama, jumlah = urut[0]

    if jumlah <= 0:
        return {
            "label": "-",
            "keterangan": "Belum ada metode yang dominan pada periode ini."
        }

    label = "Setara" if nama == "setara" else nama.title()
    return {
        "label": label,
        "keterangan": (
            f"Metode ini paling sering menjadi acuan pada item yang direkomendasikan di periode ini."
        )
    }


@require_http_methods(["POST"])
def rekomendasi_proses(request):
    gate = _require_login(request)
    if gate:
        return JsonResponse({"ok": False, "msg": "Unauthorized"}, status=401)

    try:
        payload = json.loads(request.body.decode("utf-8")) if (
            request.content_type and "application/json" in request.content_type
        ) else request.POST
    except Exception:
        return JsonResponse({"ok": False, "msg": "Payload tidak valid."}, status=400)

    def _label_metode(status, metode, kandidat):
        kandidat = kandidat or []
        if str(status or "").lower() == "setara":
            if kandidat:
                return f"Setara ({', '.join([str(x).title() for x in kandidat])})"
            return "Setara"

        teks = str(metode or "").strip()
        return teks.title() if teks else "-"

    def _label_status_kebutuhan(status):
        raw = str(status or "").strip().lower()
        if raw == "aman":
            return "Aman"
        if raw == "belum_terbeli":
            return "Belum Terbeli"
        if raw == "sangat_perlu_restock":
            return "Sangat Perlu Restock"
        return "Perlu Restock"

    def _label_prioritas_item(nilai):
        nilai = _angka_aman(nilai, 0)
        if nilai >= 70:
            return "Tinggi"
        if nilai >= 35:
            return "Sedang"
        return "Rendah"

    def _status_kebutuhan_item(stok, stok_minimal, terjual, rekomendasi_beli, qty_final_beli):
        stok = _angka_aman(stok, 0)
        stok_minimal = _angka_aman(stok_minimal, 0)
        terjual = _angka_aman(terjual, 0)
        rekomendasi_beli = _angka_aman(rekomendasi_beli, 0)
        qty_final_beli = _angka_aman(qty_final_beli, 0)

        if rekomendasi_beli <= 0:
            return "aman"

        if qty_final_beli <= 0:
            return "belum_terbeli"

        if stok <= 0 and (terjual > 0 or stok_minimal > 0):
            return "sangat_perlu_restock"

        return "perlu_restock"

    def _alasan_singkat_item(stok, stok_minimal, terjual, rekomendasi_beli, qty_final_beli):
        stok = _angka_aman(stok, 0)
        stok_minimal = _angka_aman(stok_minimal, 0)
        terjual = _angka_aman(terjual, 0)
        rekomendasi_beli = _angka_aman(rekomendasi_beli, 0)
        qty_final_beli = _angka_aman(qty_final_beli, 0)

        if rekomendasi_beli <= 0:
            return "Stok masih mencukupi pada periode ini."

        if qty_final_beli <= 0:
            return "Butuh restock, tetapi belum masuk pembelian karena dana tidak cukup."

        if stok <= 0 and terjual > 0:
            return "Stok habis dan penjualan masih berjalan."

        if stok_minimal > 0 and stok < stok_minimal:
            return "Stok berada di bawah batas minimal."

        if qty_final_beli < rekomendasi_beli:
            return "Pembelian disesuaikan dengan batas dana."

        return "Perlu menambah stok untuk menjaga ketersediaan."

    def _penjelasan_metode_sederhana_item(
        status_metode,
        metode_terbaik_item,
        kandidat_terbaik,
        hasil_mamdani,
        hasil_tsukamoto,
        hasil_sugeno,
        rekomendasi_fuzzy,
        rekomendasi_beli,
        kebutuhan_stok_minimal,
    ):
        label = _label_metode(status_metode, metode_terbaik_item, kandidat_terbaik)

        if str(status_metode or "").lower() == "setara":
            return (
                f"Hasil fuzzy setara pada metode {label}. "
                f"Mamdani menghasilkan {hasil_mamdani}, Tsukamoto menghasilkan {hasil_tsukamoto}, "
                f"dan Sugeno menghasilkan {hasil_sugeno}. "
                f"Sistem mengambil nilai fuzzy setara sebagai rekomendasi ideal {rekomendasi_fuzzy} barang. "
                f"Stok minimal digunakan sebagai bagian dari input dan pembanding, bukan sebagai penambah hasil akhir. "
                f"Rekomendasi ideal akhir menjadi {rekomendasi_beli} barang."
            )

        return (
            f"Metode fuzzy terbaik untuk barang ini adalah {label}. "
            f"Mamdani menghasilkan {hasil_mamdani}, Tsukamoto menghasilkan {hasil_tsukamoto}, "
            f"dan Sugeno menghasilkan {hasil_sugeno}. "
            f"Sistem memilih hasil {label} sebagai rekomendasi fuzzy {rekomendasi_fuzzy} barang. "
            f"Stok minimal digunakan sebagai input dan pembanding kebutuhan, bukan ditambahkan langsung ke hasil akhir. "
            f"Rekomendasi ideal akhir menjadi {rekomendasi_beli} barang."
        )

    def _penjelasan_sederhana_item(
        stok,
        stok_minimal,
        terjual,
        rata_harian,
        proyeksi_30_hari,
        target_stok_aman,
        kebutuhan_total,
        rekomendasi_beli,
        qty_final_beli,
        dana_final_beli,
    ):
        stok = _angka_aman(stok, 0)
        stok_minimal = _angka_aman(stok_minimal, 0)
        terjual = _angka_aman(terjual, 0)
        proyeksi_30_hari = _angka_aman(proyeksi_30_hari, 0)
        target_stok_aman = _angka_aman(target_stok_aman, 0)
        kebutuhan_total = _angka_aman(kebutuhan_total, 0)
        rekomendasi_beli = _angka_aman(rekomendasi_beli, 0)
        qty_final_beli = _angka_aman(qty_final_beli, 0)
        dana_final_beli = _angka_aman(dana_final_beli, 0)

        if rekomendasi_beli <= 0:
            return (
                f"Barang terjual {terjual} pada periode ini dengan stok saat ini {stok}. "
                f"Stok minimal {stok_minimal}. Karena stok masih mencukupi, sistem tidak menyarankan pembelian."
            )

        if qty_final_beli <= 0:
            return (
                f"Target stok aman adalah {target_stok_aman} barang. "
                f"Kebutuhan ideal {kebutuhan_total} barang, tetapi barang belum masuk pembelian karena dana belum cukup."
            )

        if qty_final_beli < rekomendasi_beli:
            return (
                f"Proyeksi kebutuhan 30 hari {proyeksi_30_hari} barang dan stok minimal {stok_minimal}. "
                f"Sistem menyarankan {rekomendasi_beli} barang, tetapi final beli menjadi {qty_final_beli} barang "
                f"karena menyesuaikan batas dana."
            )

        return (
            f"Stok saat ini {stok}, stok minimal {stok_minimal}, dan terjual {terjual} pada periode ini. "
            f"Sistem menyarankan pembelian {rekomendasi_beli} barang. "
            f"Final beli {qty_final_beli} barang dengan dana {_format_rupiah_angka(dana_final_beli)}."
        )

    tanggal_mulai = _teks_aman(payload.get("tanggal_mulai"))
    tanggal_selesai = _teks_aman(payload.get("tanggal_selesai"))
    jenis_barang = _teks_aman(payload.get("jenis_barang"), "SEMUA")
    dana_maksimal = _angka_aman(payload.get("dana_maksimal"), 0)

    if not tanggal_mulai:
        return JsonResponse({"ok": False, "msg": "Tanggal mulai wajib diisi."}, status=400)

    if not tanggal_selesai:
        return JsonResponse({"ok": False, "msg": "Tanggal selesai wajib diisi."}, status=400)

    try:
        periode_mulai = date.fromisoformat(tanggal_mulai)
        periode_selesai = date.fromisoformat(tanggal_selesai)
    except Exception:
        return JsonResponse({"ok": False, "msg": "Format tanggal tidak valid."}, status=400)

    if periode_mulai > periode_selesai:
        return JsonResponse({"ok": False, "msg": "Tanggal mulai tidak boleh lebih besar dari tanggal selesai."}, status=400)

    if dana_maksimal <= 0:
        return JsonResponse({"ok": False, "msg": "Dana maksimal wajib diisi."}, status=400)

    daftar_barang = _ambil_data_barang_rekomendasi(
        periode_mulai,
        periode_selesai,
        jenis_barang,
    )

    if not daftar_barang:
        return JsonResponse({"ok": False, "msg": "Data barang tidak ditemukan."}, status=404)

    jumlah_hari_periode = (periode_selesai - periode_mulai).days + 1

    domain_global = _ambil_domain_global_rekomendasi(periode_mulai, periode_selesai)
    domain_penjualan = domain_global["penjualan"]
    domain_stok = domain_global["stok"]

    hasil_items = []
    total_error_mamdani = 0
    total_error_tsukamoto = 0
    total_error_sugeno = 0

    for barang in daftar_barang:
        id_barang = _angka_aman(barang.get("id_barang"), 0)
        kode_barang = _teks_aman(barang.get("kode_barang"))
        nama_barang = _teks_aman(barang.get("nama_barang"))
        jenis_barang_item = _teks_aman(barang.get("jenis_barang"))
        satuan = _teks_aman(barang.get("satuan", barang.get("satuan_barang", "")))
        nama_barang_label = _teks_aman(
            barang.get("nama_barang_label"),
            f"{nama_barang} - {satuan}" if satuan else nama_barang,
        )

        stok_saat_ini = _angka_aman(barang.get("stok_saat_ini"), 0)
        stok_minimal = _angka_aman(barang.get("stok_minimal"), 0)
        harga_beli = _angka_aman(barang.get("harga_beli"), 0)
        terjual_periode = _angka_aman(barang.get("terjual_periode"), 0)

        id_supplier = _angka_aman(barang.get("id_supplier"), 0)
        nama_supplier = _teks_aman(barang.get("nama_supplier"))
        nama_perusahaan = _teks_aman(barang.get("nama_perusahaan"))
        supplier_label = _teks_aman(barang.get("supplier_label"))

        if harga_beli <= 0:
            continue

        kebutuhan_info = _hitung_kebutuhan_hitung_manual(
            terjual_periode=terjual_periode,
            stok_saat_ini=stok_saat_ini,
            jumlah_hari=jumlah_hari_periode,
            stok_minimal=stok_minimal,
        )

        kebutuhan_total = _angka_aman(kebutuhan_info.get("kebutuhan_total"), 0)
        kebutuhan_stok_minimal = _angka_aman(kebutuhan_info.get("kebutuhan_stok_minimal"), 0)
        target_stok_aman = _angka_aman(kebutuhan_info.get("target_stok_aman"), 0)
        proyeksi_30_hari = _angka_aman(kebutuhan_info.get("proyeksi_30_hari"), 0)
        target_fuzzy = _angka_aman(kebutuhan_info.get("target_fuzzy"), kebutuhan_total)

        # Kalau stok masih aman, jangan ditampilkan di rekomendasi.
        if kebutuhan_total <= 0:
            continue

        alpha_penjualan = _membership_variabel(terjual_periode, domain_penjualan)
        alpha_stok = _membership_variabel(stok_saat_ini, domain_stok)

        target_maks = max(1, target_fuzzy)

        aturan = _aturan_fuzzy(
                alpha_penjualan,
                alpha_stok,
                target_maks,
            )

        fuzzy_item = _evaluasi_fuzzy_item(
                {
                    "target_fuzzy": target_maks,
                    "target_stok_aman": target_stok_aman,
                    "kebutuhan_total": kebutuhan_total,
                    "kebutuhan_stok_minimal": kebutuhan_stok_minimal,
                },
                aturan,
            )

        hasil_mamdani = max(0, _angka_aman(fuzzy_item.get("hasil_mamdani"), 0))
        hasil_tsukamoto = max(0, _angka_aman(fuzzy_item.get("hasil_tsukamoto"), 0))
        hasil_sugeno = max(0, _angka_aman(fuzzy_item.get("hasil_sugeno"), 0))

        error_mamdani = _angka_aman(fuzzy_item.get("error_mamdani"), 0)
        error_tsukamoto = _angka_aman(fuzzy_item.get("error_tsukamoto"), 0)
        error_sugeno = _angka_aman(fuzzy_item.get("error_sugeno"), 0)

        total_error_mamdani += error_mamdani
        total_error_tsukamoto += error_tsukamoto
        total_error_sugeno += error_sugeno

        metode_terbaik_item = _teks_aman(
                fuzzy_item.get("metode_terbaik_item"),
                "setara",
            )
        status_metode = _teks_aman(
                fuzzy_item.get("status_metode"),
                "setara",
            )
        kandidat_terbaik = fuzzy_item.get("kandidat_terbaik") or []

        rekomendasi_fuzzy = max(
                0,
                _angka_aman(fuzzy_item.get("rekomendasi_fuzzy"), 0),
            )

        nilai_rekomendasi_fuzzy = rekomendasi_fuzzy

        rekomendasi_beli = max(
            0,
            _angka_aman(fuzzy_item.get("rekomendasi_beli"), 0),
        )

        dasar_rekomendasi = _teks_aman(
            fuzzy_item.get("dasar_rekomendasi"),
            "hasil_metode_fuzzy_terbaik",
        )

        rumus_rekomendasi_fuzzy = _teks_aman(
                fuzzy_item.get("rumus_rekomendasi_fuzzy"),
                "",
            )

        rumus_mamdani = _teks_aman(fuzzy_item.get("rumus_mamdani"), "")
        rumus_tsukamoto = _teks_aman(fuzzy_item.get("rumus_tsukamoto"), "")
        rumus_sugeno = _teks_aman(fuzzy_item.get("rumus_sugeno"), "")
        rumus_metode_terbaik = _teks_aman(fuzzy_item.get("rumus_metode_terbaik"), "")

        item = {
            "id_barang": id_barang,
            "kode": kode_barang,
            "kode_barang": kode_barang,

            "nama": nama_barang,
            "nama_barang": nama_barang,
            "nama_barang_label": nama_barang_label,
            "jenis_barang": jenis_barang_item,
            "satuan": satuan,
            "satuan_barang": satuan,

            "id_supplier": int(id_supplier or 0),
            "supplier_id": int(id_supplier or 0),
            "nama_supplier": nama_supplier,
            "nama_perusahaan": nama_perusahaan,
            "supplier_label": supplier_label,

            "stok_saat_ini": stok_saat_ini,
            "stok_minimal": stok_minimal,
            "terjual_periode": terjual_periode,
            "harga_beli": harga_beli,

            "rata_harian": kebutuhan_info.get("rata_harian", 0),
            "proyeksi_30_hari": proyeksi_30_hari,
            "target_stok_aman": target_stok_aman,
            "target_fuzzy": target_maks,
            "kebutuhan_stok_minimal": kebutuhan_stok_minimal,
            "kebutuhan_barang": kebutuhan_total,
            "kebutuhan_total": kebutuhan_total,

            "hasil_mamdani": hasil_mamdani,
            "hasil_tsukamoto": hasil_tsukamoto,
            "hasil_sugeno": hasil_sugeno,

            "qty_mamdani": hasil_mamdani,
            "qty_tsukamoto": hasil_tsukamoto,
            "qty_sugeno": hasil_sugeno,

            "rekomendasi_fuzzy": rekomendasi_fuzzy,
            "nilai_rekomendasi_fuzzy": nilai_rekomendasi_fuzzy,
            "dasar_rekomendasi": dasar_rekomendasi,

            "rumus_mamdani": rumus_mamdani,
            "rumus_tsukamoto": rumus_tsukamoto,
            "rumus_sugeno": rumus_sugeno,
            "rumus_metode_terbaik": rumus_metode_terbaik,
            "rumus_rekomendasi_fuzzy": rumus_rekomendasi_fuzzy,

            "error_mamdani": error_mamdani,
            "error_tsukamoto": error_tsukamoto,
            "error_sugeno": error_sugeno,

            "status_metode": status_metode,
            "metode_terbaik_item": metode_terbaik_item,
            "kandidat_terbaik": kandidat_terbaik,
            "metode_terpilih_label": _label_metode(status_metode, metode_terbaik_item, kandidat_terbaik),

            "rekomendasi_beli": rekomendasi_beli,
            "rekomendasi_ideal": rekomendasi_beli,
            "target_restock": rekomendasi_beli,

            "qty_final_beli": 0,
            "dana_final_beli": 0,

            "aturan": aturan,
            "membership_penjualan": alpha_penjualan,
            "membership_stok": alpha_stok,
        }

        item.update(_hitung_qty_dan_dana_per_metode(item))

        skor, penjelasan_prioritas = _prioritas_dari_item(item)
        item["prioritas"] = skor
        item["label_prioritas"] = _label_prioritas_item(skor)
        item["penjelasan_prioritas"] = penjelasan_prioritas

        hasil_items.append(item)

    if not hasil_items:
        return JsonResponse({
            "ok": False,
            "msg": "Tidak ada barang yang perlu direstock pada periode ini."
        }, status=404)

    hasil_items = sorted(
        hasil_items,
        key=lambda x: (
            -_angka_aman(x.get("prioritas"), 0),
            -_angka_aman(x.get("rekomendasi_beli"), 0),
            -_angka_aman(x.get("terjual_periode"), 0),
            x.get("kode_barang", ""),
        ),
    )

    total_dana_digunakan, sisa_dana = _alokasi_budget_final(hasil_items, dana_maksimal)

    # Hitung ulang prioritas setelah qty_final_beli terisi.
    for item in hasil_items:
        skor, penjelasan_prioritas = _prioritas_dari_item(item)
        item["prioritas"] = skor
        item["label_prioritas"] = _label_prioritas_item(skor)
        item["penjelasan_prioritas"] = penjelasan_prioritas

    hasil_items = sorted(
        hasil_items,
        key=lambda x: (
            -_angka_aman(x.get("prioritas"), 0),
            -_angka_aman(x.get("qty_final_beli"), 0),
            -_angka_aman(x.get("rekomendasi_beli"), 0),
            -_angka_aman(x.get("terjual_periode"), 0),
            x.get("kode_barang", ""),
        ),
    )

    for item in hasil_items:
        rekomendasi_beli = _angka_aman(item.get("rekomendasi_beli"), 0)
        qty_final_beli = _angka_aman(item.get("qty_final_beli"), 0)
        dana_final_beli = _angka_aman(item.get("dana_final_beli"), 0)

        status_kebutuhan = _status_kebutuhan_item(
            item.get("stok_saat_ini"),
            item.get("stok_minimal"),
            item.get("terjual_periode"),
            rekomendasi_beli,
            qty_final_beli,
        )

        item["status_kebutuhan"] = status_kebutuhan
        item["status_kebutuhan_label"] = _label_status_kebutuhan(status_kebutuhan)

        item["alasan_singkat"] = _alasan_singkat_item(
            item.get("stok_saat_ini"),
            item.get("stok_minimal"),
            item.get("terjual_periode"),
            rekomendasi_beli,
            qty_final_beli,
        )

        item["penjelasan_metode_sederhana"] = _penjelasan_metode_sederhana_item(
            item.get("status_metode"),
            item.get("metode_terbaik_item"),
            item.get("kandidat_terbaik"),
            item.get("hasil_mamdani"),
            item.get("hasil_tsukamoto"),
            item.get("hasil_sugeno"),
            item.get("rekomendasi_fuzzy"),
            rekomendasi_beli,
            item.get("kebutuhan_stok_minimal"),
        )

        item["penjelasan_sederhana"] = _penjelasan_sederhana_item(
            item.get("stok_saat_ini"),
            item.get("stok_minimal"),
            item.get("terjual_periode"),
            item.get("rata_harian"),
            item.get("proyeksi_30_hari"),
            item.get("target_stok_aman"),
            item.get("kebutuhan_total"),
            rekomendasi_beli,
            qty_final_beli,
            dana_final_beli,
        )

        item["info_alokasi"] = list(item.get("rumus_alokasi") or item.get("info_alokasi") or [])

    evaluasi_global = _evaluasi_metode_global(
        total_error_mamdani=total_error_mamdani,
        total_error_tsukamoto=total_error_tsukamoto,
        total_error_sugeno=total_error_sugeno,
    )

    metode_terbaik_global = evaluasi_global["metode_terbaik_global"]
    status_global = evaluasi_global["status_global"]
    kandidat_global = evaluasi_global["kandidat_global"]

    metode_terbaik_global_label = _label_metode(
        status_global,
        metode_terbaik_global,
        kandidat_global,
    )

    if status_global == "setara":
        keterangan_global = (
            f"Evaluasi global menunjukkan hasil terbaik yang setara. "
            f"Total error Mamdani = {total_error_mamdani}, "
            f"Tsukamoto = {total_error_tsukamoto}, "
            f"Sugeno = {total_error_sugeno}."
        )
    else:
        keterangan_global = (
            f"Metode terbaik global adalah {metode_terbaik_global_label} karena memiliki total error paling kecil. "
            f"Mamdani = {total_error_mamdani}, Tsukamoto = {total_error_tsukamoto}, Sugeno = {total_error_sugeno}."
        )

    total_hasil_mamdani = sum(_angka_aman(x.get("hasil_mamdani"), 0) for x in hasil_items)
    total_hasil_tsukamoto = sum(_angka_aman(x.get("hasil_tsukamoto"), 0) for x in hasil_items)
    total_hasil_sugeno = sum(_angka_aman(x.get("hasil_sugeno"), 0) for x in hasil_items)
    total_rekomendasi_fuzzy = sum(_angka_aman(x.get("rekomendasi_fuzzy"), 0) for x in hasil_items)
    total_hasil_rekomendasi = sum(_angka_aman(x.get("rekomendasi_beli"), 0) for x in hasil_items)
    total_final_beli = sum(_angka_aman(x.get("qty_final_beli"), 0) for x in hasil_items)

    barang_prioritas_utama = None

    for item in hasil_items:
        if _angka_aman(item.get("qty_final_beli"), 0) > 0:
            barang_prioritas_utama = item
            break

    if not barang_prioritas_utama and hasil_items:
        barang_prioritas_utama = hasil_items[0]

    barang_prioritas_utama_label = "-"
    barang_prioritas_utama_alasan = "-"

    if barang_prioritas_utama:
        barang_prioritas_utama_label = (
            f"{_teks_aman(barang_prioritas_utama.get('kode_barang'))} - "
            f"{_teks_aman(barang_prioritas_utama.get('nama_barang_label', barang_prioritas_utama.get('nama_barang')))}"
        ).strip(" -")

        barang_prioritas_utama_alasan = _teks_aman(
            barang_prioritas_utama.get("penjelasan_prioritas"),
            barang_prioritas_utama.get("alasan_singkat", "-"),
        )

    return JsonResponse({
        "ok": True,
        "msg": "Rekomendasi berhasil diproses.",
        "id_rekomendasi": 0,
        "input": {
            "periode_label": f"{periode_mulai.strftime('%d-%m-%Y')} s/d {periode_selesai.strftime('%d-%m-%Y')}",
            "periode_mulai": str(periode_mulai),
            "periode_selesai": str(periode_selesai),
            "jenis_barang": jenis_barang,
            "dana_maksimal": dana_maksimal,
            "dana_maksimal_format": _format_rupiah_angka(dana_maksimal),
        },
        "evaluasi_metode": {
            "metode_terbaik_global": metode_terbaik_global,
            "status_global": status_global,
            "kandidat_global": kandidat_global,
            "metode_terbaik_global_label": metode_terbaik_global_label,
            "total_error_mamdani": total_error_mamdani,
            "total_error_tsukamoto": total_error_tsukamoto,
            "total_error_sugeno": total_error_sugeno,
            "keterangan_global": keterangan_global,
        },
        "ringkasan": {
            "total_hasil_mamdani": total_hasil_mamdani,
            "total_hasil_tsukamoto": total_hasil_tsukamoto,
            "total_hasil_sugeno": total_hasil_sugeno,
            "total_rekomendasi_fuzzy": total_rekomendasi_fuzzy,
            "total_hasil_rekomendasi": total_hasil_rekomendasi,

            "total_final_beli": total_final_beli,
            "total_dana_digunakan": total_dana_digunakan,
            "total_dana_digunakan_format": _format_rupiah_angka(total_dana_digunakan),
            "sisa_dana": sisa_dana,
            "sisa_dana_format": _format_rupiah_angka(sisa_dana),

            "barang_prioritas_utama_label": barang_prioritas_utama_label,
            "barang_prioritas_utama_alasan": barang_prioritas_utama_alasan,

            "barang_prioritas": {
                "kode_barang": barang_prioritas_utama.get("kode_barang", "") if barang_prioritas_utama else "",
                "nama_barang": barang_prioritas_utama.get("nama_barang_label", barang_prioritas_utama.get("nama_barang", "")) if barang_prioritas_utama else "",
                "prioritas": barang_prioritas_utama.get("prioritas", 0) if barang_prioritas_utama else 0,
                "penjelasan": barang_prioritas_utama.get("penjelasan_prioritas", "") if barang_prioritas_utama else "",
            },
        },
        "items": hasil_items,
    })

# =========================================================
# ROUTES: HOME / LOGIN
# =========================================================
def index(request):
    return redirect("kasir:home") if _is_logged_in(request) else redirect("kasir:login")


def home(request):
    gate = _require_login(request)
    if gate:
        return gate

    role = _normalize_role(request.session.get("kasir_role"))
    user_id = int(request.session.get("kasir_user_id") or 0)
    u = TbUser.objects.filter(id_user=user_id).first()

    try:
        hari_ini = timezone.localdate(_get_timezone_wib())
    except Exception:
        hari_ini = timezone.localdate()

    from datetime import timedelta
    qs_trx_hari_ini = _filter_date_range_helper(TbTransaksi.objects.all(), hari_ini, hari_ini + timedelta(days=1))
    if role != "pemilik":
        qs_trx_hari_ini = qs_trx_hari_ini.filter(id_kasir=user_id)

    transaksi_hari_ini = int(qs_trx_hari_ini.aggregate(jumlah=Count("id_transaksi")).get("jumlah") or 0)
    pendapatan_hari_ini = _format_rupiah_display(
        qs_trx_hari_ini.aggregate(total=Sum("total_harga")).get("total") or 0
    )

    stok_menipis_list = _ambil_stok_urgent_beranda(jumlah_data=10, hari_penjualan=30)
    stok_menipis_count = len(stok_menipis_list)

    qs_trx_terakhir = TbTransaksi.objects.all().order_by("-tanggal_waktu", "-id_transaksi")
    if role != "pemilik":
        qs_trx_terakhir = qs_trx_terakhir.filter(id_kasir=user_id)

    peta_user = {}
    for user_obj in TbUser.objects.all().only("id_user", "nama", "username"):
        uid = int(getattr(user_obj, "id_user", 0) or 0)
        peta_user[uid] = (
            (getattr(user_obj, "nama", "") or "").strip()
            or (getattr(user_obj, "username", "") or "").strip()
        )

    transaksi_terakhir_list = []
    for trx in qs_trx_terakhir[:5]:
        dt = getattr(trx, "tanggal_waktu", None)
        if dt:
            dt_wib = _to_wib(dt)
            try:
                jam_text = dt_wib.strftime("%H:%M")
            except Exception:
                jam_text = "-"
        else:
            jam_text = "-"

        transaksi_terakhir_list.append({
            "kode": _ambil_kode_transaksi(trx),
            "jam": jam_text,
            "kasir": _ambil_nama_kasir(trx, peta_user),
        })

    context = {
        "username": (u.username if u else request.session.get("kasir_username")),
        "nama": (u.nama if u else (request.session.get("kasir_nama") or "")),
        "role": role,
        "status": (getattr(u, "status", "") if u else ""),
        "user_photo": (u.foto.url if (u and getattr(u, "foto", None)) else None),
        "transaksi_hari_ini": transaksi_hari_ini,
        "pendapatan_hari_ini": pendapatan_hari_ini,
        "stok_menipis_count": stok_menipis_count,
        "stok_menipis_list": stok_menipis_list,
        "transaksi_terakhir_list": transaksi_terakhir_list,
    }
    return render(request, "kasir/home.html", context)


@require_http_methods(["GET", "POST"])
def login_view(request):
    if request.method == "GET":
        if _is_logged_in(request):
            return redirect("kasir:home")
        return render(request, "kasir/login.html")

    username = request.POST.get("username", "").strip()
    password = request.POST.get("password", "").strip()
    remember = request.POST.get("remember") == "on"

    if not username or not password:
        messages.error(request, "Username dan password wajib diisi.")
        return render(request, "kasir/login.html")

    user = TbUser.objects.filter(username=username).first()
    if not user:
        messages.error(request, "Username atau password salah.")
        return render(request, "kasir/login.html")

    _auto_nonaktif_kasir_jika_lama_tidak_login(user)

    status_val = _normalize_status(getattr(user, "status", "") or "")
    role_val = _normalize_role(getattr(user, "role", "") or "")

    if status_val != "aktif":
        if role_val == "kasir":
            messages.error(request, "Akun kasir tidak aktif. Hubungi pemilik.")
        else:
            messages.error(request, "Akun kamu tidak aktif. Hubungi pemilik.")
        return render(request, "kasir/login.html")

    stored = (user.password or "").strip()
    if not _MD5_RE.match(stored):
        messages.error(request, "Password di database tidak valid (bukan MD5 32 karakter).")
        return render(request, "kasir/login.html")

    if _md5(password).lower() != stored.lower():
        messages.error(request, "Username atau password salah.")
        return render(request, "kasir/login.html")

    request.session["kasir_user_id"] = getattr(user, "id_user", user.pk)
    request.session["kasir_username"] = user.username
    request.session["kasir_role"] = _normalize_role(getattr(user, "role", "") or "")
    request.session["kasir_nama"] = getattr(user, "nama", "") or ""
    request.session.set_expiry(60 * 60 * 24 * 30 if remember else 0)

    sekarang = timezone.now()
    user.last_login = sekarang
    user.updated_at = sekarang
    user.save(update_fields=["last_login", "updated_at"])

    return redirect("kasir:home")


def logout_view(request):
    request.session.flush()
    return redirect("kasir:login")


@require_http_methods(["GET", "POST"])
def lupa_password(request):
    if request.method == "GET":
        return render(request, "kasir/lupapassword.html")

    username = request.POST.get("username", "").strip()
    pw_lama = request.POST.get("password_lama", "").strip()
    pw_baru = request.POST.get("password_baru", "").strip()
    pw_konfirmasi = request.POST.get("konfirmasi_password", "").strip()

    if not username or not pw_lama or not pw_baru or not pw_konfirmasi:
        messages.error(request, "Semua field wajib diisi.")
        return render(request, "kasir/lupapassword.html")

    if pw_baru != pw_konfirmasi:
        messages.error(request, "Ulangi password baru tidak sama.")
        return render(request, "kasir/lupapassword.html")

    if len(pw_baru) < 4:
        messages.error(request, "Password baru minimal 4 karakter.")
        return render(request, "kasir/lupapassword.html")

    user = TbUser.objects.filter(username=username).first()
    if not user:
        messages.error(request, "Username tidak ditemukan.")
        return render(request, "kasir/lupapassword.html")

    stored = (user.password or "").strip()
    if not _MD5_RE.match(stored):
        messages.error(request, "Password di database tidak valid.")
        return render(request, "kasir/lupapassword.html")

    if _md5(pw_lama).lower() != stored.lower():
        messages.error(request, "Password lama salah.")
        return render(request, "kasir/lupapassword.html")

    user.password = _md5(pw_baru)
    user.updated_at = timezone.now()
    user.save(update_fields=["password", "updated_at"])

    messages.success(request, "Password berhasil diubah. Silakan login.")
    return redirect("kasir:login")


@require_http_methods(["POST"])
def akun_edit(request):
    gate = _require_login(request)
    if gate:
        return gate

    user_id = request.session.get("kasir_user_id")
    user = TbUser.objects.filter(id_user=user_id).first()
    if not user:
        messages.error(request, "User tidak ditemukan.")
        return redirect("kasir:home")

    nama_baru = request.POST.get("nama", "").strip()
    username_baru = request.POST.get("username", "").strip()

    if not username_baru:
        messages.error(request, "Username wajib diisi.")
        return redirect("kasir:home")

    exists = TbUser.objects.filter(username=username_baru).exclude(id_user=user.id_user).exists()
    if exists:
        messages.error(request, "Username sudah dipakai.")
        return redirect("kasir:home")

    user.nama = nama_baru
    user.username = username_baru

    hapus_foto = request.POST.get("hapus_foto", "0") == "1"
    if hapus_foto:
        if getattr(user, "foto", None):
            try:
                user.foto.delete(save=False)
            except Exception:
                pass
        user.foto = None

    foto = request.FILES.get("foto")
    if foto:
        user.foto = foto

    pw_baru = (request.POST.get("password_baru") or "").strip()
    pw_baru2 = (request.POST.get("password_baru2") or "").strip()
    if pw_baru or pw_baru2:
        if len(pw_baru) < 4:
            messages.error(request, "Password baru minimal 4 karakter.")
            return redirect("kasir:home")
        if pw_baru != pw_baru2:
            messages.error(request, "Konfirmasi password baru tidak sama.")
            return redirect("kasir:home")
        user.password = _md5(pw_baru)

    user.updated_at = timezone.now()
    user.save()

    nama_tampil_final = (user.nama or "").strip() or (user.username or "").strip()

    _sinkronkan_nama_kasir_di_riwayat(
        id_kasir=user.id_user,
        nama_baru=nama_tampil_final,
        username_baru=(user.username or "").strip(),
    )

    request.session["kasir_username"] = user.username
    request.session["kasir_role"] = (getattr(user, "role", "") or "").strip().lower()
    request.session["kasir_nama"] = user.nama or ""

    messages.success(request, "Profil berhasil diperbarui.")
    return redirect("kasir:home")


@require_http_methods(["POST"])
def akun_tambah(request):
    gate = _require_login(request)
    if gate:
        return gate

    if not _is_pemilik(request):
        messages.error(request, "Akses ditolak.")
        return redirect("kasir:home")

    nama = request.POST.get("nama", "").strip()
    username = request.POST.get("username", "").strip()
    password = request.POST.get("password", "").strip()
    password2 = request.POST.get("password2", "").strip()

    role = _normalize_role(request.POST.get("role", "kasir"))
    status = _normalize_status(request.POST.get("status", "aktif"))

    if not username or not password or not password2:
        messages.error(request, "Username dan password wajib diisi.")
        return redirect("kasir:home")

    if password != password2:
        messages.error(request, "Konfirmasi password harus sama.")
        return redirect("kasir:home")

    if len(password) < 4:
        messages.error(request, "Password minimal 4 karakter.")
        return redirect("kasir:home")

    if TbUser.objects.filter(username=username).exists():
        messages.error(request, "Username sudah dipakai.")
        return redirect("kasir:home")

    now = timezone.now()
    TbUser.objects.create(
        nama=nama,
        username=username,
        password=_md5(password),
        role=role,
        status=status,
        last_login=None,
        created_at=now,
        updated_at=now,
    )

    role_label = "Pemilik" if role == "pemilik" else "Kasir"
    status_label = "Aktif" if status == "aktif" else "Tidak Aktif"

    messages.success(
        request,
        f"Akun {role_label} berhasil ditambahkan dengan status {status_label}."
    )
    return redirect("kasir:home")


# =========================================================
# TRANSAKSI PAGE + SIMPAN
# =========================================================
def transaksi(request):
    gate = _require_login(request)
    if gate:
        return gate

    role = _normalize_role(request.session.get("kasir_role"))
    user_id = request.session.get("kasir_user_id")
    u = TbUser.objects.filter(id_user=user_id).first()

    barang_qs = TbBarang.objects.all().order_by("jenis_barang", "kode_barang")

    barang_list = []
    for b in barang_qs:
        harga_beli_val = _safe_decimal_to_int(getattr(b, "harga_beli", 0))
        harga_val = _safe_decimal_to_int(getattr(b, "harga", 0))

        barang_list.append({
            "id_barang": int(getattr(b, "id_barang", 0) or 0),
            "kategori": (getattr(b, "jenis_barang", "") or "").strip(),
            "kode": (getattr(b, "kode_barang", "") or "").strip(),
            "nama": (getattr(b, "nama_barang", "") or "").strip(),
            "berat": (getattr(b, "satuan", "") or "").strip(),
            "stok": int(getattr(b, "stok", 0) or 0),
            "harga_beli": harga_beli_val,
            "harga": harga_val,
        })

    context = {
        "role": role,
        "username": (u.username if u else request.session.get("kasir_username")),
        "nama": (u.nama if u else (request.session.get("kasir_nama") or "")),
        "tx_barang_json": json.dumps(barang_list),
        "tx_barang_create_url": "/barang/create/",
        "tx_barang_delete_all_url": "/barang/delete-all/",
        "midtrans_client_key": settings.MIDTRANS_CLIENT_KEY,
    }
    return render(request, "kasir/transaksi.html", context)


@require_http_methods(["POST"])
def transaksi_simpan(request):
    gate = _require_login(request)
    if gate:
        return JsonResponse({"ok": False, "msg": "Unauthorized"}, status=401)

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"ok": False, "msg": "Payload JSON tidak valid."}, status=400)

    no_jual_payload = str(
        payload.get("no_jual")
        or payload.get("noJual")
        or payload.get("nomor_jual")
        or payload.get("nomorJual")
        or ""
    ).strip()

    if not no_jual_payload:
        no_jual_payload = _generate_no_jual_otomatis()

    payload["no_jual"] = no_jual_payload

    items = payload.get("items") or []
    if not items:
        return JsonResponse({"ok": False, "msg": "Item transaksi kosong."}, status=400)

    user_id = int(request.session.get("kasir_user_id") or 0)
    user = TbUser.objects.filter(id_user=user_id).first()

    if not user:
        return JsonResponse({"ok": False, "msg": "User kasir tidak ditemukan."}, status=404)

    kode_list = [
        str(it.get("kode", "")).strip()
        for it in items
        if str(it.get("kode", "")).strip()
    ]

    if not kode_list:
        return JsonResponse({"ok": False, "msg": "Kode barang kosong."}, status=400)

    now = timezone.now()
    tanggal_waktu_transaksi = _build_tanggal_waktu_from_payload(payload)

    nama_kasir_value = (
        (getattr(user, "nama", "") or "").strip()
        or (request.session.get("kasir_username") or "").strip()
    )

    trx = None

    try:
        with db_transaction.atomic():
            qs = TbBarang.objects.select_for_update().filter(kode_barang__in=kode_list)
            barang_map = {
                (b.kode_barang or "").strip(): b
                for b in qs
            }

            total_item = 0
            total_harga = 0

            # VALIDASI DULU
            for it in items:
                kode = str(it.get("kode", "")).strip()
                qty = _to_int(it.get("qty"), 0)

                if not kode or qty <= 0:
                    return JsonResponse({
                        "ok": False,
                        "msg": "Item tidak valid. Kode barang atau jumlah kosong."
                    }, status=400)

                barang = barang_map.get(kode)

                if not barang:
                    return JsonResponse({
                        "ok": False,
                        "msg": f"Barang {kode} tidak ditemukan."
                    }, status=404)

                stok_now = int(getattr(barang, "stok", 0) or 0)

                if qty > stok_now:
                    return JsonResponse({
                        "ok": False,
                        "msg": f"Stok {kode} tidak cukup. Sisa {stok_now}."
                    }, status=400)

                harga_jual = _normalize_nominal_penuh(
                    _safe_decimal_to_int(getattr(barang, "harga", 0))
                )

                subtotal = harga_jual * qty

                total_item += qty
                total_harga += subtotal

            # SIMPAN TRANSAKSI UTAMA
            trx_kwargs = _build_transaksi_create_kwargs(
                request=request,
                payload=payload,
                total_item=total_item,
                total_harga=total_harga,
                tanggal_waktu_transaksi=tanggal_waktu_transaksi,
            )

            trx = TbTransaksi.objects.create(**trx_kwargs)

            sumber_penjualan = str(
                payload.get("sumber_penjualan")
                or payload.get("sumberPenjualan")
                or "offline"
            ).strip().lower()

            if sumber_penjualan not in ("offline", "online"):
                sumber_penjualan = "offline"

            metode_bayar = str(
                payload.get("metode_bayar")
                or payload.get("metodeBayar")
                or payload.get("metode")
                or "tunai"
            ).strip().lower()

            metode_bayar = metode_bayar.replace("-", "_").replace(" ", "_")

            marketplace = str(payload.get("marketplace") or "").strip().lower()

            if sumber_penjualan == "online":
                metode_bayar = "marketplace"

                if marketplace not in ("shopee", "tiktok", "tokopedia", "lazada"):
                    marketplace = ""

            else:
                sumber_penjualan = "offline"
                marketplace = ""

                if metode_bayar not in ("tunai", "qris"):
                    metode_bayar = "tunai"

            if sumber_penjualan == "online":
                nominal_bayar = 0
                kembalian = 0

            elif metode_bayar == "qris":
                nominal_bayar = int(total_harga or 0)
                kembalian = 0

            else:
                nominal_bayar = _normalize_nominal_penuh(
                    _to_int(payload.get("bayar", payload.get("nominal_bayar", 0)), 0)
                )

                kembalian = _normalize_nominal_penuh(
                    _to_int(payload.get("kembali", payload.get("kembalian", 0)), 0)
                )

                if nominal_bayar > 0 and nominal_bayar >= total_harga:
                    kembalian = nominal_bayar - total_harga

            _simpan_ke_kolom_trx_opsional_text(
                trx,
                ["no_jual", "no_transaksi", "nomor_jual", "kode_transaksi"],
                no_jual_payload,
            )

            _simpan_ke_kolom_trx_opsional_text(
                trx,
                ["nama_kasir", "kasir_nama", "nama_user", "nama"],
                nama_kasir_value,
            )

            _simpan_ke_kolom_trx_opsional_text(
                trx,
                ["sumber_penjualan", "jenis_transaksi", "tipe_transaksi"],
                sumber_penjualan,
            )

            _simpan_ke_kolom_trx_opsional_text(
                trx,
                ["metode_bayar", "metode_pembayaran", "jenis_bayar", "payment_method", "cara_bayar"],
                metode_bayar,
            )

            _simpan_ke_kolom_trx_opsional_text(
                trx,
                ["marketplace", "nama_marketplace", "platform_marketplace", "platform"],
                marketplace,
            )

            _simpan_ke_kolom_trx_opsional(
                trx,
                ["nominal_bayar", "uang_bayar", "bayar", "jumlah_bayar"],
                nominal_bayar,
            )

            _simpan_ke_kolom_trx_opsional(
                trx,
                ["kembalian", "uang_kembali", "total_kembalian", "kembali"],
                kembalian,
            )

            # SIMPAN DETAIL + KURANGI STOK
            for it in items:
                kode = str(it.get("kode", "")).strip()
                qty = _to_int(it.get("qty"), 0)
                barang = barang_map[kode]

                kode_snapshot = (getattr(barang, "kode_barang", "") or "").strip()
                nama_snapshot = (getattr(barang, "nama_barang", "") or "").strip()
                jenis_snapshot = (getattr(barang, "jenis_barang", "") or "").strip()
                satuan_snapshot = (getattr(barang, "satuan", "") or "").strip()

                harga_beli_snapshot = _normalize_nominal_penuh(
                    _safe_decimal_to_int(getattr(barang, "harga_beli", 0))
                )

                harga_satuan = _normalize_nominal_penuh(
                    _safe_decimal_to_int(
                        it.get("harga_satuan")
                        or it.get("harga")
                        or it.get("harga_jual")
                        or getattr(barang, "harga", 0)
                    )
                )

                if harga_satuan <= 0:
                    harga_satuan = _normalize_nominal_penuh(
                        _safe_decimal_to_int(getattr(barang, "harga", 0))
                    )

                subtotal = harga_satuan * qty

                detail_data = {
                    "id_transaksi": int(trx.id_transaksi),
                    "id_barang": int(getattr(barang, "id_barang", 0) or 0),
                    "kode_barang": kode_snapshot,
                    "nama_barang": nama_snapshot,
                    "qty": qty,
                    "harga_satuan": harga_satuan,
                    "subtotal": subtotal,
                    "kode_barang_snapshot": kode_snapshot,
                    "nama_barang_snapshot": nama_snapshot,
                    "jenis_barang_snapshot": jenis_snapshot,
                    "satuan_snapshot": satuan_snapshot,
                    "harga_beli_snapshot": harga_beli_snapshot,
                }

                detail_obj = TbTransaksiDetail.objects.create(
                    **_filter_create_kwargs(TbTransaksiDetail, detail_data)
                )

                _simpan_ke_kolom_detail_opsional_text(
                    detail_obj,
                    ["kode_barang_snapshot", "kode_barang", "kode"],
                    kode_snapshot,
                )

                _simpan_ke_kolom_detail_opsional_text(
                    detail_obj,
                    ["nama_barang_snapshot", "nama_barang", "barang_nama", "nama"],
                    nama_snapshot,
                )

                _simpan_ke_kolom_detail_opsional_text(
                    detail_obj,
                    ["jenis_barang_snapshot"],
                    jenis_snapshot,
                )

                _simpan_ke_kolom_detail_opsional_text(
                    detail_obj,
                    ["satuan_snapshot", "satuan_barang_snapshot", "besaran_snapshot", "satuan", "berat", "ukuran"],
                    satuan_snapshot,
                )

                _simpan_ke_kolom_detail_opsional_int(
                    detail_obj,
                    ["harga_beli_snapshot"],
                    harga_beli_snapshot,
                )

                barang.stok = int(getattr(barang, "stok", 0) or 0) - qty

                update_fields = ["stok"]
                if "updated_at" in _model_field_names(TbBarang):
                    barang.updated_at = now
                    update_fields.append("updated_at")

                barang.save(update_fields=update_fields)

    except Exception as e:
        return JsonResponse({
            "ok": False,
            "msg": f"Gagal menyimpan transaksi: {e}"
        }, status=500)

    return JsonResponse({
        "ok": True,
        "msg": "Transaksi tersimpan.",
        "id_transaksi": int(trx.id_transaksi),
        "no_jual": no_jual_payload,
        "nama_kasir": nama_kasir_value,
    })

@require_http_methods(["POST"])
def transaksi_midtrans_token(request):
    gate = _require_login(request)
    if gate:
        return JsonResponse({"ok": False, "msg": "Unauthorized"}, status=401)

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"ok": False, "msg": "Payload JSON tidak valid."}, status=400)

    items = payload.get("items") or []
    if not items:
        return JsonResponse({"ok": False, "msg": "Item transaksi kosong."}, status=400)

    total = 0
    item_details = []

    for item in items:
        kode = str(item.get("kode") or "").strip()
        nama = str(item.get("nama") or item.get("nama_barang") or kode or "Barang").strip()

        try:
            qty = int(item.get("qty") or 0)
            harga = int(
                str(item.get("harga") or item.get("harga_satuan") or 0)
                .replace(".", "")
                .replace(",", "")
            )
        except Exception:
            qty = 0
            harga = 0

        if not kode or qty <= 0 or harga <= 0:
            return JsonResponse({
                "ok": False,
                "msg": "Data item pembayaran online tidak valid."
            }, status=400)

        total += qty * harga

        item_details.append({
            "id": kode[:50],
            "price": harga,
            "quantity": qty,
            "name": nama[:50],
        })

    if total <= 0:
        return JsonResponse({"ok": False, "msg": "Total transaksi tidak valid."}, status=400)

    no_jual = str(payload.get("no_jual") or payload.get("noJual") or "").strip()
    if not no_jual:
        no_jual = _generate_no_jual_otomatis()

    order_id = f"{no_jual}-{int(timezone.now().timestamp())}"

    snap = midtransclient.Snap(
        is_production=settings.MIDTRANS_IS_PRODUCTION,
        server_key=settings.MIDTRANS_SERVER_KEY,
        client_key=settings.MIDTRANS_CLIENT_KEY,
    )

    parameter = {
        "transaction_details": {
            "order_id": order_id,
            "gross_amount": total,
        },
        "item_details": item_details,
        "customer_details": {
            "first_name": request.session.get("kasir_username") or "Kasir",
        },
    }

    try:
        snap_response = snap.create_transaction(parameter)
    except Exception as e:
        return JsonResponse({
            "ok": False,
            "msg": f"Gagal membuat pembayaran Midtrans: {e}"
        }, status=500)

    return JsonResponse({
        "ok": True,
        "token": snap_response.get("token"),
        "redirect_url": snap_response.get("redirect_url"),
        "order_id": order_id,
        "no_jual": no_jual,
        "total": total,
    })

@require_http_methods(["GET"])
def barang(request):
    gate = _require_login(request)
    if gate:
        return gate

    if not _is_pemilik(request):
        messages.error(request, "Akses ditolak.")
        return redirect("kasir:home")

    user_id = int(request.session.get("kasir_user_id") or 0)
    user_login = TbUser.objects.filter(id_user=user_id).first()

    daftar_kategori = TbKategori.objects.all().order_by("nama_kategori")

    return render(request, "kasir/barang.html", {
        "username": (
            user_login.username
            if user_login
            else request.session.get("kasir_username")
        ),
        "nama": (
            user_login.nama
            if user_login
            else (request.session.get("kasir_nama") or "")
        ),
        "role": request.session.get("kasir_role", ""),
        "status": (
            getattr(user_login, "status", "")
            if user_login
            else ""
        ),
        "user_photo": (
            user_login.foto.url
            if (user_login and getattr(user_login, "foto", None))
            else None
        ),
        "daftar_kategori": daftar_kategori,
    })
# =========================================================
# BARANG JSON + CRUD + IMPORT/EXPORT
# =========================================================
@require_http_methods(["GET"])
def barang_json(request):
    gate = _require_login(request)
    if gate:
        return JsonResponse({"ok": False, "msg": "Unauthorized"}, status=401)

    qs = TbBarang.objects.all().order_by("jenis_barang", "kode_barang")

    data = []

    for b in qs:
        supplier_info = _ambil_supplier_barang_info(b)

        data.append({
            "id_barang": int(getattr(b, "id_barang", 0) or 0),
            "id_kategori": int(getattr(b, "id_kategori", 0) or 0),

            "id_supplier": int(supplier_info.get("id_supplier", 0) or 0),
            "nama_supplier": supplier_info.get("nama_supplier", ""),
            "nama_perusahaan": supplier_info.get("nama_perusahaan", ""),
            "supplier_label": supplier_info.get("supplier_label", ""),

            "kategori": (getattr(b, "jenis_barang", "") or "").strip(),
            "jenis_barang": (getattr(b, "jenis_barang", "") or "").strip(),

            "kode": (getattr(b, "kode_barang", "") or "").strip(),
            "kode_barang": (getattr(b, "kode_barang", "") or "").strip(),

            "nama": (getattr(b, "nama_barang", "") or "").strip(),
            "nama_barang": (getattr(b, "nama_barang", "") or "").strip(),

            "berat": (getattr(b, "satuan", "") or "").strip(),
            "satuan": (getattr(b, "satuan", "") or "").strip(),

            "stok": int(getattr(b, "stok", 0) or 0),
            "stok_saat_ini": int(getattr(b, "stok", 0) or 0),
            "stok_minimal": int(getattr(b, "stok_minimal", 0) or 0),

            "harga_beli": _safe_decimal_to_int(getattr(b, "harga_beli", 0)),
            "harga": _safe_decimal_to_int(getattr(b, "harga", 0)),
            "harga_jual": _safe_decimal_to_int(getattr(b, "harga", 0)),
        })

    daftar_kategori = []

    try:
        kategori_qs = TbKategori.objects.all().order_by("nama_kategori")

        for k in kategori_qs:
            nama_kategori = _normalisasi_nama_kategori(
                getattr(k, "nama_kategori", "") or ""
            )

            if not nama_kategori:
                continue

            kode_kategori = ""

            try:
                kode_kategori = str(
                    getattr(k, "kode_kategori", "") or ""
                ).strip().upper()
            except Exception:
                kode_kategori = ""

            if not kode_kategori:
                try:
                    kode_kategori = _ambil_kolom_opsional_db_text(
                        TbKategori._meta.db_table,
                        "id_kategori",
                        getattr(k, "id_kategori", 0),
                        ["kode_kategori"]
                    ).strip().upper()
                except Exception:
                    kode_kategori = ""

            if not kode_kategori:
                kode_kategori = _kode_kategori_dari_nama(nama_kategori)

            daftar_kategori.append({
                "id_kategori": int(getattr(k, "id_kategori", 0) or 0),
                "nama_kategori": nama_kategori,
                "kategori": nama_kategori,
                "kode_kategori": kode_kategori,
                "prefix": kode_kategori,
            })

    except Exception:
        daftar_kategori = []

    return JsonResponse({
        "ok": True,
        "data": data,
        "kategori": daftar_kategori,
        "daftar_kategori": daftar_kategori,
    })


def _norm_header(s: str) -> str:
    s = (s or "").strip().lower()
    s = s.replace(" ", "_").replace("-", "_")
    return s


def _map_headers(headers):
    h = [_norm_header(x) for x in headers]

    def find(*cands):
        for c in cands:
            if c in h:
                return h.index(c)
        return -1

    return {
        "kode": find("kode_barang", "kode", "kd_barang", "kdbarang"),
        "jenis": find("jenis_barang", "jenis", "kategori"),
        "nama": find("nama_barang", "nama", "barang"),
        "satuan": find("satuan", "unit", "berat"),
        "stok": find("stok_saat_ini", "stok", "stok_saatini"),
        "harga_beli": find("harga_beli", "hargabeli", "harga_modal", "hargamodal"),
        "harga": find("harga_jual", "harga", "hargajual"),
    }


def _upsert_barang(rows):
    now = timezone.now()
    inserted = 0
    updated = 0
    skipped = 0

    kode_list = [str(r.get("kode", "")).strip() for r in rows if str(r.get("kode", "")).strip()]
    existing = TbBarang.objects.filter(kode_barang__in=kode_list)
    existing_map = {(x.kode_barang or "").strip(): x for x in existing}

    for r in rows:
        kode = (str(r.get("kode", "") or "")).strip()
        if not kode:
            skipped += 1
            continue

        jenis = (str(r.get("jenis", "") or "")).strip()
        nama = (str(r.get("nama", "") or "")).strip()
        satuan = (str(r.get("satuan", "") or "")).strip()

        stok = _to_int(r.get("stok"), 0)
        harga_beli = _normalize_nominal_penuh(_to_int(r.get("harga_beli"), 0))
        harga = _normalize_nominal_penuh(_to_int(r.get("harga"), 0))

        obj = existing_map.get(kode)
        if obj:
            obj.jenis_barang = jenis
            obj.nama_barang = nama
            obj.satuan = satuan
            obj.stok = stok
            obj.harga_beli = harga_beli
            obj.harga = harga
            obj.updated_at = now
            obj.save(update_fields=[
                "jenis_barang", "nama_barang", "satuan",
                "stok", "harga_beli", "harga", "updated_at",
            ])
            updated += 1
        else:
            TbBarang.objects.create(
                kode_barang=kode,
                jenis_barang=jenis,
                nama_barang=nama,
                satuan=satuan,
                stok=stok,
                harga_beli=harga_beli,
                harga=harga,
                created_at=now,
                updated_at=now,
            )
            inserted += 1

    return inserted, updated, skipped


@require_http_methods(["POST"])
def barang_import(request):
    gate = _require_login(request)
    if gate:
        return JsonResponse({"ok": False, "msg": "Unauthorized"}, status=401)
    if not _is_pemilik(request):
        return JsonResponse({"ok": False, "msg": "Akses ditolak."}, status=403)

    f = request.FILES.get("file")
    if not f:
        return JsonResponse({"ok": False, "msg": "File tidak ditemukan."}, status=400)

    name = (f.name or "").lower().strip()
    is_csv = name.endswith(".csv")
    is_xlsx = name.endswith(".xlsx")
    if not (is_csv or is_xlsx):
        return JsonResponse({"ok": False, "msg": "Format harus .csv atau .xlsx"}, status=400)

    rows = []

    try:
        if is_csv:
            raw = f.read()
            text = raw.decode("utf-8-sig", errors="ignore")
            reader = csv.reader(io.StringIO(text))
            all_rows = list(reader)

            if not all_rows:
                return JsonResponse({"ok": False, "msg": "CSV kosong."}, status=400)

            headers = all_rows[0]
            idx = _map_headers(headers)
            must = ["kode", "jenis", "nama", "satuan", "stok", "harga_beli", "harga"]
            if any(idx[k] == -1 for k in must):
                return JsonResponse({"ok": False, "msg": "Header CSV tidak cocok."}, status=400)

            for r in all_rows[1:]:
                if not r or all(str(x or "").strip() == "" for x in r):
                    continue
                rows.append({
                    "kode": r[idx["kode"]] if idx["kode"] < len(r) else "",
                    "jenis": r[idx["jenis"]] if idx["jenis"] < len(r) else "",
                    "nama": r[idx["nama"]] if idx["nama"] < len(r) else "",
                    "satuan": r[idx["satuan"]] if idx["satuan"] < len(r) else "",
                    "stok": r[idx["stok"]] if idx["stok"] < len(r) else 0,
                    "harga_beli": r[idx["harga_beli"]] if idx["harga_beli"] < len(r) else 0,
                    "harga": r[idx["harga"]] if idx["harga"] < len(r) else 0,
                })
        else:
            wb = load_workbook(f, data_only=True)
            ws = wb.active
            data = list(ws.iter_rows(values_only=True))

            if not data:
                return JsonResponse({"ok": False, "msg": "Excel kosong."}, status=400)

            headers = [str(x or "") for x in data[0]]
            idx = _map_headers(headers)
            must = ["kode", "jenis", "nama", "satuan", "stok", "harga_beli", "harga"]
            if any(idx[k] == -1 for k in must):
                return JsonResponse({"ok": False, "msg": "Header Excel tidak cocok."}, status=400)

            for r in data[1:]:
                if not r or all(str(x or "").strip() == "" for x in r):
                    continue
                rows.append({
                    "kode": r[idx["kode"]] if idx["kode"] < len(r) else "",
                    "jenis": r[idx["jenis"]] if idx["jenis"] < len(r) else "",
                    "nama": r[idx["nama"]] if idx["nama"] < len(r) else "",
                    "satuan": r[idx["satuan"]] if idx["satuan"] < len(r) else "",
                    "stok": r[idx["stok"]] if idx["stok"] < len(r) else 0,
                    "harga_beli": r[idx["harga_beli"]] if idx["harga_beli"] < len(r) else 0,
                    "harga": r[idx["harga"]] if idx["harga"] < len(r) else 0,
                })

    except Exception as e:
        return JsonResponse({"ok": False, "msg": f"Gagal baca file: {e}"}, status=400)

    if not rows:
        return JsonResponse({"ok": False, "msg": "Tidak ada data yang bisa diimport."}, status=400)

    with db_transaction.atomic():
        inserted, updated, skipped = _upsert_barang(rows)

    return JsonResponse({
        "ok": True,
        "msg": f"Import berhasil. Insert: {inserted}, Update: {updated}, Skip: {skipped}",
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
    })


@require_http_methods(["GET"])
def barang_export(request):
    gate = _require_login(request)
    if gate:
        return JsonResponse({"ok": False, "msg": "Unauthorized"}, status=401)
    if not _is_pemilik(request):
        return JsonResponse({"ok": False, "msg": "Akses ditolak."}, status=403)

    fmt = (request.GET.get("format") or "xlsx").lower().strip()
    qs = TbBarang.objects.all().order_by("jenis_barang", "kode_barang")

    header_keys = [
        "kode_barang",
        "jenis_barang",
        "nama_barang",
        "satuan",
        "stok_saat_ini",
        "harga_beli",
        "harga_jual",
    ]

    header_labels = [
        "Kode Barang",
        "Jenis Barang",
        "Nama Barang",
        "Satuan",
        "Stok Saat Ini",
        "Harga Beli",
        "Harga Jual",
    ]

    data_rows = []
    for b in qs:
        data_rows.append([
            (b.kode_barang or "").strip(),
            (b.jenis_barang or "").strip(),
            (b.nama_barang or "").strip(),
            (b.satuan or "").strip(),
            int(getattr(b, "stok", 0) or 0),
            _safe_decimal_to_int(getattr(b, "harga_beli", 0)),
            _safe_decimal_to_int(getattr(b, "harga", 0)),
        ])

    if fmt == "csv":
        output = io.StringIO()
        w = csv.writer(output)
        w.writerow(header_keys)
        for row in data_rows:
            w.writerow(row)

        resp = HttpResponse(output.getvalue(), content_type="text/csv; charset=utf-8")
        resp["Content-Disposition"] = 'attachment; filename="tb_barang_export.csv"'
        return resp

    wb = Workbook()
    ws = wb.active
    ws.title = "tb_barang"

    thin_side = Side(style="thin", color="000000")
    medium_side = Side(style="medium", color="000000")

    thin_border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )

    header_border = Border(
        left=medium_side,
        right=medium_side,
        top=medium_side,
        bottom=medium_side,
    )

    header_fill = PatternFill("solid", fgColor="D9EAD3")
    header_font = Font(bold=True, color="000000")
    body_font = Font(bold=False, color="000000")

    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    text_alignment = Alignment(horizontal="left", vertical="center")
    number_alignment = Alignment(horizontal="right", vertical="center")

    for col_idx, label in enumerate(header_labels, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = header_alignment

    for row_idx, row_data in enumerate(data_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.border = thin_border

            if col_idx in (5, 6, 7):
                cell.alignment = number_alignment
                cell.number_format = '#,##0'
            else:
                cell.alignment = text_alignment

    for col_idx in range(1, len(header_labels) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0

        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col_idx).value
            cell_text = "" if cell_value is None else str(cell_value)
            if len(cell_text) > max_length:
                max_length = len(cell_text)

        adjusted_width = max_length + 3
        if adjusted_width < 12:
            adjusted_width = 12
        if adjusted_width > 35:
            adjusted_width = 35

        ws.column_dimensions[col_letter].width = adjusted_width

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    resp = HttpResponse(
        out.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = 'attachment; filename="tb_barang_export.xlsx"'
    return resp

def _ambil_nama_kategori_dari_id(id_kategori):
    try:
        id_kategori = int(id_kategori or 0)
    except Exception:
        id_kategori = 0

    if id_kategori <= 0:
        return ""

    obj = TbKategori.objects.filter(id_kategori=id_kategori).first()
    if not obj:
        return ""

    return _teks_aman(getattr(obj, "nama_kategori", ""))


def _normalisasi_data_barang_dari_payload(payload, barang_lama=None):
    barang_lama = barang_lama or None

    kode_barang = _teks_aman(payload.get("kode_barang", getattr(barang_lama, "kode_barang", "")))
    nama_barang = _teks_aman(payload.get("nama_barang", getattr(barang_lama, "nama_barang", "")))
    satuan = _teks_aman(payload.get("satuan", getattr(barang_lama, "satuan", "")))

    id_kategori = _to_int(payload.get("id_kategori", getattr(barang_lama, "id_kategori", 0)), 0)
    nama_kategori = _ambil_nama_kategori_dari_id(id_kategori)

    # jenis_barang tetap dipertahankan untuk kompatibilitas sistem lama
    jenis_barang_input = _teks_aman(payload.get("jenis_barang", ""))
    jenis_barang_lama = _teks_aman(getattr(barang_lama, "jenis_barang", ""))
    jenis_barang = jenis_barang_input or nama_kategori or jenis_barang_lama

    stok = _to_int(payload.get("stok_saat_ini", payload.get("stok", getattr(barang_lama, "stok", 0))), 0)
    stok_minimal = _to_int(payload.get("stok_minimal", getattr(barang_lama, "stok_minimal", 0)), 0)

    harga_beli = _normalize_nominal_penuh(
        _to_int(payload.get("harga_beli", getattr(barang_lama, "harga_beli", 0)), 0)
    )
    harga = _normalize_nominal_penuh(
        _to_int(payload.get("harga_jual", payload.get("harga", getattr(barang_lama, "harga", 0))), 0)
    )

    id_supplier = _to_int(
    payload.get("id_supplier")
    or payload.get("supplier_id")
    or payload.get("supplier"),
    0
)

    return {
        "kode_barang": kode_barang,
        "nama_barang": nama_barang,
        "satuan": satuan,
        "id_kategori": id_kategori if id_kategori > 0 else None,
        "id_supplier": id_supplier if id_supplier > 0 else None,
        "jenis_barang": jenis_barang,
        "stok": stok,
        "stok_minimal": stok_minimal,
        "harga_beli": harga_beli,
        "harga": harga,
    }

@require_http_methods(["POST"])
def barang_create(request):
    gate = _require_login(request)
    if gate:
        return JsonResponse({"ok": False, "msg": "Unauthorized"}, status=401)
    if not _is_pemilik(request):
        return JsonResponse({"ok": False, "msg": "Akses ditolak."}, status=403)

    try:
        payload = json.loads(request.body.decode("utf-8")) if (
            request.content_type and "application/json" in request.content_type
        ) else request.POST
    except Exception:
        return JsonResponse({"ok": False, "msg": "Payload tidak valid."}, status=400)

    data_barang = _normalisasi_data_barang_dari_payload(payload)

    if not data_barang["kode_barang"]:
        return JsonResponse({"ok": False, "msg": "Kode barang wajib diisi."}, status=400)
    if not data_barang["nama_barang"]:
        return JsonResponse({"ok": False, "msg": "Nama barang wajib diisi."}, status=400)

    if TbBarang.objects.filter(kode_barang=data_barang["kode_barang"]).exists():
        return JsonResponse({"ok": False, "msg": "Kode barang sudah dipakai."}, status=400)

    now = timezone.now()

    barang = TbBarang.objects.create(
        kode_barang=data_barang["kode_barang"],
        id_kategori=data_barang["id_kategori"],
        id_supplier=data_barang["id_supplier"],
        jenis_barang=data_barang["jenis_barang"],
        nama_barang=data_barang["nama_barang"],
        satuan=data_barang["satuan"],
        stok=data_barang["stok"],
        stok_minimal=data_barang["stok_minimal"],
        harga_beli=data_barang["harga_beli"],
        harga=data_barang["harga"],
        created_at=now,
        updated_at=now,
    )

    return JsonResponse({
        "ok": True,
        "msg": "Barang berhasil ditambahkan.",
        "data": {
            "id_barang": int(barang.id_barang),
        }
    })


@require_http_methods(["POST"])
def barang_edit(request):
    gate = _require_login(request)
    if gate:
        return JsonResponse({"ok": False, "msg": "Unauthorized"}, status=401)
    if not _is_pemilik(request):
        return JsonResponse({"ok": False, "msg": "Akses ditolak."}, status=403)

    try:
        payload = json.loads(request.body.decode("utf-8")) if (
            request.content_type and "application/json" in request.content_type
        ) else request.POST
    except Exception:
        return JsonResponse({"ok": False, "msg": "Payload tidak valid."}, status=400)

    id_barang = _to_int(payload.get("id_barang"), 0)
    barang = TbBarang.objects.filter(id_barang=id_barang).first()
    if not barang:
        return JsonResponse({"ok": False, "msg": "Barang tidak ditemukan."}, status=404)
    

    data_barang = _normalisasi_data_barang_dari_payload(payload, barang_lama=barang)

    kode_exists = (
        TbBarang.objects
        .filter(kode_barang=data_barang["kode_barang"])
        .exclude(id_barang=id_barang)
        .exists()
    )
    if kode_exists:
        return JsonResponse({"ok": False, "msg": "Kode barang sudah dipakai."}, status=400)

    barang.kode_barang = data_barang["kode_barang"]
    barang.id_kategori = data_barang["id_kategori"]
    barang.id_supplier = data_barang["id_supplier"]
    barang.jenis_barang = data_barang["jenis_barang"]
    barang.nama_barang = data_barang["nama_barang"]
    barang.satuan = data_barang["satuan"]
    barang.stok = data_barang["stok"]
    barang.stok_minimal = data_barang["stok_minimal"]
    barang.harga_beli = data_barang["harga_beli"]
    barang.harga = data_barang["harga"]
    barang.updated_at = timezone.now()
    barang.save()

    return JsonResponse({"ok": True, "msg": "Barang berhasil diupdate."})


@require_http_methods(["POST", "DELETE"])
def barang_delete(request):
    gate = _require_login(request)
    if gate:
        return JsonResponse({"ok": False, "msg": "Unauthorized"}, status=401)
    if not _is_pemilik(request):
        return JsonResponse({"ok": False, "msg": "Akses ditolak."}, status=403)

    try:
        payload = json.loads(request.body.decode("utf-8")) if (
            request.content_type and "application/json" in request.content_type
        ) else request.POST
    except Exception:
        payload = {}

    id_barang = _to_int(payload.get("id_barang"), 0)
    barang = TbBarang.objects.filter(id_barang=id_barang).first()
    if not barang:
        return JsonResponse({"ok": False, "msg": "Barang tidak ditemukan."}, status=404)

    kode_snapshot = (getattr(barang, "kode_barang", "") or "").strip()
    nama_snapshot = (getattr(barang, "nama_barang", "") or "").strip()
    jenis_snapshot = (getattr(barang, "jenis_barang", "") or "").strip()

    try:
        with db_transaction.atomic():
            detail_qs = TbTransaksiDetail.objects.select_for_update().filter(id_barang=id_barang)
            detail_list = list(detail_qs)

            for d in detail_list:
                _simpan_ke_kolom_detail_opsional_text(
                    d,
                    ["kode_barang_snapshot", "kode_barang", "kode"],
                    kode_snapshot,
                )
                _simpan_ke_kolom_detail_opsional_text(
                    d,
                    ["nama_barang_snapshot", "nama_barang", "barang_nama", "nama"],
                    nama_snapshot,
                )
                _simpan_ke_kolom_detail_opsional_text(
                    d,
                    ["jenis_barang_snapshot"],
                    jenis_snapshot,
                )

            if detail_list:
                detail_qs.update(id_barang=None)

            barang.delete()

    except Exception as e:
        return JsonResponse({"ok": False, "msg": f"Gagal hapus barang: {e}"}, status=500)

    return JsonResponse({"ok": True, "msg": "Barang berhasil dihapus. Riwayat transaksi tetap aman."})

@require_http_methods(["POST", "DELETE"])
def barang_delete_all(request):
    gate = _require_login(request)
    if gate:
        return JsonResponse({"ok": False, "msg": "Unauthorized"}, status=401)
    if not _is_pemilik(request):
        return JsonResponse({"ok": False, "msg": "Akses ditolak."}, status=403)

    try:
        with db_transaction.atomic():
            semua_barang = list(
                TbBarang.objects.all().only("id_barang", "kode_barang", "nama_barang", "jenis_barang")
            )

            if not semua_barang:
                return JsonResponse({"ok": True, "msg": "Data barang sudah kosong."})

            barang_map = {
                int(getattr(b, "id_barang", 0) or 0): {
                    "kode": (getattr(b, "kode_barang", "") or "").strip(),
                    "nama": (getattr(b, "nama_barang", "") or "").strip(),
                    "jenis": (getattr(b, "jenis_barang", "") or "").strip(),
                }
                for b in semua_barang
                if int(getattr(b, "id_barang", 0) or 0) > 0
            }

            if barang_map:
                detail_qs = TbTransaksiDetail.objects.select_for_update().filter(
                    id_barang__in=list(barang_map.keys())
                )
                detail_list = list(detail_qs)

                for d in detail_list:
                    did = _safe_decimal_to_int(getattr(d, "id_barang", 0))
                    snap = barang_map.get(did, {"kode": "", "nama": "", "jenis": ""})

                    _simpan_ke_kolom_detail_opsional_text(
                        d,
                        ["kode_barang_snapshot", "kode_barang", "kode"],
                        snap["kode"],
                    )
                    _simpan_ke_kolom_detail_opsional_text(
                        d,
                        ["nama_barang_snapshot", "nama_barang", "barang_nama", "nama"],
                        snap["nama"],
                    )
                    _simpan_ke_kolom_detail_opsional_text(
                        d,
                        ["jenis_barang_snapshot"],
                        snap["jenis"],
                    )

                if detail_list:
                    detail_qs.update(id_barang=None)

            deleted_count, _ = TbBarang.objects.all().delete()

    except Exception as e:
        return JsonResponse({"ok": False, "msg": f"Gagal menghapus semua barang: {e}"}, status=500)

    return JsonResponse({
        "ok": True,
        "msg": f"Semua barang berhasil dihapus. Total: {deleted_count}. Riwayat transaksi tetap aman."
    })


# =========================================================
# MASTER DATA - KATEGORI / JENIS BARANG
# =========================================================

def _normalisasi_nama_kategori(value):
    nama = str(value or "").strip().upper()
    nama = re.sub(r"\s+", " ", nama)
    return nama


def _prefix_kategori(nama_kategori):
    nama = _normalisasi_nama_kategori(nama_kategori)

    prefix_map = {
        "FUNGISIDA": "FNG",
        "INSEKTISIDA": "INS",
        "HERBISIDA": "HRB",
        "NUTRISI DAN ZPT": "NUT",
        "PUPUK": "PPK",
    }

    if nama in prefix_map:
        return prefix_map[nama]

    huruf = re.sub(r"[^A-Z0-9]", "", nama)
    if len(huruf) >= 3:
        return huruf[:3]

    return (huruf or "KTG").ljust(3, "X")[:3]

def _kode_kategori_dari_nama(nama_kategori):
    nama = _normalisasi_nama_kategori(nama_kategori)

    prefix_map = {
        "FUNGISIDA": "FNG",
        "INSEKTISIDA": "INS",
        "HERBISIDA": "HRB",
        "NUTRISI DAN ZPT": "NUT",
        "PUPUK": "PPK",
    }

    if nama in prefix_map:
        return prefix_map[nama]

    huruf = re.sub(r"[^A-Z0-9]", "", nama)
    if len(huruf) >= 3:
        return huruf[:3]

    return (huruf or "KTG").ljust(3, "X")[:3]


def _ambil_kode_kategori_obj(kategori):
    kode = ""

    try:
        kode = str(getattr(kategori, "kode_kategori", "") or "").strip().upper()
    except Exception:
        kode = ""

    if kode:
        return kode

    try:
        kode = _ambil_kolom_opsional_db_text(
            TbKategori._meta.db_table,
            "id_kategori",
            getattr(kategori, "id_kategori", 0),
            ["kode_kategori"]
        )
        kode = str(kode or "").strip().upper()
    except Exception:
        kode = ""

    if kode:
        return kode

    return _kode_kategori_dari_nama(getattr(kategori, "nama_kategori", ""))


def _simpan_kode_kategori_db(id_kategori, kode_kategori):
    try:
        id_kategori = int(id_kategori or 0)
    except Exception:
        id_kategori = 0

    kode_kategori = str(kode_kategori or "").strip().upper()

    if id_kategori <= 0 or not kode_kategori:
        return

    try:
        if _db_has_column(TbKategori._meta.db_table, "kode_kategori"):
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    UPDATE tb_kategori
                    SET kode_kategori = %s
                    WHERE id_kategori = %s
                    """,
                    [kode_kategori, id_kategori]
                )
    except Exception:
        pass


def _buat_kategori_dengan_kode(nama_kategori, kode_kategori):
    nama_kategori = _normalisasi_nama_kategori(nama_kategori)
    kode_kategori = str(kode_kategori or "").strip().upper()

    if not kode_kategori:
        kode_kategori = _prefix_kategori(nama_kategori)

    field_kategori = _model_field_names(TbKategori)

    if "kode_kategori" in field_kategori:
        return TbKategori.objects.create(
            nama_kategori=nama_kategori,
            kode_kategori=kode_kategori
        )

    with connection.cursor() as cursor:
        cursor.execute(
            """
            INSERT INTO tb_kategori (nama_kategori, kode_kategori)
            VALUES (%s, %s)
            """,
            [nama_kategori, kode_kategori]
        )
        id_baru = cursor.lastrowid

    return TbKategori.objects.get(id_kategori=id_baru)


def _kategori_to_dict(kategori):
    nama = _normalisasi_nama_kategori(getattr(kategori, "nama_kategori", "") or "")
    kode_kategori = _ambil_kode_kategori_obj(kategori)

    jumlah_barang_by_id = 0
    jumlah_barang_by_nama = 0

    try:
        jumlah_barang_by_id = TbBarang.objects.filter(
            id_kategori=getattr(kategori, "id_kategori", 0)
        ).count()
    except Exception:
        jumlah_barang_by_id = 0

    try:
        jumlah_barang_by_nama = TbBarang.objects.filter(
            jenis_barang__iexact=nama
        ).count()
    except Exception:
        jumlah_barang_by_nama = 0

    return {
        "id_kategori": int(getattr(kategori, "id_kategori", 0) or 0),
        "nama_kategori": nama,
        "kode_kategori": kode_kategori,
        "prefix": kode_kategori,
        "jumlah_barang": max(jumlah_barang_by_id, jumlah_barang_by_nama),
    }


def _sinkron_kategori_dari_barang():
    """
    Membentuk dan menyinkronkan kategori dari jenis_barang di tb_barang.
    Jika tb_kategori punya kolom kode_kategori, kolom itu ikut diisi.
    """
    try:
        daftar_jenis = (
            TbBarang.objects
            .exclude(jenis_barang__isnull=True)
            .exclude(jenis_barang__exact="")
            .values_list("jenis_barang", flat=True)
            .distinct()
        )

        for jenis in daftar_jenis:
            nama = _normalisasi_nama_kategori(jenis)
            if not nama:
                continue

            kode_kategori = _kode_kategori_dari_nama(nama)

            kategori = (
                TbKategori.objects
                .filter(nama_kategori__iexact=nama)
                .first()
            )

            if not kategori:
                kategori = _buat_kategori_dengan_kode(nama, kode_kategori)
            else:
                _simpan_kode_kategori_db(kategori.id_kategori, kode_kategori)

            try:
                TbBarang.objects.filter(jenis_barang__iexact=nama).update(
                    id_kategori=kategori.id_kategori,
                    jenis_barang=nama,
                )
            except Exception:
                pass

    except Exception:
        pass


@require_http_methods(["GET"])
def kategori(request):
    gate = _require_login(request)
    if gate:
        return gate

    if not _is_pemilik(request):
        messages.error(request, "Akses ditolak.")
        return redirect("kasir:home")

    _sinkron_kategori_dari_barang()

    return render(request, "kasir/kategori.html", {
        "role": request.session.get("kasir_role", ""),
        "username": request.session.get("kasir_username", ""),
        "nama": request.session.get("kasir_nama", ""),
    })


@require_http_methods(["GET"])
def kategori_json(request):
    gate = _require_login(request)
    if gate:
        return JsonResponse({"ok": False, "msg": "Unauthorized"}, status=401)

    if not _is_pemilik(request):
        return JsonResponse({"ok": False, "msg": "Akses ditolak."}, status=403)

    _sinkron_kategori_dari_barang()

    kategori_qs = TbKategori.objects.all().order_by("nama_kategori")
    data = [_kategori_to_dict(k) for k in kategori_qs]

    return JsonResponse({
        "ok": True,
        "data": data,
    })


@require_http_methods(["POST"])
def kategori_create(request):
    gate = _require_login(request)
    if gate:
        return JsonResponse({"ok": False, "msg": "Unauthorized"}, status=401)

    if not _is_pemilik(request):
        return JsonResponse({"ok": False, "msg": "Akses ditolak."}, status=403)

    payload = _json_body(request)

    nama = _normalisasi_nama_kategori(
        payload.get("nama_kategori")
        or payload.get("nama")
        or payload.get("jenis_barang")
    )

    if not nama:
        return JsonResponse({
            "ok": False,
            "msg": "Nama kategori wajib diisi."
        }, status=400)

    kode_kategori = str(
        payload.get("kode_kategori")
        or payload.get("prefix")
        or ""
    ).strip().upper()

    if not kode_kategori:
        kode_kategori = _kode_kategori_dari_nama(nama)

    if TbKategori.objects.filter(nama_kategori__iexact=nama).exists():
        return JsonResponse({
            "ok": False,
            "msg": "Kategori sudah ada."
        }, status=400)

    try:
        with db_transaction.atomic():
            kategori = _buat_kategori_dengan_kode(nama, kode_kategori)

            TbBarang.objects.filter(jenis_barang__iexact=nama).update(
                id_kategori=kategori.id_kategori,
                jenis_barang=nama,
            )

    except Exception as e:
        return JsonResponse({
            "ok": False,
            "msg": f"Gagal tambah kategori: {e}"
        }, status=500)

    return JsonResponse({
        "ok": True,
        "msg": "Kategori berhasil ditambahkan.",
        "data": _kategori_to_dict(kategori),
    })


@require_http_methods(["POST", "PUT"])
def kategori_edit(request):
    gate = _require_login(request)
    if gate:
        return JsonResponse({"ok": False, "msg": "Unauthorized"}, status=401)

    if not _is_pemilik(request):
        return JsonResponse({"ok": False, "msg": "Akses ditolak."}, status=403)

    payload = _json_body(request)

    id_kategori = _to_int(
        payload.get("id_kategori")
        or payload.get("id")
        or payload.get("pk"),
        0
    )

    nama_baru = _normalisasi_nama_kategori(
        payload.get("nama_kategori")
        or payload.get("nama")
        or payload.get("jenis_barang")
    )

    if id_kategori <= 0:
        return JsonResponse({
            "ok": False,
            "msg": "ID kategori tidak valid."
        }, status=400)

    if not nama_baru:
        return JsonResponse({
            "ok": False,
            "msg": "Nama kategori wajib diisi."
        }, status=400)

    kode_kategori = str(
        payload.get("kode_kategori")
        or payload.get("prefix")
        or ""
    ).strip().upper()

    if not kode_kategori:
        kode_kategori = _kode_kategori_dari_nama(nama_baru)

    try:
        kategori = TbKategori.objects.get(id_kategori=id_kategori)
    except TbKategori.DoesNotExist:
        return JsonResponse({
            "ok": False,
            "msg": "Kategori tidak ditemukan."
        }, status=404)

    nama_lama = _normalisasi_nama_kategori(kategori.nama_kategori)

    duplicate = (
        TbKategori.objects
        .filter(nama_kategori__iexact=nama_baru)
        .exclude(id_kategori=id_kategori)
        .exists()
    )

    if duplicate:
        return JsonResponse({
            "ok": False,
            "msg": "Nama kategori sudah dipakai kategori lain."
        }, status=400)

    try:
        with db_transaction.atomic():
            kategori.nama_kategori = nama_baru

            field_kategori = _model_field_names(TbKategori)

            if "kode_kategori" in field_kategori:
                kategori.kode_kategori = kode_kategori
                kategori.save(update_fields=["nama_kategori", "kode_kategori"])
            else:
                kategori.save(update_fields=["nama_kategori"])
                _simpan_kode_kategori_db(id_kategori, kode_kategori)

            TbBarang.objects.filter(id_kategori=id_kategori).update(
                jenis_barang=nama_baru
            )

            TbBarang.objects.filter(jenis_barang__iexact=nama_lama).update(
                id_kategori=id_kategori,
                jenis_barang=nama_baru,
            )

    except Exception as e:
        return JsonResponse({
            "ok": False,
            "msg": f"Gagal edit kategori: {e}"
        }, status=500)

    kategori = TbKategori.objects.get(id_kategori=id_kategori)

    return JsonResponse({
        "ok": True,
        "msg": "Kategori berhasil diperbarui.",
        "data": _kategori_to_dict(kategori),
    })


@require_http_methods(["POST", "DELETE"])
def kategori_delete(request):
    gate = _require_login(request)
    if gate:
        return JsonResponse({"ok": False, "msg": "Unauthorized"}, status=401)

    if not _is_pemilik(request):
        return JsonResponse({"ok": False, "msg": "Akses ditolak."}, status=403)

    payload = _json_body(request)

    id_kategori = _to_int(
        payload.get("id_kategori")
        or payload.get("id")
        or payload.get("pk"),
        0
    )

    if id_kategori <= 0:
        return JsonResponse({
            "ok": False,
            "msg": "ID kategori tidak valid."
        }, status=400)

    try:
        kategori = TbKategori.objects.get(id_kategori=id_kategori)
    except TbKategori.DoesNotExist:
        return JsonResponse({
            "ok": False,
            "msg": "Kategori tidak ditemukan."
        }, status=404)

    nama = _normalisasi_nama_kategori(kategori.nama_kategori)

    jumlah_barang_by_id = TbBarang.objects.filter(id_kategori=id_kategori).count()
    jumlah_barang_by_nama = TbBarang.objects.filter(jenis_barang__iexact=nama).count()
    jumlah_barang = max(jumlah_barang_by_id, jumlah_barang_by_nama)

    if jumlah_barang > 0:
        return JsonResponse({
            "ok": False,
            "msg": (
                f"Kategori tidak bisa dihapus karena masih dipakai oleh "
                f"{jumlah_barang} barang. Pindahkan/edit barangnya dulu."
            )
        }, status=400)

    try:
        kategori.delete()
    except Exception as e:
        return JsonResponse({
            "ok": False,
            "msg": f"Gagal hapus kategori: {e}"
        }, status=500)

    return JsonResponse({
        "ok": True,
        "msg": "Kategori berhasil dihapus."
    })


@require_http_methods(["POST", "DELETE"])
def kategori_delete_all(request):
    gate = _require_login(request)
    if gate:
        return JsonResponse({"ok": False, "msg": "Unauthorized"}, status=401)

    if not _is_pemilik(request):
        return JsonResponse({"ok": False, "msg": "Akses ditolak."}, status=403)

    kategori_list = list(TbKategori.objects.all())

    if not kategori_list:
        return JsonResponse({
            "ok": True,
            "msg": "Data kategori sudah kosong."
        })

    kategori_dipakai = []

    for kategori in kategori_list:
        id_kategori = int(getattr(kategori, "id_kategori", 0) or 0)
        nama = _normalisasi_nama_kategori(kategori.nama_kategori)

        jumlah_by_id = TbBarang.objects.filter(id_kategori=id_kategori).count()
        jumlah_by_nama = TbBarang.objects.filter(jenis_barang__iexact=nama).count()
        jumlah = max(jumlah_by_id, jumlah_by_nama)

        if jumlah > 0:
            kategori_dipakai.append(f"{nama} ({jumlah} barang)")

    if kategori_dipakai:
        return JsonResponse({
            "ok": False,
            "msg": (
                "Tidak bisa hapus semua kategori karena masih ada kategori "
                "yang dipakai barang: " + ", ".join(kategori_dipakai)
            )
        }, status=400)

    try:
        deleted_count, _ = TbKategori.objects.all().delete()
    except Exception as e:
        return JsonResponse({
            "ok": False,
            "msg": f"Gagal hapus semua kategori: {e}"
        }, status=500)

    return JsonResponse({
        "ok": True,
        "msg": f"Semua kategori berhasil dihapus. Total: {deleted_count}."
    })


# =========================================================
# SUPPLIER + PEMBELIAN
# =========================================================
def _json_body(request):
    try:
        return json.loads(request.body.decode("utf-8") or "{}")
    except Exception:
        return {}


def _generate_no_pembelian_otomatis():
    tz_wib = _get_timezone_wib()
    now = timezone.now()
    try:
        now = timezone.localtime(now, tz_wib) if tz_wib else timezone.localtime(now)
    except Exception:
        pass
    return f"PO-{now.strftime('%d%m%y%H%M%S')}"


@require_http_methods(["GET"])
def pembelian_halaman(request):
    gate = _require_login(request)
    if gate:
        return gate

    if not _is_pemilik(request):
        messages.error(request, "Akses pembelian hanya untuk pemilik.")
        return redirect("kasir:home")

    return render(request, "kasir/pembelian.html", {
        "role": request.session.get("kasir_role", ""),
        "nama": request.session.get("kasir_nama", ""),
    })


@require_http_methods(["GET"])
def supplier_json(request):
    gate = _require_login(request)
    if gate:
        return JsonResponse({"ok": False, "msg": "Belum login."}, status=401)

    if not _is_pemilik(request):
        return JsonResponse({"ok": False, "msg": "Akses ditolak."}, status=403)

    data = []
    qs = TbSupplier.objects.all().order_by("status_supplier", "nama_supplier")

    for s in qs:
        id_supplier = int(getattr(s, "id_supplier", 0) or 0)

        nama_perusahaan = _teks_aman(getattr(s, "nama_perusahaan", ""))
        if not nama_perusahaan:
            nama_perusahaan = _ambil_kolom_opsional_db_text(
                TbSupplier._meta.db_table,
                "id_supplier",
                id_supplier,
                ["nama_perusahaan", "perusahaan", "nama_mitra", "mitra_supplier"],
            )

        data.append({
            "id_supplier": id_supplier,
            "nama_supplier": _teks_aman(getattr(s, "nama_supplier", "")),
            "nama_perusahaan": nama_perusahaan,
            "no_hp": _teks_aman(getattr(s, "no_hp", "")),
            "alamat": _teks_aman(getattr(s, "alamat", "")),
            "status_supplier": _normalize_status(getattr(s, "status_supplier", "aktif")),
        })

    return JsonResponse({"ok": True, "data": data})


@require_http_methods(["POST"])
def supplier_create(request):
    gate = _require_login(request)
    if gate:
        return JsonResponse({"ok": False, "msg": "Belum login."}, status=401)

    if not _is_pemilik(request):
        return JsonResponse({"ok": False, "msg": "Akses ditolak."}, status=403)

    payload = _json_body(request)

    nama_supplier = _teks_aman(payload.get("nama_supplier"))
    nama_perusahaan = _teks_aman(payload.get("nama_perusahaan"))
    no_hp = _teks_aman(payload.get("no_hp"))
    alamat = _teks_aman(payload.get("alamat"))
    status_supplier = _normalize_status(payload.get("status_supplier") or "aktif")

    if not nama_supplier:
        return JsonResponse({"ok": False, "msg": "Nama supplier wajib diisi."}, status=400)

    if not nama_perusahaan:
        return JsonResponse({"ok": False, "msg": "Nama perusahaan / mitra wajib diisi."}, status=400)

    create_data = {
        "nama_supplier": nama_supplier,
        "nama_perusahaan": nama_perusahaan,
        "no_hp": no_hp,
        "alamat": alamat,
        "status_supplier": status_supplier,
    }

    field_names = _model_field_names(TbSupplier)
    now = timezone.now()

    if "created_at" in field_names:
        create_data["created_at"] = now
    if "updated_at" in field_names:
        create_data["updated_at"] = now

    supplier = TbSupplier.objects.create(**_filter_create_kwargs(TbSupplier, create_data))

    id_supplier = int(getattr(supplier, "id_supplier", 0) or 0)

    if "nama_perusahaan" not in field_names and _db_has_column(TbSupplier._meta.db_table, "nama_perusahaan"):
        _force_update_db_column(
            table_name=TbSupplier._meta.db_table,
            pk_col="id_supplier",
            pk_val=id_supplier,
            col="nama_perusahaan",
            val=nama_perusahaan,
        )

    return JsonResponse({
        "ok": True,
        "msg": "Supplier berhasil ditambahkan.",
        "id_supplier": id_supplier,
        "data": {
            "id_supplier": id_supplier,
            "nama_supplier": nama_supplier,
            "nama_perusahaan": nama_perusahaan,
            "no_hp": no_hp,
            "alamat": alamat,
            "status_supplier": status_supplier,
        }
    })


@require_http_methods(["POST"])
def supplier_edit(request):
    gate = _require_login(request)
    if gate:
        return JsonResponse({"ok": False, "msg": "Belum login."}, status=401)

    if not _is_pemilik(request):
        return JsonResponse({"ok": False, "msg": "Akses ditolak."}, status=403)

    payload = _json_body(request)

    id_supplier = _to_int(payload.get("id_supplier"), 0)
    supplier = TbSupplier.objects.filter(id_supplier=id_supplier).first()

    if not supplier:
        return JsonResponse({"ok": False, "msg": "Supplier tidak ditemukan."}, status=404)

    nama_supplier = _teks_aman(payload.get("nama_supplier"))
    nama_perusahaan = _teks_aman(payload.get("nama_perusahaan"))
    no_hp = _teks_aman(payload.get("no_hp"))
    alamat = _teks_aman(payload.get("alamat"))
    status_supplier = _normalize_status(payload.get("status_supplier") or "aktif")

    if not nama_supplier:
        return JsonResponse({"ok": False, "msg": "Nama supplier wajib diisi."}, status=400)

    if not nama_perusahaan:
        return JsonResponse({"ok": False, "msg": "Nama perusahaan / mitra wajib diisi."}, status=400)

    supplier.nama_supplier = nama_supplier
    supplier.nama_perusahaan = nama_perusahaan
    supplier.no_hp = no_hp
    supplier.alamat = alamat
    supplier.status_supplier = status_supplier
    supplier.save()

    return JsonResponse({
        "ok": True,
        "msg": "Supplier berhasil diperbarui.",
        "data": {
            "id_supplier": supplier.id_supplier,
            "nama_supplier": supplier.nama_supplier,
            "nama_perusahaan": supplier.nama_perusahaan,
            "no_hp": supplier.no_hp,
            "alamat": supplier.alamat,
            "status_supplier": supplier.status_supplier,
        }
    })

@require_http_methods(["POST"])
def supplier_delete(request):
    gate = _require_login(request)
    if gate:
        return JsonResponse({"ok": False, "msg": "Belum login."}, status=401)

    if not _is_pemilik(request):
        return JsonResponse({"ok": False, "msg": "Akses ditolak."}, status=403)

    payload = _json_body(request)
    id_supplier = _to_int(payload.get("id_supplier"), 0)

    supplier = TbSupplier.objects.filter(id_supplier=id_supplier).first()

    if not supplier:
        return JsonResponse({"ok": False, "msg": "Supplier tidak ditemukan."}, status=404)

    sudah_dipakai = TbPembelian.objects.filter(id_supplier=id_supplier).exists()

    if sudah_dipakai:
        return JsonResponse({
            "ok": False,
            "msg": "Supplier sudah digunakan pada data pembelian, jadi tidak bisa dihapus. Ubah status menjadi tidak aktif saja."
        }, status=400)

    supplier.delete()

    return JsonResponse({
        "ok": True,
        "msg": "Supplier berhasil dihapus."
    })


@require_http_methods(["GET"])
def pembelian_json(request):
    gate = _require_login(request)
    if gate:
        return JsonResponse({
            "ok": False,
            "msg": "Belum login."
        }, status=401)

    if not _is_pemilik(request):
        return JsonResponse({
            "ok": False,
            "msg": "Akses ditolak."
        }, status=403)

    supplier_map = {}

    for s in TbSupplier.objects.all():
        id_supplier = _to_int(getattr(s, "id_supplier", 0), 0)

        nama_supplier = _teks_aman(getattr(s, "nama_supplier", ""))
        nama_perusahaan = _teks_aman(getattr(s, "nama_perusahaan", ""))

        if not nama_perusahaan:
            nama_perusahaan = _ambil_kolom_opsional_db_text(
                TbSupplier._meta.db_table,
                "id_supplier",
                id_supplier,
                ["nama_perusahaan", "perusahaan", "nama_mitra", "mitra_supplier"],
            )

        supplier_map[id_supplier] = {
            "nama_supplier": nama_supplier,
            "nama_perusahaan": nama_perusahaan,
        }

    data = []
    qs = TbPembelian.objects.all().order_by("-id_pembelian")

    for p in qs:
        id_pembelian = _to_int(getattr(p, "id_pembelian", 0), 0)
        id_supplier = _to_int(getattr(p, "id_supplier", 0), 0)

        tanggal_pembelian = getattr(p, "tanggal_pembelian", None)
        tanggal_diterima = getattr(p, "tanggal_diterima", None)

        supplier_info = supplier_map.get(id_supplier, {})

        detail_items = []

        details = TbDetailPembelian.objects.filter(
            id_pembelian=id_pembelian
        )

        for d in details:
            id_barang = _to_int(getattr(d, "id_barang", 0), 0)

            barang = TbBarang.objects.filter(
                id_barang=id_barang
            ).first()

            qty = _to_int(getattr(d, "qty", 0), 0)
            harga_beli = _safe_decimal_to_int(getattr(d, "harga_beli", 0))
            subtotal = _safe_decimal_to_int(getattr(d, "subtotal", 0))

            if subtotal <= 0:
                subtotal = qty * harga_beli

            detail_items.append({
                "id_barang": id_barang,
                "kode_barang": _teks_aman(getattr(barang, "kode_barang", "")) if barang else "",
                "nama_barang": _teks_aman(getattr(barang, "nama_barang", "")) if barang else "",
                "satuan": _teks_aman(getattr(barang, "satuan", "")) if barang else "-",
                "qty": qty,
                "harga_beli": harga_beli,
                "subtotal": subtotal,
            })

        total_item = _to_int(getattr(p, "total_item", 0), 0)
        total_harga = _safe_decimal_to_int(getattr(p, "total_harga", 0))

        if total_item <= 0:
            total_item = sum(_to_int(item.get("qty"), 0) for item in detail_items)

        if total_harga <= 0:
            total_harga = sum(_to_int(item.get("subtotal"), 0) for item in detail_items)

        data.append({
            "id_pembelian": id_pembelian,
            "no_pembelian": _teks_aman(getattr(p, "no_pembelian", "")),
            "id_supplier": id_supplier,
            "nama_supplier": supplier_info.get("nama_supplier", "-"),
            "nama_perusahaan": supplier_info.get("nama_perusahaan", ""),
            "tanggal_pembelian": tanggal_pembelian.strftime("%Y-%m-%d") if tanggal_pembelian else "",
            "status_pembelian": _teks_aman(getattr(p, "status_pembelian", "")) or "dipesan",
            "total_item": total_item,
            "total_harga": total_harga,
            "tanggal_diterima": tanggal_diterima.strftime("%Y-%m-%d %H:%M:%S") if tanggal_diterima else "",
            "items": detail_items,
        })

    return JsonResponse({
        "ok": True,
        "data": data
    })


@require_http_methods(["POST"])
def pembelian_simpan(request):
    gate = _require_login(request)
    if gate:
        return JsonResponse({
            "ok": False,
            "msg": "Belum login."
        }, status=401)

    if not _is_pemilik(request):
        return JsonResponse({
            "ok": False,
            "msg": "Akses ditolak."
        }, status=403)

    payload = _json_body(request)

    id_supplier_payload = _to_int(payload.get("id_supplier"), 0)
    items = payload.get("items") or []
    langsung_terima = bool(payload.get("langsung_terima"))
    sumber = _teks_aman(payload.get("sumber"), "manual").lower()

    if not isinstance(items, list) or not items:
        return JsonResponse({
            "ok": False,
            "msg": "Detail pembelian masih kosong."
        }, status=400)

    no_pembelian_base = (
        _teks_aman(payload.get("no_pembelian"))
        or _generate_no_pembelian_otomatis()
    )

    tanggal_pembelian = _build_tanggal_waktu_from_payload({
        "tanggal": payload.get("tanggal_pembelian")
    })

    detail_valid = []

    for row in items:
        id_barang = _to_int(row.get("id_barang"), 0)
        qty = max(1, _to_int(row.get("qty"), 0))
        harga_beli = max(0, _to_int(row.get("harga_beli"), 0))

        if id_barang <= 0:
            return JsonResponse({
                "ok": False,
                "msg": "Ada barang yang belum valid di detail pembelian."
            }, status=400)

        if harga_beli <= 0:
            return JsonResponse({
                "ok": False,
                "msg": "Harga beli barang harus lebih dari 0."
            }, status=400)

        barang = TbBarang.objects.filter(id_barang=id_barang).first()

        if not barang:
            return JsonResponse({
                "ok": False,
                "msg": f"Barang dengan id {id_barang} tidak ditemukan."
            }, status=400)

        if sumber == "rekomendasi":
            id_supplier_barang = _ambil_id_supplier_dari_barang(barang)

            if id_supplier_barang <= 0:
                return JsonResponse({
                    "ok": False,
                    "msg": (
                        f"Barang {getattr(barang, 'kode_barang', '')} "
                        f"{getattr(barang, 'nama_barang', '')} belum punya supplier utama."
                    )
                }, status=400)

            id_supplier_final = int(id_supplier_barang)
        else:
            id_supplier_final = int(id_supplier_payload or 0)

            if id_supplier_final <= 0:
                return JsonResponse({
                    "ok": False,
                    "msg": "Supplier wajib dipilih."
                }, status=400)

        supplier = TbSupplier.objects.filter(id_supplier=id_supplier_final).first()

        if not supplier:
            return JsonResponse({
                "ok": False,
                "msg": f"Supplier dengan id {id_supplier_final} tidak ditemukan."
            }, status=400)

        subtotal = qty * harga_beli

        detail_valid.append({
            "id_supplier": id_supplier_final,
            "id_barang": id_barang,
            "qty": qty,
            "harga_beli": harga_beli,
            "subtotal": subtotal,
        })

    if not detail_valid:
        return JsonResponse({
            "ok": False,
            "msg": "Detail pembelian masih kosong."
        }, status=400)

    grouped = defaultdict(list)

    for d in detail_valid:
        grouped[int(d["id_supplier"])].append(d)

    created_pembelian = []

    with db_transaction.atomic():
        now = timezone.now()

        total_group = len(grouped)

        for urutan, (id_supplier_group, detail_group) in enumerate(grouped.items(), start=1):
            supplier = TbSupplier.objects.filter(id_supplier=id_supplier_group).first()

            if not supplier:
                return JsonResponse({
                    "ok": False,
                    "msg": f"Supplier dengan id {id_supplier_group} tidak ditemukan."
                }, status=400)

            total_item = sum(_to_int(d["qty"], 0) for d in detail_group)
            total_harga = sum(_to_int(d["subtotal"], 0) for d in detail_group)

            if total_group > 1:
                no_pembelian = f"{no_pembelian_base}-{urutan}"
            else:
                no_pembelian = no_pembelian_base

            pembelian_data = {
                "no_pembelian": no_pembelian,
                "tanggal_pembelian": tanggal_pembelian,
                "id_supplier": int(id_supplier_group),
                "id_user": int(request.session.get("kasir_user_id") or 0),
                "status_pembelian": "diterima" if langsung_terima else "dipesan",
                "total_item": total_item,
                "total_harga": total_harga,
            }

            if langsung_terima:
                pembelian_data["tanggal_diterima"] = now

            field_names = _model_field_names(TbPembelian)

            if "created_at" in field_names:
                pembelian_data["created_at"] = now

            if "updated_at" in field_names:
                pembelian_data["updated_at"] = now

            pembelian = TbPembelian.objects.create(
                **_filter_create_kwargs(TbPembelian, pembelian_data)
            )

            for d in detail_group:
                detail_data = {
                    "id_pembelian": int(getattr(pembelian, "id_pembelian", 0) or 0),
                    "id_barang": d["id_barang"],
                    "qty": d["qty"],
                    "harga_beli": d["harga_beli"],
                    "subtotal": d["subtotal"],
                }

                TbDetailPembelian.objects.create(
                    **_filter_create_kwargs(TbDetailPembelian, detail_data)
                )

                if langsung_terima:
                    barang = TbBarang.objects.select_for_update().filter(
                        id_barang=d["id_barang"]
                    ).first()

                    if barang:
                        stok_lama = _to_int(getattr(barang, "stok", 0), 0)
                        barang.stok = stok_lama + d["qty"]

                        barang_fields = _model_field_names(TbBarang)
                        update_fields = ["stok"]

                        if "updated_at" in barang_fields:
                            barang.updated_at = now
                            update_fields.append("updated_at")

                        barang.save(update_fields=update_fields)

            created_pembelian.append({
                "id_pembelian": int(getattr(pembelian, "id_pembelian", 0) or 0),
                "no_pembelian": no_pembelian,
                "id_supplier": int(id_supplier_group),
                "nama_supplier": _nama_supplier_display(supplier),
                "total_item": total_item,
                "total_harga": total_harga,
            })

    if len(created_pembelian) > 1:
        msg = (
            f"Pembelian berhasil dibuat menjadi {len(created_pembelian)} nota "
            f"sesuai Supplier Utama masing-masing barang."
        )
    elif langsung_terima:
        msg = "Pembelian berhasil disimpan dan stok langsung bertambah."
    else:
        msg = "Pembelian berhasil disimpan."

    return JsonResponse({
        "ok": True,
        "msg": msg,
        "jumlah_nota": len(created_pembelian),
        "data": created_pembelian,
        "langsung_terima": langsung_terima,
        "sumber": sumber,
    })

@require_http_methods(["POST"])
def pembelian_terima(request):
    gate = _require_login(request)
    if gate:
        return JsonResponse({"ok": False, "msg": "Belum login."}, status=401)

    if not _is_pemilik(request):
        return JsonResponse({"ok": False, "msg": "Akses ditolak."}, status=403)

    payload = _json_body(request)
    id_pembelian = _to_int(payload.get("id_pembelian"), 0)

    with db_transaction.atomic():
        pembelian = TbPembelian.objects.select_for_update().filter(id_pembelian=id_pembelian).first()
        if not pembelian:
            return JsonResponse({"ok": False, "msg": "Data pembelian tidak ditemukan."}, status=404)

        status_now = _teks_aman(getattr(pembelian, "status_pembelian", "")).lower()
        if status_now == "diterima":
            return JsonResponse({"ok": False, "msg": "Pembelian ini sudah diterima sebelumnya."}, status=400)

        details = list(TbDetailPembelian.objects.select_for_update().filter(id_pembelian=id_pembelian))

        for d in details:
            id_barang = _to_int(getattr(d, "id_barang", 0), 0)
            qty = max(0, _to_int(getattr(d, "qty", 0), 0))

            if id_barang <= 0 or qty <= 0:
                continue

            barang = TbBarang.objects.select_for_update().filter(id_barang=id_barang).first()
            if not barang:
                continue

            stok_lama = _to_int(getattr(barang, "stok", 0), 0)
            barang.stok = stok_lama + qty

            barang_fields = _model_field_names(TbBarang)
            if "updated_at" in barang_fields:
                barang.updated_at = timezone.now()

            update_fields = ["stok"]
            if "updated_at" in barang_fields:
                update_fields.append("updated_at")

            barang.save(update_fields=update_fields)

        pembelian.status_pembelian = "diterima"
        pembelian.tanggal_diterima = timezone.now()

        pembelian_fields = _model_field_names(TbPembelian)
        if "updated_at" in pembelian_fields:
            pembelian.updated_at = timezone.now()

        update_fields = ["status_pembelian", "tanggal_diterima"]
        if "updated_at" in pembelian_fields:
            update_fields.append("updated_at")

        pembelian.save(update_fields=update_fields)

    return JsonResponse({"ok": True, "msg": "Pembelian diterima, stok berhasil ditambahkan."})

@require_http_methods(["POST"])
def pembelian_delete(request):
    gate = _require_login(request)
    if gate:
        return JsonResponse({
            "ok": False,
            "msg": "Belum login."
        }, status=401)

    if not _is_pemilik(request):
        return JsonResponse({
            "ok": False,
            "msg": "Akses ditolak."
        }, status=403)

    payload = _json_body(request)
    id_pembelian = _to_int(payload.get("id_pembelian"), 0)

    if id_pembelian <= 0:
        return JsonResponse({
            "ok": False,
            "msg": "ID pembelian tidak valid."
        }, status=400)

    pembelian = TbPembelian.objects.filter(id_pembelian=id_pembelian).first()

    if not pembelian:
        return JsonResponse({
            "ok": False,
            "msg": "Data pembelian tidak ditemukan."
        }, status=404)

    status = str(getattr(pembelian, "status_pembelian", "") or "").lower()

    if status == "diterima":
        return JsonResponse({
            "ok": False,
            "msg": "Pembelian yang sudah diterima tidak bisa dihapus karena stok sudah bertambah."
        }, status=400)

    with db_transaction.atomic():
        TbDetailPembelian.objects.filter(id_pembelian=id_pembelian).delete()
        pembelian.delete()

    return JsonResponse({
        "ok": True,
        "msg": "Pembelian berhasil dihapus."
    })

# =========================================================
# RIWAYAT HELPER
# =========================================================
def _nama_bulan_indonesia(bulan: int) -> str:
    daftar = [
        "",
        "Januari",
        "Februari",
        "Maret",
        "April",
        "Mei",
        "Juni",
        "Juli",
        "Agustus",
        "September",
        "Oktober",
        "November",
        "Desember",
    ]
    if 1 <= int(bulan) <= 12:
        return daftar[int(bulan)]
    return "-"


def _parse_bulan_filter(teks_bulan: str):
    teks_bulan = (teks_bulan or "").strip()
    try:
        hari_ini = timezone.localdate(_get_timezone_wib())
    except Exception:
        hari_ini = timezone.localdate()

    if not teks_bulan:
        return date(hari_ini.year, hari_ini.month, 1)

    try:
        hasil = datetime.strptime(teks_bulan, "%Y-%m").date()
        return date(hasil.year, hasil.month, 1)
    except Exception:
        return date(hari_ini.year, hari_ini.month, 1)


def _range_bulan(tahun: int, bulan: int):
    awal = date(int(tahun), int(bulan), 1)
    if int(bulan) == 12:
        akhir = date(int(tahun) + 1, 1, 1)
    else:
        akhir = date(int(tahun), int(bulan) + 1, 1)
    return awal, akhir

def _range_tahun(tahun: int):
    awal = date(int(tahun), 1, 1)
    akhir = date(int(tahun) + 1, 1, 1)
    return awal, akhir

def _bulan_pendapatan_labels():
    return ["Jan", "Feb", "Mar", "Apr", "Mei", "Jun", "Jul", "Agu", "Sep", "Okt", "Nov", "Des"]

# =========================================================
# RIWAYAT HELPER BULANAN
# =========================================================
def _ambil_ringkasan_bulanan(qs_bulan, peta_user: dict):
    transaksi_bulan = []
    total_item_bulan = 0
    total_penjualan_bulan = 0
    kasir_unik = set()

    kasir_text = "-"
    kasir_ringkas = "-"

    trx_list = list(qs_bulan.order_by("-tanggal_waktu", "-id_transaksi"))
    trx_ids = [
        int(getattr(t, "id_transaksi", 0) or 0)
        for t in trx_list
        if getattr(t, "id_transaksi", None)
    ]

    detail_all = TbTransaksiDetail.objects.filter(
        id_transaksi__in=trx_ids
    ).order_by("id_transaksi", "id_detail")

    detail_map = defaultdict(list)
    id_barang_set = set()

    for d in detail_all:
        trx_id = int(getattr(d, "id_transaksi", 0) or 0)
        detail_map[trx_id].append(d)

        try:
            bid = int(getattr(d, "id_barang", 0) or 0)
            if bid > 0:
                id_barang_set.add(bid)
        except Exception:
            pass

    barang_map = {}
    barang_satuan_map = {}

    if id_barang_set:
        for b in TbBarang.objects.filter(id_barang__in=list(id_barang_set)):
            bid = int(getattr(b, "id_barang", 0) or 0)

            barang_map[bid] = (
                getattr(b, "nama_barang", "") or ""
            ).strip()

            barang_satuan_map[bid] = (
                getattr(b, "satuan", "") or ""
            ).strip()

    for trx in trx_list:
        trx_id = int(getattr(trx, "id_transaksi", 0) or 0)
        dt = getattr(trx, "tanggal_waktu", None)

        if dt:
            dt_wib = _to_wib(dt)
            try:
                tanggal_text = dt_wib.strftime("%d %b %Y")
                jam_text = dt_wib.strftime("%H:%M")
            except Exception:
                tanggal_text, jam_text = "-", "-"
        else:
            tanggal_text, jam_text = "-", "-"

        kasir_nama = _ambil_nama_kasir(trx, peta_user)
        if kasir_nama and kasir_nama != "-":
            kasir_unik.add(kasir_nama)

        detail_items = []
        nama_barang_list = []
        total_item_detail = 0
        total_harga_detail = 0
        besaran_list = []

        for d in detail_map.get(trx_id, []):
            qty = _safe_decimal_to_int(getattr(d, "qty", 0))
            harga_satuan = _normalize_nominal_penuh(
                _safe_decimal_to_int(getattr(d, "harga_satuan", 0))
            )
            subtotal = _normalize_nominal_penuh(
                _safe_decimal_to_int(getattr(d, "subtotal", 0))
            )

            nama_barang = _ambil_snapshot_nama_barang_detail(d, barang_map)
            besaran = _ambil_snapshot_satuan_barang_detail(d, barang_satuan_map)

            total_item_detail += qty
            total_harga_detail += subtotal

            if nama_barang and nama_barang != "-" and nama_barang not in nama_barang_list:
                nama_barang_list.append(nama_barang)
            
            if besaran and besaran != "-" and besaran not in besaran_list:
                besaran_list.append(besaran)

            detail_items.append({
                "nama_barang": nama_barang,
                "besaran": besaran,
                "qty": qty,
                "harga_satuan": _format_rupiah_display(harga_satuan),
                "subtotal": _format_rupiah_display(subtotal),
            })

        total_item_header = _safe_decimal_to_int(getattr(trx, "total_item", 0))
        total_harga_header = _normalize_nominal_penuh(
            _safe_decimal_to_int(getattr(trx, "total_harga", 0))
        )

        total_item_final = total_item_detail if total_item_detail > 0 else total_item_header
        total_harga_final = total_harga_detail if total_harga_detail > 0 else total_harga_header

        nominal_bayar, kembalian = _ambil_nominal_bayar_tampil(
            trx,
            total_harga_final,
        )

        total_item_bulan += total_item_final
        total_penjualan_bulan += total_harga_final

        transaksi_bulan.append({
            "id_transaksi": trx_id,
            "kode_transaksi": _ambil_kode_transaksi(trx),
            "tanggal": tanggal_text,
            "jam": jam_text,
            "kasir": kasir_nama,
            "jenis_transaksi": _label_jenis_transaksi(trx),
            "marketplace": _label_marketplace(trx),
            "nama_barang_list": nama_barang_list,
            "besaran_list": besaran_list,
            "total_item": total_item_final,
            "total_belanja": _format_rupiah_display(total_harga_final),
            "nominal_bayar": _format_rupiah_display(nominal_bayar),
            "kembalian": _format_rupiah_display(kembalian),
            "items": detail_items,
        })

    kasir_list = sorted(kasir_unik)

    if not kasir_list:
        kasir_text = "-"
        kasir_ringkas = "-"
    elif len(kasir_list) == 1:
        kasir_text = kasir_list[0]
        kasir_ringkas = kasir_list[0]
    elif len(kasir_list) == 2:
        kasir_text = ", ".join(kasir_list)
        kasir_ringkas = ", ".join(kasir_list)
    else:
        kasir_text = ", ".join(kasir_list)
        kasir_ringkas = f"{kasir_list[0]}, {kasir_list[1]}, +{len(kasir_list) - 2}"

    return {
        "jumlah_transaksi": len(transaksi_bulan),
        "total_item": total_item_bulan,
        "total_penjualan": total_penjualan_bulan,
        "kasir": kasir_text,
        "kasir_ringkas": kasir_ringkas,
        "transaksi": transaksi_bulan,
    }


# =========================================================
# RIWAYAT PAGE
# =========================================================
def riwayat(request):
    gate = _require_login(request)
    if gate:
        return gate

    role = _normalize_role(request.session.get("kasir_role"))
    user_id = int(request.session.get("kasir_user_id") or 0)
    user_login = TbUser.objects.filter(id_user=user_id).first()

    tampilan = (request.GET.get("tampilan") or "harian").strip().lower()

    if role != "pemilik":
        tampilan = "harian"
    else:
        if tampilan not in ("harian", "bulanan"):
            tampilan = "harian"

    semua_user = TbUser.objects.all().only("id_user", "nama", "username")
    peta_user = {}
    opsi_kasir = []

    for u in semua_user:
        uid = int(getattr(u, "id_user", 0) or 0)
        nama_tampil = (getattr(u, "nama", "") or "").strip() or (getattr(u, "username", "") or "").strip()
        peta_user[uid] = nama_tampil
        opsi_kasir.append({"id": uid, "nama": nama_tampil})

    qs_dasar = TbTransaksi.objects.all().order_by("-tanggal_waktu", "-id_transaksi")
    if role != "pemilik":
        qs_dasar = qs_dasar.filter(id_kasir=user_id)

    konteks = {
        "username": (user_login.username if user_login else request.session.get("kasir_username")),
        "nama": (user_login.nama if user_login else (request.session.get("kasir_nama") or "")),
        "role": role,
        "status": (getattr(user_login, "status", "") if user_login else ""),
        "user_photo": (user_login.foto.url if (user_login and getattr(user_login, "foto", None)) else None),
        "tampilan": tampilan,
    }

    if tampilan == "harian":
        try:
            tanggal_filter = timezone.localdate(_get_timezone_wib())
        except Exception:
            tanggal_filter = timezone.localdate()

        kasir_id_filter = _to_int(request.GET.get("kasir_id"), 0)

        from datetime import timedelta
        qs_harian = _filter_date_range_helper(qs_dasar, tanggal_filter, tanggal_filter + timedelta(days=1))

        if role == "pemilik" and kasir_id_filter > 0:
            qs_harian = qs_harian.filter(id_kasir=kasir_id_filter)

        total_harian = qs_harian.aggregate(
            jumlah_transaksi=Count("id_transaksi"),
            total_penjualan=Sum("total_harga"),
        )

        daftar_riwayat = []
        for trx in qs_harian[:200]:
            dt = getattr(trx, "tanggal_waktu", None)
            if dt:
                dt_wib = _to_wib(dt)
                try:
                    tanggal_text = dt_wib.strftime("%d %b %Y")
                    jam_text = dt_wib.strftime("%H:%M")
                except Exception:
                    tanggal_text, jam_text = "-", "-"
            else:
                tanggal_text, jam_text = "-", "-"

            daftar_riwayat.append({
                "id_transaksi": int(getattr(trx, "id_transaksi", 0) or 0),
                "kode_transaksi": _ambil_kode_transaksi(trx),
                "tanggal": tanggal_text,
                "jam": jam_text,
                "kasir": _ambil_nama_kasir(trx, peta_user),
                "jenis_transaksi": _label_jenis_transaksi(trx),
                "marketplace": _label_marketplace(trx),
                "total_penjualan": _format_rupiah_display(getattr(trx, "total_harga", 0) or 0),
            })

        konteks.update({
            "tanggal_dipilih": tanggal_filter.strftime("%d %b %Y"),
            "tanggal_filter_value": tanggal_filter.strftime("%Y-%m-%d"),
            "opsi_kasir": opsi_kasir,
            "kasir_id_terpilih": kasir_id_filter,
            "nama_kasir": (
                peta_user.get(kasir_id_filter, "Semua Kasir")
                if (role == "pemilik" and kasir_id_filter > 0)
                else (
                    "Semua Kasir"
                    if role == "pemilik"
                    else (request.session.get("kasir_nama") or request.session.get("kasir_username") or "-")
                )
            ),
            "jumlah_transaksi_harian": int(total_harian.get("jumlah_transaksi") or 0),
            "total_penjualan_harian": _format_rupiah_display(total_harian.get("total_penjualan") or 0),
            "daftar_riwayat": daftar_riwayat,
        })
        return render(request, "kasir/riwayat.html", konteks)

    tahun_filter_bulanan = _to_int(request.GET.get("tahun"), 0)
    tampilkan_kolom_tahun = tahun_filter_bulanan <= 0

    qs_bulanan_source = qs_dasar.exclude(tanggal_waktu__isnull=True)

    daftar_tahun_bulanan = sorted(
        {
            int(_to_wib(getattr(trx, "tanggal_waktu", None)).year)
            for trx in qs_bulanan_source
            if getattr(trx, "tanggal_waktu", None)
        },
        reverse=True
    )

    rekap_dict = defaultdict(lambda: {"jumlah_transaksi_bulan": 0, "total_penjualan_bulan": 0})
    for trx in qs_bulanan_source.only("id_transaksi", "tanggal_waktu", "total_harga"):
        dt = trx.tanggal_waktu
        if not dt:
            continue
        dt_wib = _to_wib(dt)
        if not dt_wib:
            continue
        bulan_rekap = dt_wib.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        rekap_dict[bulan_rekap]["jumlah_transaksi_bulan"] += 1
        rekap_dict[bulan_rekap]["total_penjualan_bulan"] += int(trx.total_harga or 0)

    rekap_bulanan = []
    for bulan_rekap, vals in sorted(rekap_dict.items(), reverse=True):
        rekap_bulanan.append({
            "bulan_rekap": bulan_rekap,
            "jumlah_transaksi_bulan": vals["jumlah_transaksi_bulan"],
            "total_penjualan_bulan": vals["total_penjualan_bulan"],
        })

    daftar_bulanan_semua = []
    jumlah_transaksi_bulanan_total = 0
    total_penjualan_bulanan_total_raw = 0

    for item_bulan in rekap_bulanan:
        bulan_obj = item_bulan.get("bulan_rekap")
        if not bulan_obj:
            continue

        bulan_obj = _to_wib(bulan_obj)
        tahun = int(bulan_obj.year)
        bulan = int(bulan_obj.month)

        if tahun_filter_bulanan > 0 and tahun != tahun_filter_bulanan:
            continue

        awal_bulan, akhir_bulan = _range_bulan(tahun, bulan)

        qs_bulan = _filter_date_range_helper(qs_bulanan_source, awal_bulan, akhir_bulan)

        ringkas = _ambil_ringkasan_bulanan(qs_bulan, peta_user)

        jumlah_transaksi_bulanan_total += ringkas["jumlah_transaksi"]
        total_penjualan_bulanan_total_raw += ringkas["total_penjualan"]

        daftar_bulanan_semua.append({
            "tahun": tahun,
            "tahun_display": str(tahun),
            "bulan_angka": bulan,
            "nama_bulan": _nama_bulan_indonesia(bulan),
            "jumlah_transaksi_bulan": ringkas["jumlah_transaksi"],
            "total_item_bulan": ringkas["total_item"],
            "total_penjualan_bulan": _format_rupiah_display(ringkas["total_penjualan"]),
            "kasir_bulan": ringkas.get("kasir_ringkas", ringkas["kasir"]),
            "kasir_id": 0 if role == "pemilik" else user_id,
        })

    konteks.update({
        "daftar_bulanan_semua": daftar_bulanan_semua,
        "daftar_tahun_bulanan": daftar_tahun_bulanan,
        "tahun_filter_bulanan": tahun_filter_bulanan if tahun_filter_bulanan > 0 else "",
        "tampilkan_kolom_tahun": tampilkan_kolom_tahun,
        "jumlah_bulan_terfilter": len(daftar_bulanan_semua),
        "jumlah_transaksi_bulanan_total": jumlah_transaksi_bulanan_total,
        "total_penjualan_bulanan_total": _format_rupiah_display(total_penjualan_bulanan_total_raw),
    })
    return render(request, "kasir/riwayat.html", konteks)

# =========================================================
# RIWAYAT DETAIL TRANSAKSI
# =========================================================
@require_http_methods(["GET"])
def riwayat_detail(request, id_transaksi):
    gate = _require_login(request)
    if gate:
        return JsonResponse({"ok": False, "msg": "Unauthorized"}, status=401)

    try:
        role = _normalize_role(request.session.get("kasir_role"))
        user_id = int(request.session.get("kasir_user_id") or 0)

        trx = TbTransaksi.objects.filter(id_transaksi=id_transaksi).first()
        if not trx:
            return JsonResponse({"ok": False, "msg": "Transaksi tidak ditemukan."}, status=404)

        if role != "pemilik":
            id_kasir_trx = int(getattr(trx, "id_kasir", 0) or 0)
            if id_kasir_trx != user_id:
                return JsonResponse({"ok": False, "msg": "Akses ditolak."}, status=403)

        peta_user = {}
        for u in TbUser.objects.all().only("id_user", "nama", "username"):
            uid = int(getattr(u, "id_user", 0) or 0)
            peta_user[uid] = (
                (getattr(u, "nama", "") or "").strip()
                or (getattr(u, "username", "") or "").strip()
            )

        detail_qs = TbTransaksiDetail.objects.filter(
            id_transaksi=int(id_transaksi)
        ).order_by("id_detail")

        id_barang_list = []
        for d in detail_qs:
            try:
                bid = int(getattr(d, "id_barang", 0) or 0)
                if bid > 0:
                    id_barang_list.append(bid)
            except Exception:
                pass

        barang_map = {}
        barang_satuan_map = {}

        if id_barang_list:
            for b in TbBarang.objects.filter(id_barang__in=id_barang_list):
                bid = int(getattr(b, "id_barang", 0) or 0)

                barang_map[bid] = (
                    getattr(b, "nama_barang", "") or ""
                ).strip()

                barang_satuan_map[bid] = (
                    getattr(b, "satuan", "") or ""
                ).strip()

        items = []
        total_item_detail = 0
        total_belanja_detail = 0

        for d in detail_qs:
            qty = _safe_decimal_to_int(getattr(d, "qty", 0))
            harga_satuan = _normalize_nominal_penuh(
                _safe_decimal_to_int(getattr(d, "harga_satuan", 0))
            )
            subtotal = _normalize_nominal_penuh(
                _safe_decimal_to_int(getattr(d, "subtotal", 0))
            )

            total_item_detail += qty
            total_belanja_detail += subtotal

            items.append({
                "nama_barang": _ambil_snapshot_nama_barang_detail(d, barang_map),
                "besaran": _ambil_snapshot_satuan_barang_detail(d, barang_satuan_map),
                "qty": qty,
                "harga_satuan": _format_rupiah_display(harga_satuan),
                "subtotal": _format_rupiah_display(subtotal),
            })


        dt = getattr(trx, "tanggal_waktu", None)
        if dt:
            dt_wib = _to_wib(dt)
            try:
                tanggal_text = dt_wib.strftime("%d %b %Y")
                jam_text = dt_wib.strftime("%H:%M")
            except Exception:
                tanggal_text, jam_text = "-", "-"
        else:
            tanggal_text, jam_text = "-", "-"

        trx_id = int(getattr(trx, "id_transaksi", 0) or 0)

        total_belanja_header = _normalize_nominal_penuh(
            _safe_decimal_to_int(getattr(trx, "total_harga", 0))
        )
        total_item_header = _safe_decimal_to_int(getattr(trx, "total_item", 0))

        total_item_final = total_item_detail if total_item_detail > 0 else total_item_header
        total_belanja_final = total_belanja_detail if total_belanja_detail > 0 else total_belanja_header

        nominal_bayar, kembalian = _ambil_nominal_bayar_tampil(
            trx,
            total_belanja_final,
        )

        return JsonResponse({
            "ok": True,
            "data": {
                "id_transaksi": trx_id,
                "kode_transaksi": _ambil_kode_transaksi(trx),
                "tanggal": tanggal_text,
                "jam": jam_text,
                "kasir": _ambil_nama_kasir(trx, peta_user),
                "jenis_transaksi": _label_jenis_transaksi(trx),
                "marketplace": _label_marketplace(trx),
                "total_item": total_item_final,
                "total_belanja": _format_rupiah_display(total_belanja_final),
                "nominal_bayar": _format_rupiah_display(nominal_bayar),
                "kembalian": _format_rupiah_display(kembalian),
                "items": items,
            }
        })

    except Exception as e:
        return JsonResponse({
            "ok": False,
            "msg": f"Gagal memuat detail transaksi: {e}"
        }, status=500)


# =========================================================
# RIWAYAT DETAIL BULANAN (UNTUK POPUP)
# =========================================================
@require_http_methods(["GET"])
def riwayat_bulanan_detail(request, tahun, bulan):
    gate = _require_login(request)
    if gate:
        return JsonResponse({"ok": False, "msg": "Unauthorized"}, status=401)

    try:
        role = _normalize_role(request.session.get("kasir_role"))
        user_id = int(request.session.get("kasir_user_id") or 0)

        tahun = int(tahun)
        bulan = int(bulan)
        if bulan < 1 or bulan > 12:
            return JsonResponse({"ok": False, "msg": "Bulan tidak valid."}, status=400)

        kasir_id = _to_int(request.GET.get("kasir_id"), 0)

        semua_user = TbUser.objects.all().only("id_user", "nama", "username")
        peta_user = {}
        for u in semua_user:
            uid = int(getattr(u, "id_user", 0) or 0)
            peta_user[uid] = (
                (getattr(u, "nama", "") or "").strip()
                or (getattr(u, "username", "") or "").strip()
            )

        qs_dasar = TbTransaksi.objects.all().order_by("-tanggal_waktu", "-id_transaksi")
        if role != "pemilik":
            qs_dasar = qs_dasar.filter(id_kasir=user_id)
        elif kasir_id > 0:
            qs_dasar = qs_dasar.filter(id_kasir=kasir_id)

        awal_bulan, akhir_bulan = _range_bulan(tahun, bulan)

        qs_bulan = _filter_date_range_helper(qs_dasar, awal_bulan, akhir_bulan)

        hasil = _ambil_ringkasan_bulanan(qs_bulan, peta_user)

        return JsonResponse({
            "ok": True,
            "data": {
                "tahun": tahun,
                "bulan": bulan,
                "nama_bulan": f"{_nama_bulan_indonesia(bulan)} {tahun}",
                "jumlah_transaksi": hasil["jumlah_transaksi"],
                "total_item": hasil["total_item"],
                "total_penjualan": _format_rupiah_display(
                    _normalize_nominal_penuh(hasil["total_penjualan"])
                ),
                "kasir": hasil["kasir"],
                "transaksi": hasil["transaksi"],
            }
        })

    except Exception as e:
        return JsonResponse({
            "ok": False,
            "msg": f"Gagal memuat detail bulanan: {e}"
        }, status=500)

    
def pendapatan(request):
    gate = _require_login(request)
    if gate:
        return gate

    if not _is_pemilik(request):
        messages.error(request, "Akses ditolak.")
        return redirect("kasir:home")

    role = _normalize_role(request.session.get("kasir_role"))
    user_id = int(request.session.get("kasir_user_id") or 0)
    user_login = TbUser.objects.filter(id_user=user_id).first()

    tahun_filter = _to_int(request.GET.get("tahun"), 0)
    if tahun_filter < 0:
        tahun_filter = 0

    peta_user = {}
    for u in TbUser.objects.all().only("id_user", "nama", "username"):
        uid = int(getattr(u, "id_user", 0) or 0)
        peta_user[uid] = (
            (getattr(u, "nama", "") or "").strip()
            or (getattr(u, "username", "") or "").strip()
        )

    qs_dasar = (
        TbTransaksi.objects
        .exclude(tanggal_waktu__isnull=True)
        .order_by("-tanggal_waktu", "-id_transaksi")
    )

    daftar_tahun = sorted(
        {
            int(_to_wib(trx.tanggal_waktu).year)
            for trx in qs_dasar
            if getattr(trx, "tanggal_waktu", None)
        },
        reverse=True,
    )

    chart_labels = []
    chart_values = []

    label_tertinggi = "Bulan Tertinggi"
    label_terendah = "Bulan Terendah"
    periode_tertinggi = ""
    periode_terendah = ""
    nominal_tertinggi = _format_rupiah_display(0)
    nominal_terendah = _format_rupiah_display(0)
    chart_title = "Tren Pendapatan Bulanan"
    chart_subtitle = "Perbandingan pendapatan setiap bulan"

    if tahun_filter > 0:
        chart_labels = _bulan_pendapatan_labels()
        chart_values = []

        for bulan in range(1, 13):
            awal_bulan, akhir_bulan = _range_bulan(tahun_filter, bulan)

            qs_bulan = _filter_date_range_helper(qs_dasar, awal_bulan, akhir_bulan)

            hasil_bulan = _ambil_ringkasan_bulanan(qs_bulan, peta_user)
            total_bulan = _normalize_nominal_penuh(hasil_bulan.get("total_penjualan", 0) or 0)

            if total_bulan < 0:
                total_bulan = 0

            chart_values.append(int(total_bulan))

        bulan_positif = [(i + 1, v) for i, v in enumerate(chart_values) if int(v or 0) > 0]

        if bulan_positif:
            bulan_tertinggi_ke, nilai_tertinggi = max(bulan_positif, key=lambda x: x[1])
            periode_tertinggi = chart_labels[bulan_tertinggi_ke - 1]
            nominal_tertinggi = _format_rupiah_display(nilai_tertinggi)

        if len(bulan_positif) >= 2:
            bulan_terendah_ke, nilai_terendah = min(bulan_positif, key=lambda x: x[1])
            periode_terendah = chart_labels[bulan_terendah_ke - 1]
            nominal_terendah = _format_rupiah_display(nilai_terendah)

        label_tertinggi = "Bulan Tertinggi"
        label_terendah = "Bulan Terendah"
        chart_title = "Tren Pendapatan Bulanan"
        chart_subtitle = f"Perbandingan pendapatan setiap bulan pada tahun {tahun_filter}"

    else:
        tahun_urut_naik = sorted(daftar_tahun)
        chart_labels = [str(th) for th in tahun_urut_naik]
        chart_values = []

        for th in tahun_urut_naik:
            awal_tahun, akhir_tahun = _range_tahun(th)

            qs_tahun = _filter_date_range_helper(qs_dasar, awal_tahun, akhir_tahun)

            hasil_tahun = _ambil_ringkasan_bulanan(qs_tahun, peta_user)
            total_tahun = _normalize_nominal_penuh(hasil_tahun.get("total_penjualan", 0) or 0)

            if total_tahun < 0:
                total_tahun = 0

            chart_values.append(int(total_tahun))

        tahun_positif = [
            (chart_labels[i], v)
            for i, v in enumerate(chart_values)
            if int(v or 0) > 0
        ]

        if tahun_positif:
            periode_tertinggi, nilai_tertinggi = max(tahun_positif, key=lambda x: x[1])
            nominal_tertinggi = _format_rupiah_display(nilai_tertinggi)

        if len(tahun_positif) >= 2:
            periode_terendah, nilai_terendah = min(tahun_positif, key=lambda x: x[1])
            nominal_terendah = _format_rupiah_display(nilai_terendah)

        label_tertinggi = "Tahun Tertinggi"
        label_terendah = "Tahun Terendah"
        chart_title = "Tren Pendapatan Tahunan"
        chart_subtitle = "Perbandingan pendapatan setiap tahun"

    context = {
        "username": (user_login.username if user_login else request.session.get("kasir_username")),
        "nama": (user_login.nama if user_login else (request.session.get("kasir_nama") or "")),
        "role": role,
        "status": (getattr(user_login, "status", "") if user_login else ""),
        "user_photo": (user_login.foto.url if (user_login and getattr(user_login, "foto", None)) else None),

        "tahun_filter": tahun_filter if tahun_filter > 0 else 0,
        "daftar_tahun": daftar_tahun,

        "label_tertinggi": label_tertinggi,
        "label_terendah": label_terendah,
        "periode_tertinggi": periode_tertinggi,
        "periode_terendah": periode_terendah,
        "nominal_tertinggi": nominal_tertinggi,
        "nominal_terendah": nominal_terendah,

        "chart_title": chart_title,
        "chart_subtitle": chart_subtitle,
        "chart_labels_json": json.dumps(chart_labels, ensure_ascii=False),
        "chart_values_json": json.dumps(chart_values),
    }

    return render(request, "kasir/pendapatan.html", context)

@require_http_methods(["GET"])
def download_riwayat_bulanan(request):
    gate = _require_login(request)
    if gate:
        return gate

    if not _is_pemilik(request):
        return HttpResponse("Akses ditolak.", status=403)

    tahun_filter = _to_int(request.GET.get("tahun"), 0)

    peta_user = {}
    for u in TbUser.objects.all().only("id_user", "nama", "username"):
        uid = int(getattr(u, "id_user", 0) or 0)
        peta_user[uid] = (
            (getattr(u, "nama", "") or "").strip()
            or (getattr(u, "username", "") or "").strip()
        )

    qs_transaksi = (
        TbTransaksi.objects
        .exclude(tanggal_waktu__isnull=True)
        .order_by("tanggal_waktu", "id_transaksi")
    )

    if tahun_filter > 0:
        qs_transaksi = qs_transaksi.filter(tanggal_waktu__year=tahun_filter)

    trx_list = list(qs_transaksi)

    trx_ids = [
        int(getattr(t, "id_transaksi", 0) or 0)
        for t in trx_list
        if getattr(t, "id_transaksi", None)
    ]

    detail_all = TbTransaksiDetail.objects.filter(
        id_transaksi__in=trx_ids
    ).order_by("id_transaksi", "id_detail")

    detail_map = defaultdict(list)
    id_barang_set = set()

    for d in detail_all:
        trx_id = int(getattr(d, "id_transaksi", 0) or 0)
        detail_map[trx_id].append(d)

        try:
            bid = int(getattr(d, "id_barang", 0) or 0)
            if bid > 0:
                id_barang_set.add(bid)
        except Exception:
            pass

    barang_map = {}
    barang_satuan_map = {}

    if id_barang_set:
        for b in TbBarang.objects.filter(id_barang__in=list(id_barang_set)):
            bid = int(getattr(b, "id_barang", 0) or 0)

            barang_map[bid] = (
                getattr(b, "nama_barang", "") or ""
            ).strip()

            barang_satuan_map[bid] = (
                getattr(b, "satuan", "") or ""
            ).strip()

    wb = Workbook()
    ws = wb.active
    ws.title = "Riwayat Bulanan"

    headers = [
        "No. Jual",
        "Tanggal",
        "Jam",
        "Bulan",
        "Tahun",
        "Kasir",
        "Jenis Transaksi",
        "Marketplace",
        "Nama Barang",
        "Besaran",
        "Qty",
        "Harga Satuan",
        "Subtotal",
        "Total Item Transaksi",
        "Total Belanja Transaksi",
        "Nominal Bayar",
        "Kembalian",
    ]

    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="4E8B3A")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(border_style="thin", color="DDDDDD")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for trx in trx_list:
        trx_id = int(getattr(trx, "id_transaksi", 0) or 0)

        dt = getattr(trx, "tanggal_waktu", None)
        if dt:
            dt_wib = _to_wib(dt)
            try:
                tanggal_text = dt_wib.strftime("%d %b %Y")
                jam_text = dt_wib.strftime("%H:%M")
                bulan_text = _nama_bulan_indonesia(int(dt_wib.month))
                tahun_text = int(dt_wib.year)
            except Exception:
                tanggal_text, jam_text, bulan_text, tahun_text = "-", "-", "-", "-"
        else:
            tanggal_text, jam_text, bulan_text, tahun_text = "-", "-", "-", "-"

        detail_list = detail_map.get(trx_id, [])

        total_item_detail = 0
        total_harga_detail = 0

        detail_rows = []

        for d in detail_list:
            qty = _safe_decimal_to_int(getattr(d, "qty", 0))
            harga_satuan = _normalize_nominal_penuh(
                _safe_decimal_to_int(getattr(d, "harga_satuan", 0))
            )
            subtotal = _normalize_nominal_penuh(
                _safe_decimal_to_int(getattr(d, "subtotal", 0))
            )

            total_item_detail += qty
            total_harga_detail += subtotal

            detail_rows.append({
                "nama_barang": _ambil_snapshot_nama_barang_detail(d, barang_map),
                "besaran": _ambil_snapshot_satuan_barang_detail(d, barang_satuan_map),
                "qty": qty,
                "harga_satuan": harga_satuan,
                "subtotal": subtotal,
            })

        total_item_header = _safe_decimal_to_int(getattr(trx, "total_item", 0))
        total_harga_header = _normalize_nominal_penuh(
            _safe_decimal_to_int(getattr(trx, "total_harga", 0))
        )

        total_item_final = total_item_detail if total_item_detail > 0 else total_item_header
        total_harga_final = total_harga_detail if total_harga_detail > 0 else total_harga_header

        nominal_bayar, kembalian = _ambil_nominal_bayar_tampil(
            trx,
            total_harga_final,
        )

        if not detail_rows:
            detail_rows = [{
                "nama_barang": "-",
                "besaran": "-",
                "qty": 0,
                "harga_satuan": 0,
                "subtotal": 0,
            }]

        for item in detail_rows:
            ws.append([
                _ambil_kode_transaksi(trx),
                tanggal_text,
                jam_text,
                bulan_text,
                tahun_text,
                _ambil_nama_kasir(trx, peta_user),
                _label_jenis_transaksi(trx),
                _label_marketplace(trx),
                item["nama_barang"],
                item["besaran"],
                item["qty"],
                item["harga_satuan"],
                item["subtotal"],
                total_item_final,
                total_harga_final,
                nominal_bayar,
                kembalian,
            ])

    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, int):
                cell.number_format = '#,##0'

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        ws.column_dimensions[col_letter].width = min(max_length + 3, 36)

    filename = "riwayat_bulanan"
    if tahun_filter > 0:
        filename += f"_{tahun_filter}"
    filename += ".xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


# =========================================================
# retur
# =========================================================

def _rp_dictfetchall(cursor):
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


def _rp_baca_json_request(request):
    try:
        return json.loads(request.body.decode("utf-8") or "{}")
    except Exception:
        return {}


def _rp_int(nilai, default=0):
    try:
        if nilai is None or nilai == "":
            return default
        return int(float(nilai))
    except Exception:
        return default


def _rp_decimal(nilai, default=0):
    try:
        if nilai is None or nilai == "":
            return default

        teks = str(nilai).strip()
        teks = teks.replace("Rp", "").replace("rp", "")
        teks = teks.replace(".", "").replace(",", "")
        teks = teks.replace(" ", "")

        if teks == "":
            return default

        return int(float(teks))
    except Exception:
        return default


def _rp_tolak_json_pemilik(request):
    gate = _require_login(request)
    if gate:
        return JsonResponse(
            {"ok": False, "message": "Silakan login terlebih dahulu."},
            status=401
        )

    if not _is_pemilik(request):
        return JsonResponse(
            {"ok": False, "message": "Akses ditolak. Menu ini hanya untuk pemilik."},
            status=403
        )

    return None


def _rp_tolak_page_pemilik(request):
    gate = _require_login(request)
    if gate:
        return gate

    if not _is_pemilik(request):
        return redirect("kasir:home")

    return None


def _rp_kolom_stok_barang():
    """
    Deteksi kolom stok di tb_barang.
    Aman untuk database yang memakai stok_saat_ini atau stok.
    """
    with connection.cursor() as cursor:
        cursor.execute("SHOW COLUMNS FROM tb_barang")
        kolom = [row[0] for row in cursor.fetchall()]

    if "stok_saat_ini" in kolom:
        return "stok_saat_ini"

    if "stok" in kolom:
        return "stok"

    raise Exception("Kolom stok di tb_barang tidak ditemukan. Gunakan stok_saat_ini atau stok.")


def _rp_update_stok_barang(cursor, id_barang, perubahan):
    """
    perubahan negatif = stok berkurang
    perubahan positif = stok bertambah
    """
    id_barang = _rp_int(id_barang, 0)
    perubahan = _rp_int(perubahan, 0)

    if id_barang <= 0 or perubahan == 0:
        return

    kolom_stok = _rp_kolom_stok_barang()

    cursor.execute(
        f"""
        UPDATE tb_barang
        SET `{kolom_stok}` = COALESCE(`{kolom_stok}`, 0) + %s
        WHERE id_barang = %s
        """,
        [perubahan, id_barang]
    )


def _rp_ambil_stok_barang(cursor, id_barang):
    id_barang = _rp_int(id_barang, 0)

    if id_barang <= 0:
        return None

    kolom_stok = _rp_kolom_stok_barang()

    cursor.execute(
        f"""
        SELECT COALESCE(`{kolom_stok}`, 0)
        FROM tb_barang
        WHERE id_barang = %s
        FOR UPDATE
        """,
        [id_barang]
    )

    row = cursor.fetchone()
    if not row:
        return None

    return _rp_int(row[0], 0)


def _rp_update_stok_barang(cursor, id_barang, perubahan):
    """
    perubahan negatif = stok berkurang
    perubahan positif = stok bertambah
    """
    kolom_stok = _rp_kolom_stok_barang()

    cursor.execute(
        f"""
        UPDATE tb_barang
        SET {kolom_stok} = {kolom_stok} + %s
        WHERE id_barang = %s
        """,
        [perubahan, id_barang]
    )


def _rp_ambil_stok_barang(cursor, id_barang):
    kolom_stok = _rp_kolom_stok_barang()

    cursor.execute(
        f"""
        SELECT {kolom_stok}
        FROM tb_barang
        WHERE id_barang = %s
        FOR UPDATE
        """,
        [id_barang]
    )

    row = cursor.fetchone()
    if not row:
        return None

    return _rp_int(row[0], 0)


@require_http_methods(["GET"])
def retur(request):
    gate = _rp_tolak_page_pemilik(request)
    if gate:
        return gate

    return render(request, "kasir/retur.html", {
        "role": request.session.get("kasir_role", ""),
        "nama": request.session.get("kasir_nama", ""),
    })


@require_http_methods(["GET"])
def pengembalian(request):
    gate = _rp_tolak_page_pemilik(request)
    if gate:
        return gate

    return render(request, "kasir/pengembalian.html", {
        "role": request.session.get("kasir_role", ""),
        "nama": request.session.get("kasir_nama", ""),
    })


@require_http_methods(["GET"])
def retur_detail_options(request):
    gate = _rp_tolak_json_pemilik(request)
    if gate:
        return gate

    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT
                dp.id_detail_pembelian,
                dp.id_pembelian,
                dp.id_barang,
                COALESCE(dp.qty, 0) AS qty,
                COALESCE(dp.harga_beli, 0) AS harga_beli,

                p.no_pembelian,
                p.tanggal_pembelian,
                p.status_pembelian,

                b.kode_barang,
                b.nama_barang,
                b.satuan,

                COALESCE(sb.nama_supplier, sp.nama_supplier, '-') AS nama_supplier,
                COALESCE(sb.nama_perusahaan, sp.nama_perusahaan, '-') AS nama_perusahaan,

                COALESCE((
                    SELECT SUM(r.jumlah_retur)
                    FROM tb_retur r
                    WHERE r.id_detail_pembelian = dp.id_detail_pembelian
                    AND r.status_retur <> 'batal'
                ), 0) AS total_sudah_retur

            FROM tb_detail_pembelian dp
            INNER JOIN tb_pembelian p
                ON p.id_pembelian = dp.id_pembelian
            INNER JOIN tb_barang b
                ON b.id_barang = dp.id_barang
            LEFT JOIN tb_supplier sb
                ON sb.id_supplier = b.id_supplier
            LEFT JOIN tb_supplier sp
                ON sp.id_supplier = p.id_supplier
            WHERE p.status_pembelian = 'diterima'
            ORDER BY p.tanggal_pembelian DESC, dp.id_detail_pembelian DESC
        """)

        rows = _rp_dictfetchall(cursor)

    hasil = []

    for row in rows:
        qty_beli = _rp_int(row.get("qty"), 0)
        sudah_retur = _rp_int(row.get("total_sudah_retur"), 0)
        sisa_bisa_retur = max(0, qty_beli - sudah_retur)

        if sisa_bisa_retur <= 0:
            continue

        row["sisa_bisa_retur"] = sisa_bisa_retur
        row["label"] = (
            f"{row.get('no_pembelian')} | "
            f"{row.get('kode_barang')} - {row.get('nama_barang')} - {row.get('satuan')} | "
            f"Supplier: {row.get('nama_supplier')} / {row.get('nama_perusahaan')} | "
            f"Sisa bisa retur {sisa_bisa_retur}"
        )

        hasil.append(row)

    return JsonResponse({
        "ok": True,
        "items": hasil
    })


@require_http_methods(["GET"])
def retur_json(request):
    gate = _rp_tolak_json_pemilik(request)
    if gate:
        return gate

    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT
                r.id_retur,
                r.id_detail_pembelian,
                r.tanggal_retur,
                COALESCE(r.jumlah_retur, 0) AS jumlah_retur,
                r.alasan,
                r.status_retur,

                COALESCE(dp.harga_beli, 0) AS harga_beli,
                (COALESCE(r.jumlah_retur, 0) * COALESCE(dp.harga_beli, 0)) AS nominal_refund_saran,

                p.no_pembelian,
                p.tanggal_pembelian,

                b.kode_barang,
                b.nama_barang,
                b.satuan,

                COALESCE(sb.nama_supplier, sp.nama_supplier, '-') AS nama_supplier,
                COALESCE(sb.nama_perusahaan, sp.nama_perusahaan, '-') AS nama_perusahaan,

                CASE
                    WHEN r.status_retur = 'diproses' THEN 'Belum Selesai'
                    WHEN r.status_retur = 'selesai' THEN 'Selesai'
                    WHEN r.status_retur = 'batal' THEN 'Batal'
                    WHEN r.status_retur = 'diajukan' THEN 'Belum Selesai'
                    ELSE COALESCE(r.status_retur, '-')
                END AS status_label

            FROM tb_retur r
            INNER JOIN tb_detail_pembelian dp
                ON dp.id_detail_pembelian = r.id_detail_pembelian
            INNER JOIN tb_pembelian p
                ON p.id_pembelian = dp.id_pembelian
            INNER JOIN tb_barang b
                ON b.id_barang = dp.id_barang
            LEFT JOIN tb_supplier sb
                ON sb.id_supplier = b.id_supplier
            LEFT JOIN tb_supplier sp
                ON sp.id_supplier = p.id_supplier
            ORDER BY r.tanggal_retur DESC, r.id_retur DESC
        """)

        rows = _rp_dictfetchall(cursor)

    return JsonResponse({
        "ok": True,
        "items": rows
    })


@require_http_methods(["POST"])
def retur_create(request):
    gate = _rp_tolak_json_pemilik(request)
    if gate:
        return gate

    data = _rp_baca_json_request(request)

    id_detail_pembelian = _rp_int(data.get("id_detail_pembelian"), 0)
    jumlah_retur = _rp_int(data.get("jumlah_retur"), 0)
    alasan = str(data.get("alasan") or "").strip()

    if id_detail_pembelian <= 0:
        return JsonResponse({
            "ok": False,
            "message": "Item pembelian wajib dipilih."
        }, status=400)

    if jumlah_retur <= 0:
        return JsonResponse({
            "ok": False,
            "message": "Jumlah retur harus lebih dari 0."
        }, status=400)

    if not alasan:
        return JsonResponse({
            "ok": False,
            "message": "Alasan retur wajib diisi."
        }, status=400)

    with db_transaction.atomic():
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT
                    COALESCE(dp.qty, 0) AS qty,
                    dp.id_barang,
                    p.status_pembelian,
                    COALESCE((
                        SELECT SUM(r.jumlah_retur)
                        FROM tb_retur r
                        WHERE r.id_detail_pembelian = dp.id_detail_pembelian
                        AND r.status_retur <> 'batal'
                    ), 0) AS total_sudah_retur
                FROM tb_detail_pembelian dp
                INNER JOIN tb_pembelian p
                    ON p.id_pembelian = dp.id_pembelian
                WHERE dp.id_detail_pembelian = %s
                FOR UPDATE
            """, [id_detail_pembelian])

            row = cursor.fetchone()

            if not row:
                return JsonResponse({
                    "ok": False,
                    "message": "Item pembelian tidak ditemukan."
                }, status=404)

            qty_pembelian = _rp_int(row[0], 0)
            id_barang = _rp_int(row[1], 0)
            status_pembelian = str(row[2] or "").strip().lower()
            sudah_retur = _rp_int(row[3], 0)

            if status_pembelian != "diterima":
                return JsonResponse({
                    "ok": False,
                    "message": "Barang hanya bisa diretur jika pembelian sudah diterima."
                }, status=400)

            sisa_bisa_retur = max(0, qty_pembelian - sudah_retur)

            if jumlah_retur > sisa_bisa_retur:
                return JsonResponse({
                    "ok": False,
                    "message": f"Jumlah retur melebihi sisa. Maksimal {sisa_bisa_retur}."
                }, status=400)

            stok_sekarang = _rp_ambil_stok_barang(cursor, id_barang)

            if stok_sekarang is None:
                return JsonResponse({
                    "ok": False,
                    "message": "Barang tidak ditemukan."
                }, status=404)

            if stok_sekarang < jumlah_retur:
                return JsonResponse({
                    "ok": False,
                    "message": (
                        f"Stok tidak cukup untuk retur. "
                        f"Stok saat ini {stok_sekarang}, jumlah retur {jumlah_retur}."
                    )
                }, status=400)

            cursor.execute("""
                INSERT INTO tb_retur
                    (
                        id_detail_pembelian,
                        tanggal_retur,
                        jumlah_retur,
                        alasan,
                        status_retur
                    )
                VALUES
                    (%s, %s, %s, %s, 'diproses')
            """, [
                id_detail_pembelian,
                timezone.now(),
                jumlah_retur,
                alasan,
            ])

            _rp_update_stok_barang(cursor, id_barang, -jumlah_retur)

    return JsonResponse({
        "ok": True,
        "message": "Retur berhasil disimpan. Stok barang sudah otomatis berkurang."
    })

@require_http_methods(["POST"])
def retur_update_status(request):
    gate = _rp_tolak_json_pemilik(request)
    if gate:
        return gate

    data = _rp_baca_json_request(request)

    id_retur = _rp_int(data.get("id_retur"), 0)
    status_baru = str(data.get("status_retur") or "").strip().lower()

    if id_retur <= 0:
        return JsonResponse({
            "ok": False,
            "message": "ID retur tidak valid."
        }, status=400)

    if status_baru != "batal":
        return JsonResponse({
            "ok": False,
            "message": "Aksi ini hanya untuk membatalkan retur."
        }, status=400)

    with db_transaction.atomic():
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT
                    r.id_retur,
                    COALESCE(r.jumlah_retur, 0) AS jumlah_retur,
                    r.status_retur,
                    dp.id_barang
                FROM tb_retur r
                INNER JOIN tb_detail_pembelian dp
                    ON dp.id_detail_pembelian = r.id_detail_pembelian
                WHERE r.id_retur = %s
                FOR UPDATE
            """, [id_retur])

            row = cursor.fetchone()

            if not row:
                return JsonResponse({
                    "ok": False,
                    "message": "Data retur tidak ditemukan."
                }, status=404)

            jumlah_retur = _rp_int(row[1], 0)
            status_lama = str(row[2] or "").strip().lower()
            id_barang = _rp_int(row[3], 0)

            if status_lama == "batal":
                return JsonResponse({
                    "ok": False,
                    "message": "Retur ini sudah dibatalkan."
                }, status=400)

            if status_lama == "selesai":
                return JsonResponse({
                    "ok": False,
                    "message": "Retur yang sudah selesai tidak bisa dibatalkan dari menu ini."
                }, status=400)

            if status_lama == "diproses":
                _rp_update_stok_barang(cursor, id_barang, jumlah_retur)

            cursor.execute("""
                UPDATE tb_retur
                SET status_retur = 'batal'
                WHERE id_retur = %s
            """, [id_retur])

    return JsonResponse({
        "ok": True,
        "message": "Retur berhasil dibatalkan. Stok barang sudah dikembalikan jika sebelumnya sudah berkurang."
    })

@require_http_methods(["POST"])
def retur_selesaikan(request):
    gate = _rp_tolak_json_pemilik(request)
    if gate:
        return gate

    data = _rp_baca_json_request(request)

    id_retur = _rp_int(data.get("id_retur"), 0)
    jenis_pengembalian = str(data.get("jenis_pengembalian") or "").strip().lower()
    qty_pengganti = _rp_int(data.get("qty_pengganti"), 0)

    if id_retur <= 0:
        return JsonResponse({
            "ok": False,
            "message": "ID retur tidak valid."
        }, status=400)

    if jenis_pengembalian not in ["barang_pengganti", "refund"]:
        return JsonResponse({
            "ok": False,
            "message": "Jenis penyelesaian retur tidak valid."
        }, status=400)

    with db_transaction.atomic():
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT
                    r.id_retur,
                    r.status_retur,
                    COALESCE(r.jumlah_retur, 0) AS jumlah_retur,
                    dp.id_barang,
                    COALESCE(dp.harga_beli, 0) AS harga_beli
                FROM tb_retur r
                INNER JOIN tb_detail_pembelian dp
                    ON dp.id_detail_pembelian = r.id_detail_pembelian
                WHERE r.id_retur = %s
                FOR UPDATE
            """, [id_retur])

            row = cursor.fetchone()

            if not row:
                return JsonResponse({
                    "ok": False,
                    "message": "Data retur tidak ditemukan."
                }, status=404)

            status_retur = str(row[1] or "").strip().lower()
            jumlah_retur_db = _rp_int(row[2], 0)
            id_barang = _rp_int(row[3], 0)
            harga_beli_db = _rp_decimal(row[4], 0)

            if status_retur == "selesai":
                return JsonResponse({
                    "ok": False,
                    "message": "Retur ini sudah selesai."
                }, status=400)

            if status_retur == "batal":
                return JsonResponse({
                    "ok": False,
                    "message": "Retur yang sudah batal tidak bisa diselesaikan."
                }, status=400)

            if status_retur != "diproses":
                return JsonResponse({
                    "ok": False,
                    "message": (
                        "Retur ini masih data lama. Silakan batalkan dulu lalu input retur ulang "
                        "agar stok mengikuti alur baru."
                    )
                }, status=400)

            cursor.execute("""
                SELECT COUNT(*)
                FROM tb_pengembalian
                WHERE id_retur = %s
            """, [id_retur])

            sudah_ada_pengembalian = _rp_int(cursor.fetchone()[0], 0)

            if sudah_ada_pengembalian > 0:
                return JsonResponse({
                    "ok": False,
                    "message": "Pengembalian untuk retur ini sudah pernah dicatat."
                }, status=400)

            if jenis_pengembalian == "barang_pengganti":
                if qty_pengganti <= 0:
                    return JsonResponse({
                        "ok": False,
                        "message": "Qty barang pengganti harus lebih dari 0."
                    }, status=400)

                nominal_refund = 0

            else:
                qty_pengganti = 0
                nominal_refund = jumlah_retur_db * harga_beli_db

                if nominal_refund <= 0:
                    return JsonResponse({
                        "ok": False,
                        "message": (
                            "Nominal refund tidak bisa dihitung karena harga beli belum valid. "
                            "Pastikan harga beli pada detail pembelian tidak kosong."
                        )
                    }, status=400)

            cursor.execute("""
                INSERT INTO tb_pengembalian
                    (
                        id_retur,
                        tanggal_pengembalian,
                        jenis_pengembalian,
                        qty_pengganti,
                        nominal_refund,
                        status_pengembalian
                    )
                VALUES
                    (%s, %s, %s, %s, %s, 'selesai')
            """, [
                id_retur,
                timezone.now(),
                jenis_pengembalian,
                qty_pengganti,
                nominal_refund,
            ])

            if jenis_pengembalian == "barang_pengganti":
                _rp_update_stok_barang(cursor, id_barang, qty_pengganti)

            cursor.execute("""
                UPDATE tb_retur
                SET status_retur = 'selesai'
                WHERE id_retur = %s
            """, [id_retur])

    if jenis_pengembalian == "barang_pengganti":
        pesan = "Retur selesai. Barang pengganti dicatat dan stok sudah bertambah."
    else:
        pesan = f"Retur selesai. Refund otomatis dicatat sebesar {_format_rupiah_angka(nominal_refund)}."

    return JsonResponse({
        "ok": True,
        "message": pesan
    })

@require_http_methods(["GET"])
def pengembalian_retur_options(request):
    gate = _rp_tolak_json_pemilik(request)
    if gate:
        return gate

    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT
                r.id_retur,
                r.tanggal_retur,
                r.jumlah_retur,
                r.status_retur,
                r.alasan,

                p.no_pembelian,

                b.kode_barang,
                b.nama_barang,
                b.satuan,

                COALESCE(sb.nama_supplier, sp.nama_supplier, '-') AS nama_supplier,
                COALESCE(sb.nama_perusahaan, sp.nama_perusahaan, '-') AS nama_perusahaan

            FROM tb_retur r
            INNER JOIN tb_detail_pembelian dp
                ON dp.id_detail_pembelian = r.id_detail_pembelian
            INNER JOIN tb_pembelian p
                ON p.id_pembelian = dp.id_pembelian
            INNER JOIN tb_barang b
                ON b.id_barang = dp.id_barang
            LEFT JOIN tb_supplier sb
                ON sb.id_supplier = b.id_supplier
            LEFT JOIN tb_supplier sp
                ON sp.id_supplier = p.id_supplier
            WHERE r.status_retur = 'diproses'
            AND NOT EXISTS (
                SELECT 1
                FROM tb_pengembalian pg
                WHERE pg.id_retur = r.id_retur
            )
            ORDER BY r.tanggal_retur DESC, r.id_retur DESC
        """)

        rows = _rp_dictfetchall(cursor)

    hasil = []

    for row in rows:
        row["label"] = (
            f"RETUR-{row.get('id_retur')} | "
            f"{row.get('kode_barang')} - {row.get('nama_barang')} - {row.get('satuan')} | "
            f"Jumlah retur {row.get('jumlah_retur')} | "
            f"{row.get('nama_supplier')} / {row.get('nama_perusahaan')}"
        )
        hasil.append(row)

    return JsonResponse({
        "ok": True,
        "items": hasil
    })


@require_http_methods(["GET"])
def pengembalian_json(request):
    gate = _rp_tolak_json_pemilik(request)
    if gate:
        return gate

    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT
                pg.id_pengembalian,
                pg.id_retur,
                pg.tanggal_pengembalian,
                pg.jenis_pengembalian,
                COALESCE(pg.qty_pengganti, 0) AS qty_pengganti,
                COALESCE(pg.nominal_refund, 0) AS nominal_refund,
                pg.status_pengembalian,

                COALESCE(r.jumlah_retur, 0) AS jumlah_retur,
                r.status_retur,
                r.alasan,

                COALESCE(p.no_pembelian, '-') AS no_pembelian,

                COALESCE(b.kode_barang, '-') AS kode_barang,
                COALESCE(b.nama_barang, '-') AS nama_barang,
                COALESCE(b.satuan, '-') AS satuan,

                COALESCE(sb.nama_supplier, sp.nama_supplier, '-') AS nama_supplier,
                COALESCE(sb.nama_perusahaan, sp.nama_perusahaan, '-') AS nama_perusahaan,

                CASE
                    WHEN pg.jenis_pengembalian = 'barang_pengganti' THEN 'Barang Pengganti'
                    WHEN pg.jenis_pengembalian = 'refund' THEN 'Refund'
                    ELSE COALESCE(pg.jenis_pengembalian, '-')
                END AS jenis_label,

                CASE
                    WHEN pg.status_pengembalian = 'selesai' THEN 'Selesai'
                    ELSE COALESCE(pg.status_pengembalian, '-')
                END AS status_label

            FROM tb_pengembalian pg
            LEFT JOIN tb_retur r
                ON r.id_retur = pg.id_retur
            LEFT JOIN tb_detail_pembelian dp
                ON dp.id_detail_pembelian = r.id_detail_pembelian
            LEFT JOIN tb_pembelian p
                ON p.id_pembelian = dp.id_pembelian
            LEFT JOIN tb_barang b
                ON b.id_barang = dp.id_barang
            LEFT JOIN tb_supplier sb
                ON sb.id_supplier = b.id_supplier
            LEFT JOIN tb_supplier sp
                ON sp.id_supplier = p.id_supplier
            ORDER BY pg.tanggal_pengembalian DESC, pg.id_pengembalian DESC
        """)

        rows = _rp_dictfetchall(cursor)

    return JsonResponse({
        "ok": True,
        "items": rows
    })


@require_http_methods(["POST"])
def pengembalian_create(request):
    return JsonResponse({
        "ok": False,
        "message": (
            "Input pengembalian manual sudah tidak digunakan. "
            "Silakan selesaikan retur dari halaman Retur Pembelian."
        )
    }, status=400)


@require_http_methods(["POST"])
def pengembalian_update_status(request):
    return JsonResponse({
        "ok": False,
        "message": (
            "Update status pengembalian manual sudah tidak digunakan. "
            "Status pengembalian otomatis selesai saat retur diselesaikan."
        )
    }, status=400)

# =========================================================
# CETAK STRUK THERMAL PRINTER RPP02N
# =========================================================
THERMAL_PRINTER_NAME = settings.THERMAL_PRINTER_NAME


def _thermal_text(value):
    """
    Encode teks ke bytes untuk printer thermal RONGTA.
    Menggunakan cp437 (PC437/ASCII murni) — codepage default printer thermal.
    Karakter di luar ASCII diganti dengan transliterasi aman.
    """
    text = str(value or "")
    # Transliterasi huruf beraksent agar tetap terbaca dengan benar
    replacements = {
        "á": "a", "à": "a", "ä": "a", "â": "a", "ã": "a",
        "é": "e", "è": "e", "ë": "e", "ê": "e",
        "í": "i", "ì": "i", "ï": "i", "î": "i",
        "ó": "o", "ò": "o", "ö": "o", "ô": "o", "õ": "o",
        "ú": "u", "ù": "u", "ü": "u", "û": "u",
        "ñ": "n", "ç": "c",
        "\u2019": "'", "\u2018": "'", "\u201c": '"', "\u201d": '"',
        "\u2013": "-", "\u2014": "-",
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    return text.encode("ascii", errors="replace")


def _thermal_rupiah(value):
    try:
        value = int(value or 0)
    except Exception:
        value = 0
    return f"Rp {value:,}".replace(",", ".")


def _thermal_line(width=32):
    return "-" * width + "\n"


def _thermal_left_right(left, right, width=32):
    left = str(left or "")
    right = str(right or "")

    if len(left) + len(right) >= width:
        return left[:width] + "\n" + right.rjust(width) + "\n"

    space = width - len(left) - len(right)
    return left + (" " * space) + right + "\n"


def _thermal_align_left():
    return b"\x1b\x61\x00"


def _thermal_align_center():
    return b"\x1b\x61\x01"


def _thermal_bold_on():
    return b"\x1b\x45\x01"


def _thermal_bold_off():
    return b"\x1b\x45\x00"


def _thermal_normal_size():
    return b"\x1d\x21\x00"


def _thermal_raw_print(data: bytes):
    if win32print is None:
        raise NotImplementedError("Direct server-side printing is not supported on this platform/host.")
    hprinter = win32print.OpenPrinter(THERMAL_PRINTER_NAME)

    try:
        win32print.StartDocPrinter(
            hprinter,
            1,
            ("Struk Transaksi", None, "RAW")
        )
        win32print.StartPagePrinter(hprinter)
        win32print.WritePrinter(hprinter, data)
        win32print.EndPagePrinter(hprinter)
        win32print.EndDocPrinter(hprinter)
    finally:
        win32print.ClosePrinter(hprinter)


def cetak_struk_transaksi(request, id_transaksi):
    login = _require_login(request)
    if login:
        return login

    try:
        transaksi = TbTransaksi.objects.filter(
            id_transaksi=id_transaksi
        ).first()

        if not transaksi:
            return JsonResponse({
                "status": "error",
                "message": "Data transaksi tidak ditemukan."
            }, status=404)

        details = TbTransaksiDetail.objects.filter(
            id_transaksi=id_transaksi
        ).order_by("id_detail")

        # Untuk printer 58mm.
        # Kalau hasil terlalu melebar, ubah ke 30 atau 31.
        width = 32

        data = b""

        # Reset printer
        data += b"\x1b@"           # ESC @ — inisialisasi ulang printer
        data += b"\x1bt\x00"       # ESC t 0 — pilih codepage PC437 (ASCII default)
        data += _thermal_normal_size()

        # =========================
        # HEADER / KOP RATA TENGAH
        # =========================
        data += _thermal_align_center()
        data += _thermal_bold_on()
        data += _thermal_text("Toko Pertanian Harmoni Agro\n")
        data += _thermal_bold_off()
        data += _thermal_text("Jl. Puspita Jaya, Krajan\n")
        data += _thermal_text("Kec. Jenangan, Ponorogo\n")
        data += _thermal_text("Telp. 085790727110\n")

        # Isi struk kembali rata kiri
        data += _thermal_align_left()
        data += _thermal_text(_thermal_line(width))

        # =========================
        # INFO TRANSAKSI
        # =========================
        no_jual = getattr(transaksi, "no_jual", "-") or "-"
        tanggal = _to_wib(getattr(transaksi, "tanggal_waktu", None))

        if tanggal:
            tanggal_text = tanggal.strftime("%d/%m/%Y")
            jam_text = tanggal.strftime("%H:%M")
        else:
            tanggal_text = "-"
            jam_text = "-"

        nama_kasir = getattr(transaksi, "nama_kasir", "-") or "-"

        data += _thermal_text(f"No: {no_jual}\n")
        data += _thermal_text(f"Tanggal: {tanggal_text}\n")
        data += _thermal_text(f"Jam: {jam_text}\n")
        data += _thermal_text(f"Kasir: {nama_kasir}\n")
        data += _thermal_text(_thermal_line(width))

        # =========================
        # DETAIL BARANG
        # =========================
        total_harga_hitung = 0

        for item in details:
            nama_barang = (
                getattr(item, "nama_barang_snapshot", None)
                or getattr(item, "nama_barang", None)
                or "-"
            )

            qty = _safe_decimal_to_int(getattr(item, "qty", 0))
            harga_satuan = _safe_decimal_to_int(getattr(item, "harga_satuan", 0))
            subtotal = _safe_decimal_to_int(getattr(item, "subtotal", 0))

            total_harga_hitung += subtotal

            # Nama barang di baris sendiri
            data += _thermal_text(str(nama_barang)[:width] + "\n")

            # Qty x harga di kiri, subtotal di kanan
            data += _thermal_text(
                _thermal_left_right(
                    f"{qty} x {_thermal_rupiah(harga_satuan)}",
                    _thermal_rupiah(subtotal),
                    width
                )
            )

        data += _thermal_text(_thermal_line(width))

        # =========================
        # TOTAL PEMBAYARAN
        # =========================
        total_harga_final = _safe_decimal_to_int(
            getattr(transaksi, "total_harga", total_harga_hitung)
        ) or total_harga_hitung

        nominal_bayar = _safe_decimal_to_int(
            getattr(transaksi, "nominal_bayar", 0)
        )

        kembalian = _safe_decimal_to_int(
            getattr(transaksi, "kembalian", 0)
        )

        metode_bayar = str(
            getattr(transaksi, "metode_bayar", "-") or "-"
        ).strip()

        if metode_bayar.lower() == "qris":
            metode_bayar_tampil = "QRIS"
        elif metode_bayar.lower() == "tunai":
            metode_bayar_tampil = "Tunai"
        elif metode_bayar.lower() == "marketplace":
            metode_bayar_tampil = "Marketplace"
        else:
            metode_bayar_tampil = metode_bayar.title()

        data += _thermal_bold_on()
        data += _thermal_text(
            _thermal_left_right(
                "Total",
                _thermal_rupiah(total_harga_final),
                width
            )
        )
        data += _thermal_bold_off()

        data += _thermal_text(
            _thermal_left_right(
                "Metode",
                metode_bayar_tampil,
                width
            )
        )

        data += _thermal_text(
            _thermal_left_right(
                "Bayar",
                _thermal_rupiah(nominal_bayar),
                width
            )
        )

        data += _thermal_text(
            _thermal_left_right(
                "Kembali",
                _thermal_rupiah(kembalian),
                width
            )
        )

        data += _thermal_text(_thermal_line(width))

        # =========================
        # FOOTER RATA TENGAH
        # TANGGAL TIDAK DITAMPILKAN LAGI DI BAWAH
        # =========================
        data += _thermal_align_center()
        data += _thermal_text("Barang yang sudah dibeli\n")
        data += _thermal_text("tidak dapat ditukar kembali.\n")
        data += _thermal_text("\n")
        data += _thermal_bold_on()
        data += _thermal_text("Terima kasih\n")
        data += _thermal_bold_off()

        data += _thermal_align_left()
        data += _thermal_text("\n\n\n")

        # Kembalikan data ESC/POS sebagai base64 ke browser.
        # Browser (JS) yang akan meneruskan ke local_print_server.exe di PC kasir.
        # Ini bekerja baik di lokal maupun saat Django berjalan di VPS.
        return JsonResponse({
            "status": "ok",
            "escpos_b64": base64.b64encode(data).decode("ascii"),
        })

    except Exception as e:
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)